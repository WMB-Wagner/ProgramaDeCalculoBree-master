from platform import java_ver
from openpyxl import Workbook, load_workbook
import numpy as np
import math
import pandas as pd
import os
from multiprocessing import Process
import threading
import time

seconds_inicial = time.time()

# Coletar dados do reator

arquivo = 'INPHASEINDIA72MH200A-1MODULO'  
path = "C:/Users/felipe.franchi/Documents/calculo_bree/bree_reactor_capacitors-Double-Spin-Box/Reator Fio/Saida/"
caminhodadoscusto = 'Reator Fio\Base de dados\CustoMaterial.xlsx'     
caminhodados = 'Resultados/' + arquivo + '.xlsx'

def coletardados(caminhodados, printar):
    wb_dados = load_workbook(caminhodados)
    dados = {}
    vetortitulo = []
    cont = 1
    for sheet in wb_dados:
        vetortitulo.append(sheet.title)
    reator = []
    for i in range(len(vetortitulo)):
        #print(vetortitulo[i])
        sheet_obj = wb_dados[vetortitulo[i]]
        max_row = sheet_obj.max_row
        max_column = sheet_obj.max_column
        reator.append([])
        for j in range(max_column):
            reator[i].append([])
            for k in range(max_row):
                reator[i][j].append(
                    str(sheet_obj.cell(row=k+1, column=j+1).value))
        if (printar == 1):
            print(reator[i])
            print('')
    return reator

reator = coletardados(caminhodados, 0)
numcilindros = len(reator[2][0])-1
numcamadas = reator[2][3]
numfios = reator[2][2]
fiocamada = reator[2][1]
diamisol = reator[2][10]
diamfionu = reator[2][11]
isolamento = reator[1][1][1]
alturaanel = reator[1][2][1]
fibrainterna = reator[2][4]
fibraexterna = reator[2][5]
diamint = float(reator[1][4][1])
diamext = float(reator[1][5][1])
ladomadeira = 30
pesoespira = reator[3][12]
numespiras = reator[3][3]
cilindrodacamada = reator[3][0]
numfio = len(reator[3][0])
alturadacamada = reator[3][6]
circunferenciaexterna = reator[3][8]
numbracos = int(str(reator[1][7][1])[0:1])
vetoraux = []
contadorbraco = 0
alturart = float(reator[1][26][1])
acumulador = 0
percentualroving = 0.6
percentualtecido = 0.25*percentualroving
percentualresina = 0.35*(percentualroving+percentualtecido)
percentualacelerador = 0.03*percentualresina
densidadealuminio = 2700
comprimentodafita55 = 0
distanciaentreespacadores = 150
diamespacador = float(reator[1][6][1])
cruzeta = reator[1][7][1]
ajusteembalagem = 40
numtotalespiras = 0
numtotalfios = 0
indutancianominal = float(reator[0][11][1])
frequencia = reator[0][0][1]
impedancianominal = 2*3.1415*int(frequencia)*float(indutancianominal)/1000
tensao = reator[0][1][1]
nbi = reator[0][2][1]
correntenominal = reator[0][13][1]
correntecc = reator[0][6][1]
duracaocc = reator[0][7][1]
correnteccd = reator[0][8][1]
perdasdc = reator[1][18][1]
perdasac = reator[1][19][1]
fatorq = reator[1][20][1]
potencianominal = float(impedancianominal)*((float(correntenominal)) ** 2)/1000
#if (isolamento == '130'):
#    isolamento = 'MYLAR'
#else:
#    isolamento = 'TEONEX'
if(((float(nbi)/0.45)+diamext) > (1.67*diamext)):
    de = (float(nbi)/0.45)+diamext
else:
    de = 1.67*diamext
mc1a = 0.5*diamext
mc1r = 1.1*diamext
altitude = reator[0][4][1]
temperaturaambiente = reator[0][9][1]
velocidadevento =reator[0][5][1]
pi = reator[5][2][1]
item = reator[5][6][1]
cliente = reator[5][0][1]
tipo = reator[5][4][1]
qtd = reator[5][1][1]
perimetrogabarito = 3.1415*diamint #- numbracos*float(cruzeta.split(' ')[5]) 
alturacruzeta = float(cruzeta.split(' ')[2])
distanciaentreaneis = 300
numeroaneis = math.ceil(alturart/distanciaentreaneis) + 1
perfilanel = 7.5    
numcolunascusto = 0
numlinhascusto = 0
numtotalespiras = 0
numtotalfios = 0
numtotalcamadas = 0
numespacadorescamada = reator[2][19]
numtotalespacadores = 0
superficiedoreator = 0
pesoroving = 0
pesoresina = 0
pesotecido = 0
pesoacelerador = 0

try:
    perfilhibrido = float(reator[6][3][1]) 
    larguraperfilhibrido = float(reator[6][2][1])
    calco = float(reator[6][1][1])        
    espirashibrido = reator[7][8]         
    hibrido = 1
    diamhibrido = reator[6][8][1]
except:
    hibrido = 0
    pass

try:
    observacoes = reator[8][0][1]
    obss = 1
except:
    obss = 0
    pass

densidadealuminio = 0.0027
densidademetalon = 0.015888
densidademadeira = 0.0004744
densidadefibradevidro = 0.0025
espessurametalon = 2

if(((float(nbi)/0.45)+diamext) > (1.67*diamext)):
    de = (float(nbi)/0.45)+diamext
else:
    de = 1.67*diamext
    mc1a = 0.5*diamext
    mc1r = 1.1*diamext
    altitude = reator[0][4][1]
    temperaturaambiente = reator[0][9][1]
    velocidadevento =reator[0][5][1]

# Coletar dados de custo

def coletardadosdecusto(caminhodadoscusto, printar):
    wb_dados = load_workbook(caminhodadoscusto)
    vetortitulo = []
    for sheet in wb_dados:
        vetortitulo.append(sheet.title)
    custo = []
    for i in range(len(vetortitulo)):
        # print(vetortitulo[i])
        sheet_obj = wb_dados[vetortitulo[i]]
        max_row = sheet_obj.max_row
        numlinhascusto = max_row
        max_column = sheet_obj.max_column
        numcolunascusto = max_column
        custo.append([])
        for j in range(max_column):
            custo[i].append([])
            for k in range(max_row):
                custo[i][j].append(
                    str(sheet_obj.cell(row=k+1, column=j+1).value))
        if (printar == 1):
            print(custo[i])
            print('')
    return custo, numlinhascusto, numcolunascusto

####################################################################################################################################
# ORDEM DE FABRICAÇÃO
####################################################################################################################################

def criarofrfe(wb_of,numtotalespacadores, pesofioaluminio, pesoroving, pesoresina, pesotecido, pesoacelerador):
    # CABECALHO
    wb_of['OF RFE'].cell(row=2, column=13).value = pi
    wb_of['OF RFE'].cell(row=6, column=7).value = item
    wb_of['OF RFE'].cell(row=6, column=4).value = pi
    wb_of['OF RFE'].cell(row=9, column=3).value = cliente
    wb_of['OF RFE'].cell(row=9, column=7).value = tipo
    wb_of['OF RFE'].cell(row=9, column=9).value = qtd        

    # DADOS PROJETO
    wb_of['OF RFE'].cell(row=20, column=13).value = numtotalespacadores

    # ASPECTOS CONSTRUTIVOS    
    # diametro externo
    wb_of['OF RFE'].cell(row=14, column=3).value = reator[1][5][1]
    # diametro interno
    wb_of['OF RFE'].cell(row=14, column=5).value = reator[1][4][1]
    wb_of['OF RFE'].cell(row=17, column=5).value = str(
        reator[1][8][1])[0:1]  # numero de bracos
    wb_of['OF RFE'].cell(row=17, column=13).value = len(
        reator[2][0])-1  # numero de cilindros
    wb_of['OF RFE'].cell(row=14, column=7).value = alturart + alturacruzeta*2

    # GABARITO
    wb_of['OF RFE'].cell(row=34, column=3).value = str(
        reator[1][7][1])[9:14]  # altura cruzeta
    wb_of['OF RFE'].cell(row=25, column=3).value = (perimetrogabarito)/ladomadeira   #arrumar
    wb_of['OF RFE'].cell(row=25, column=5).value = alturart+2*alturacruzeta 
    wb_of['OF RFE'].cell(row=28, column=3).value = numbracos
    wb_of['OF RFE'].cell(row=28, column=5).value = alturart + 2*alturacruzeta
    wb_of['OF RFE'].cell(row=28, column=11).value = alturart + alturacruzeta 
    wb_of['OF RFE'].cell(row=31, column=3).value = distanciaentreaneis
    wb_of['OF RFE'].cell(row=31, column=5).value = numeroaneis
    wb_of['OF RFE'].cell(row=31, column=7).value = perfilanel
    wb_of['OF RFE'].cell(row=31, column=9).value = diamint-2*23-2*ladomadeira      
    wb_of['OF RFE'].cell(row=34, column=5).value = diamint - 150
    
    # EMBALAGEM
    if alturart > 3000:
        wb_of['OF RFE'].cell(row=39, column=5).value = "Horizontal"
    else:
        wb_of['OF RFE'].cell(row=39, column=5).value = "Vertical"
    if diamext < 500 and alturart < 500: 
        wb_of['OF RFE'].cell(row=39, column=3).value = 3
    else:
        wb_of['OF RFE'].cell(row=39, column=3).value = 1
    wb_of['OF RFE'].cell(row=41, column=4).value = int(diamext + 200)
    wb_of['OF RFE'].cell(row=42, column=4).value = int(diamext + 200)
    wb_of['OF RFE'].cell(row=43, column=4).value = int(alturart + 200)

    # OBSERVACOES

    if(obss == 1):
        obs = observacoes.split('\n')
        for observacao in range(len(obs)):
            wb_of['OF RFE'].cell(row=47 + 2*observacao, column=2).value = obs[observacao]        

    # HIBRIDO
    if(hibrido == 1):
        wb_of['OF RFE'].cell(row=73, column=3).value = perfilhibrido  
        wb_of['OF RFE'].cell(row=73, column=5).value = diamhibrido 
        wb_of['OF RFE'].cell(row=73, column=7).value = calco  
        numcolunashibrido = (diamext - larguraperfilhibrido)/90
        wb_of['OF RFE'].cell(row=73, column=9).value = ((int(numcolunashibrido/numbracos)+1)*numbracos)     
        wb_of['OF RFE'].cell(row=73, column=11).value = float(espirashibrido[int(len(espirashibrido)-1)])
        wb_of['OF RFE'].cell(row=73, column=13).value = numbracos  

    # perfilhibrido = reator[6][3][1]    
    # larguraperfilhibrido = reator[6][2][1]  
    # calco = reator[6][1][1]    
    # espirashibrido = reator[7][9]

    #PESO

    cruz = cruzeta.split(' ')
    volumecruzeta = float(cruz[0])*float(cruz[2])*float(cruz[5])*(diamext/2) + (float(cruz[2])*(3.1415*(81 ** 2) - 3.1415*(22 ** 2)))
    pesocruzeta = 2*densidadealuminio*volumecruzeta/1000
    #print('Peso cruzeta: ' + str(pesocruzeta))
    wb_of['OF RFE'].cell(row=2, column=17).value = pesocruzeta

    pesofibradevidro = pesoroving + pesoresina + pesotecido + pesoacelerador
    #print('Peso fibra + resina: ' + str(pesofibradevidro))
    wb_of['OF RFE'].cell(row=3, column=17).value = pesofibradevidro

    #print('Peso fios: ' + str(pesofioaluminio))
    wb_of['OF RFE'].cell(row=4, column=17).value = pesofioaluminio

    pesoespacadores = (numtotalespacadores*(alturart - alturacruzeta)*(((diamespacador ** 2)*3.1415/4) -  (((diamespacador - 1) ** 2)*3.1415/4))*densidadefibradevidro/1000)*4
    #print('Peso espacadores: ' + str(pesoespacadores))
    wb_of['OF RFE'].cell(row=5, column=17).value = pesoespacadores

    pesoaneis = numeroaneis*perfilanel*23*(diamint - 23)*3.1415*densidadealuminio/1000
    #print('Peso aneis: ' + str(pesoaneis))
    wb_of['OF RFE'].cell(row=6, column=17).value = pesoaneis

    if(hibrido == 1):
        pesocoroa = (float(espirashibrido[int(len(espirashibrido)-1)]))*perfilhibrido*23*3.1415*diamext*densidadealuminio/1000
        #print('Peso coroa: ' + str(pesocoroa))
        wb_of['OF RFE'].cell(row=7, column=17).value = pesocoroa
    else:
        pesocoroa = 0

    pesomadeira = (densidademadeira*((perimetrogabarito)/ladomadeira)*ladomadeira*ladomadeira*(alturart)/1000)*1.1
    #print('Peso madeira: ' + str(pesomadeira))
    wb_of['OF RFE'].cell(row=8, column=17).value = pesomadeira

    pesometalon = densidademetalon*alturart*(50*30-((50-espessurametalon)*(30-espessurametalon)))*numbracos/1000
    #print('Peso metalon: ' + str(pesometalon))    
    wb_of['OF RFE'].cell(row=9, column=17).value = pesometalon

    pesoembalagem = (((diamext+10))*((diamext+10))*2 + ((diamext+10))*alturart*4)*0.02*0.25*densidademadeira
    #print('Peso embalagem: ' + str(pesoembalagem)) 
    wb_of['OF RFE'].cell(row=10, column=17).value = pesoembalagem

    pesoreator = (pesofioaluminio + pesocruzeta + pesofibradevidro + pesocoroa + pesoespacadores)
    #print('\nPeso reator: ' + str(pesoreator))
    wb_of['OF RFE'].cell(row=9, column=11).value = pesoreator
    wb_of['OF RFE'].cell(row=11, column=17).value = pesoreator

    pesoreatorgabarito = (pesoreator + pesometalon + pesomadeira + pesoaneis)*1.10
    #print('Peso reator + gabarito: ' + str(pesoreatorgabarito))
    wb_of['OF RFE'].cell(row=34, column=9).value = pesoreatorgabarito
    wb_of['OF RFE'].cell(row=12, column=17).value = pesoreatorgabarito

    pesobruto = (pesoreator + pesoembalagem)*1.10
    #print('Peso bruto: ' + str(pesobruto))
    #print('\n')
    wb_of['OF RFE'].cell(row=38, column=8).value = pesobruto
    wb_of['OF RFE'].cell(row=13, column=17).value = pesobruto

def criarbobinagem(wb_of):  
    isolamento = reator[1][1][1]
    pesofioaluminio = 0
    contadorbraco = 0

    # print('numcilindros: ' + str(numcilindros) + ', numcamadas: ' +
    #      str(numcamadas) + ', numfios: ' + str(numfios) + ', pesoespira: ' + str(pesoespira) + ', cilindrodacamada: ' + str(cilindrodacamada))

    # Criar paginas bobinagem    
    for cilindro in range(numcilindros-1):
        target = wb_of.copy_worksheet(wb_of['BOBINAGEM C1'])
    for cilindro in range(numcilindros-1):
        if (cilindro > 0):
            wb_of['BOBINAGEM C1' + ' Copy' +
                  str(cilindro)].title = 'BOBINAGEM C' + str(cilindro+2)
        else:
            wb_of['BOBINAGEM C1 Copy'].title = 'BOBINAGEM C' + str(cilindro+2)

    for cilindro in range(numcilindros):
        if(numcilindros == 1):
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12, column=11).value = 'RTR/RTR' 
        else:
            if(cilindro == 0):
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12, column=11).value = 'RTR/RR'     
            elif(cilindro == numcilindros - 1):
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12, column=11).value = 'RR/RTR' 
            else:
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12, column=11).value = 'RR/RR'  

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=7, column=3).value = cliente
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=7, column=5).value = tipo
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=7, column=7).value = qtd      

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=2, column=14).value = pi         

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=4,
                                                    column=4).value = 'Cilindro ' + str(cilindro+1)  # numero do cilindro
        contadorfiosbracoinf = []
        contadorfiosbracosup = []
        for lado in range(numbracos):
            contadorfiosbracoinf.append([0]*numbracos)
            contadorfiosbracosup.append([0]*numbracos)

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=3).value = fiocamada[cilindro+1]  # fio utilizado na camada
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=5).value = diamisol[cilindro+1]  # diametro do fio isolado na camada
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=7).value = numfios[cilindro+1]  # num de fios na camada

        if (isolamento == '130'):
            isolamento = 'Mylar'
        else:
            isolamento = 'Teonex'
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=9).value = isolamento  # isolamento do reator

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=13).value = (float(alturart) - float(alturadacamada[cilindro+1]))/2  # altura do anel de fibra

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=15,
                                                    column=11).value = diamisol[cilindro+1]  # diametro fio isolado
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=15,
                                                    column=13).value = diamfionu[cilindro+1]  # diametro fio nu

        if (cilindro == 0):
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=15,
                                                        column=3).value = float(diamint) - float(fibrainterna[cilindro+1]) - float(ladomadeira)  # diametro do gabarito
        else:
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=14,
                                                        column=3).value = 'Diâmetro interno do cilindro'  # diametro interno do cilindro
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=15,
                                                        column=3).value = float(diamint) + float(fibrainterna[cilindro+1])  # diametro interno do cilindro

        contadorauxfios = 0
        vetoraux.append([])
        for aux in range(len(cilindrodacamada)-1):
            if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                vetoraux[cilindro].append(aux)

        cont = 0
        pesofio = 0
        for fio in range(numfio-1):
            if (int(cilindrodacamada[fio+1]) == int(cilindro)):
                # areaaluminio = (
                #     (float(diamfionu[cilindro+1])/1000) ** 2)*3.1415/4
                # volumeespiraaluminio = areaaluminio * \
                #     (float(circunferenciaexterna[fio+1])/1000)
                # volumealuminio = volumeespiraaluminio*float(numespiras[fio+1])
                pesoaluminio = float(numespiras[fio+1])*float(pesoespira[fio+1])*1.15
                pesofio = pesofio + pesoaluminio
                #print('pesofio: ' + str(pesofio))
                #print(pesofio)

                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=24+11*cont,
                                                            column=6).value = float(alturadacamada[fio+1])  # altura da camada
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=26+11*cont,
                                                            column=6).value = float(circunferenciaexterna[fio+1])  # perimetro externo da camada
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=20+11*cont,
                                                            column=10).value = numespiras[fio + 1]  # numero de espiras
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=7, column=11).value = pesofio  
                cont = cont + 1
        #wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=7,
        #                                            column=11).value = pesofio  # peso de fio

        pesofioaluminio = pesofioaluminio + pesofio
        #print(pesofioaluminio)

        for braco in range(numbracos):
            wb_of['BOBINAGEM C' +
                  str(cilindro+1)].cell(row=23+braco, column=13).value = braco  # enumerar bracos
        if(cilindro < numcilindros-1):
            wb_of['BOBINAGEM C' +
                  str(cilindro+1)].cell(row=7, column=9).value = ((int(float(numespacadorescamada[cilindro+1])/numbracos)+1)*numbracos)    # numero de espacadores da camada
        
        #print('\n')

        for camada in range(int(numcamadas[cilindro+1])):                  
            for aux in range(len(cilindrodacamada)-1):
                for fio in range(int(numfios[cilindro+1])):
                    if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                        if (vetoraux[cilindro][camada] == aux):
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=8).value = fio+1  # numero do fio axial
                            fioinf = contadorbraco
                            if (fioinf >= numbracos):
                                fioinf = 0
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=9).value = fioinf  # braco inferior bobinagem
                            fiosup = fioinf + int((float(numespiras[aux+1]) - int(float(numespiras[aux+1])))/(1/numbracos))
                            if (fiosup >= numbracos):
                                fiosup = fiosup - numbracos
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=11).value = fiosup  # braco superior bobinagem

                            #print('cilindro: ' + str(cilindro + 1) + ', fioinf: ' + str(fioinf) + ', fiosup: ' + str(fiosup))   

                            contadorbraco = contadorbraco + 1
                            if (contadorbraco >= numbracos):
                                contadorbraco = 0
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=10).value = int(float(numespiras[aux + 1]))  # numero de espiras arredondado

                            # contador de fios nos bracos
                            #arrumar
                            for braco in range(numbracos):
                                try:
                                    if (fioinf == braco):                                    
                                        contadorfiosbracoinf[cilindro][braco] = contadorfiosbracoinf[cilindro][braco] + 1   
                                    if (fiosup == braco):
                                        contadorfiosbracosup[cilindro][braco] = contadorfiosbracosup[cilindro][braco] + 1                                 
                                except:
                                    pass                                                                                                                    

        for braco in range(numbracos):
            try:
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=23+braco,
                                                            column=12).value = contadorfiosbracoinf[cilindro][braco]
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=23+braco,
                                                            column=14).value = contadorfiosbracosup[cilindro][braco]                
            except:
                pass

    return pesofioaluminio

def criarcontrole( wb_of):
    vetoraux = []
    # Criar paginas controle
    for cilindro in range(numcilindros-1):
        target = wb_of.copy_worksheet(wb_of['CONTROLE C1'])
    for cilindro in range(numcilindros-1):
        if (cilindro > 0):
            wb_of['CONTROLE C1' + ' Copy' +
                  str(cilindro)].title = 'CONTROLE C' + str(cilindro+2)
        else:
            wb_of['CONTROLE C1 Copy'].title = 'CONTROLE C' + \
                str(cilindro+2)
        #print('CONTROLE C' + str(cilindro+2))

    for cilindro in range(numcilindros):
        vetoraux.append([])
        for aux in range(len(cilindrodacamada)-1):
            if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                vetoraux[cilindro].append(aux)

        wb_of['CONTROLE C' + str(cilindro+1)].cell(
            row=4, column=4).value = 'CILINDRO ' + str(cilindro+1)  # numero do cilindro

        wb_of['CONTROLE C' + str(cilindro+1)].cell(row=3, column=4).value = cliente
        wb_of['CONTROLE C' + str(cilindro+1)].cell(row=3, column=8).value = pi
        wb_of['CONTROLE C' + str(cilindro+1)].cell(row=3, column=11).value = qtd
        wb_of['CONTROLE C' + str(cilindro+1)].cell(row=3, column=13).value = tipo        

        #print('len(vetoraux[' + str(cilindro) + ']): ' + str(len(vetoraux[cilindro])))

        print('qtd: ' + str(qtd))
        for camada in range(len(vetoraux[cilindro])):
            for unid in range(int(qtd)):
                if(camada >= 1):
                    auxiliar_contador = 1
                else: 
                    auxiliar_contador = 0
                wb_of['CONTROLE C' + str(cilindro+1)].cell(row=5+3*camada+auxiliar_contador, column=7+unid).value = str(pi) + "-" + str(item) + "0" + str(unid+1) + ", Cilindro " + str(cilindro+1)
                #print('cilindro: ' + str(cilindro) + ', camada: ' + str(camada) + ', unid: ' + str(unid))
                #print('coordenada: x: ' + str(5+3*camada+auxiliar_contador) + ', y: ' + str(7+unid))
            if (camada == 1):
                wb_of['CONTROLE C' + str(cilindro+1)].cell(row=6,
                                                           column=5).value = float(diamint) - float(fibrainterna[cilindro+1])  # diametro interno da camada de fibra
            wb_of['CONTROLE C' + str(cilindro+1)].cell(row=7+3*camada,
                                                       column=5).value = float(alturadacamada[int(float(vetoraux[cilindro][camada]))+1])  # altura da camada
            wb_of['CONTROLE C' + str(cilindro+1)].cell(row=8+3*camada,
                                                       column=5).value = float(circunferenciaexterna[int(float(vetoraux[cilindro][camada]))+1])  # perimetro da camada

####################################################################################################################################
# LISTA DE MATERIAIS E ROTEIRO DE PRODUÇÃO
####################################################################################################################################

def criarlistademateriais(wb_lm, caminhodadoscusto):
    numtotalespiras = 0
    numtotalfios = 0
    numtotalcamadas = 0
    acumulador = 0
    isolamento = reator[1][1][1]
    comprimentodafita55 = 0
    numtotalespacadores = 0 
    numfiosaxiais = reator[2][2][1]
    for fio in range(len(numespiras)-1):
        numtotalespiras = numtotalespiras + float(numespiras[fio+1])   
    for cilindro in range(numcilindros):
        numtotalfios = numtotalfios + float(numfios[cilindro+1])
        numtotalcamadas = numtotalcamadas + float(numcamadas[cilindro+1])    
    constantert2 = (numtotalcamadas*alturart*diamint + numcilindros*alturart*diamint)/numtotalfios 

    for cilindro in range(numcilindros):
        if (isolamento == '130'):
            isolamento = '4M036'
        else:
            isolamento = '2M036-2T0'
        wb_lm['LISTA DE MATERIAIS'].cell(
            row=5+cilindro, column=4).value = 'Fio alumínio isolado ' + str(fiocamada[cilindro+1]) + ' AWG - ' + str(isolamento)  # fio do cilindro

        cont = 0
        pesofio = 0
        for fio in range(numfio-1):
            if (int(cilindrodacamada[fio+1]) == int(cilindro)):
                #print(str(cilindro) + '-' + str(fio))
                #areaaluminio = ((float(diamfionu[cilindro+1])/1000) ** 2)*3.1415/4
                #volumeespiraaluminio = areaaluminio * (float(circunferenciaexterna[fio+1])/1000)
                #volumealuminio = volumeespiraaluminio*float(numespiras[fio+1])
                pesoaluminio = float(numespiras[fio+1])*float(pesoespira[fio+1])*1.15
                pesofio = pesofio + pesoaluminio
                #print(pesofio)
                cont = cont + 1
        wb_lm['LISTA DE MATERIAIS'].cell(row=5+cilindro,
                                         column=5).value = pesofio  # peso de fio
        wb_lm['LISTA DE MATERIAIS'].cell(row=5+cilindro,
                                         column=6).value = '[kg]'

        vetoraux.append([])
        for aux in range(len(cilindrodacamada)-1):
            if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                vetoraux[cilindro].append(aux)

        # VOLUME DA CAMADA DE FIBRA
        for camada in range(len(vetoraux[cilindro])):
            if (camada == 0):
                raiointcilindro = ((float(circunferenciaexterna[int(float(vetoraux[cilindro][camada]))+1])/(
                    2*3.1415)) - float(diamisol[cilindro+1]) - float(fibrainterna[cilindro+1]))
                raioextcilindro = ((float(circunferenciaexterna[int(
                    float(vetoraux[cilindro][camada]))+1])/(2*3.1415)) - float(diamisol[cilindro+1]))
                volumefibra = (3.1415*((float(raioextcilindro)/1000) ** 2) -
                               3.1415*((float(raiointcilindro)/1000) ** 2))*(float(alturart)/1000)
                acumulador = acumulador + volumefibra

            if (camada == len(vetoraux[cilindro])-1):
                raioextcilindro = ((float(circunferenciaexterna[int(float(vetoraux[cilindro][len(
                    vetoraux[cilindro])-1]))+1])/(2*3.1415)) + float(fibraexterna[cilindro+1]))
                raiointcilindro = ((float(circunferenciaexterna[int(
                    float(vetoraux[cilindro][len(vetoraux[cilindro])-1]))+1])/(2*3.1415)))
                volumefibra = (3.1415*((float(raioextcilindro)/1000) ** 2) -
                               3.1415*((float(raiointcilindro)/1000) ** 2))*(float(alturart)/1000)
                acumulador = acumulador + volumefibra
                #print('acumulador ' + str(camada) + ': ' + str(acumulador))

                # 'FITA ADESIVA POLIESTER 50MM X 66 M - BRA'
                numerodevoltasfita = math.floor(
                    float(alturadacamada[vetoraux[cilindro][camada]+1])/150)
                comprimentodafita55 = (
                    comprimentodafita55 + numerodevoltasfita*raioextcilindro)

                if (cilindro < numcilindros - 1):
                    #numtotalespacadores = numtotalespacadores + math.floor((float(circunferenciaexterna[int(float(vetoraux[cilindro][len(vetoraux[cilindro])-1]))+1])/distanciaentreespacadores))
                    numtotalespacadores = numtotalespacadores + ((int(int(float(numespacadorescamada[cilindro+1]))/numbracos)+1)*numbracos) 
                    #print(numtotalespacadores)

    percentualroving = 7.5*numtotalfios/(numtotalcamadas*float(numfiosaxiais))
    percentualtecido = 0.015*percentualroving
    percentualresina = 0.33*(percentualroving+percentualtecido)
    percentualacelerador = 0.035*percentualresina

    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=3).value = 'RT23017038'   # roving
    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=4).value = 'Roving contínuo 4400TEX'
    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=5).value = percentualroving*acumulador*densidadefibradevidro*1000000
    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=6).value = '[kg]'
    pesoroving = percentualroving*acumulador*densidadefibradevidro*1000000
    #print('\npesoroving: ' + str(pesoroving))

    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=3).value = 'RT23017001'   # tecido
    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=4).value = 'Tecido WR-600/3 0,20m'
    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=5).value = percentualtecido*acumulador*densidadefibradevidro*1000000
    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=6).value = '[kg]'
    pesotecido = percentualtecido*acumulador*densidadefibradevidro*1000000
    #print('pesotecido: ' + str(pesotecido))

    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=3).value = 'RT25010001'   # resina
    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=4).value = 'Resina epoxi araldite MY750 BR'
    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=5).value = percentualresina*acumulador*densidadefibradevidro*1000000
    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=6).value = '[kg]'
    pesoresina = percentualresina*acumulador*densidadefibradevidro*1000000
    #print('pesoresina: ' + str(pesoresina))

    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=3).value = 'RT25020002'   # acelerador
    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=4).value = 'Acelerador DY 9577'
    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=5).value = percentualacelerador*acumulador*densidadefibradevidro*1000000
    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=6).value = '[kg]'
    pesoacelerador = percentualacelerador*acumulador*densidadefibradevidro*1000000
    #print('pesoacelerador: ' + str(pesoacelerador))
    #print('\n')

    numerodevoltasperfil = math.floor(float(alturart*1000000)/300)
    comprimentodeperfil = numerodevoltasperfil*2*3.1415*((float(diamint)+23)/2)
    volumedeperfil = comprimentodeperfil*7.5*23/(1000 ** 3)
    pesodeperfil = volumedeperfil*densidadealuminio

    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=3).value = 'RT22013007'
    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=4).value = 'Perfil de alumínio 7,5 x 23 mm - EE223'
    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=5).value = pesodeperfil
    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=6).value = '[kg]'

    #print('comprimentodafita55: ' + str(comprimentodafita55))

    #'FITA ADESIVA POLIESTER 19MM X 66 M - BRA'

    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=3).value = 'RT24015001'
    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=4).value = 'Fita adesiva poliester 19mm x 66m - BRA'
    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=5).value = math.ceil((comprimentodafita55/(1000*66)))
    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=6).value = '[un]'

    #'FITA ADESIVA POLIESTER 50MM X 66 M - BRA'

    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=3).value = 'RT23019005'
    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=4).value = 'Fita adesiva poliester 50mm x 66m - BRA'
    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=5).value = math.floor((comprimentodafita55*9.24/(1000*66))*10)
    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=6).value = '[un]'

    superficiedoreator = ((alturart+2*alturacruzeta)*3.1415*(diamint + diamext) + (((diamext ** 2)*3.1415)/4) - (((diamint ** 2)*3.1415)/4))/1000000
    sumatanea = superficiedoreator/6.2
    sumataneb = sumatanea*0.47/3.13
    sumaclada = superficiedoreator/15.5
    sumacladb = sumaclada
    diluente = (sumatanea + sumataneb)*0.2

    # SUMATANE HB S/B COMP. A

    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=3).value = '6200A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=4).value = 'Sumatane HB S/B comp. A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=5).value = sumatanea
    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=6).value = '[L]'

    # SUMATANE HB S/B COMP. B

    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=3).value = '6200B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=4).value = 'Sumatane HB S/B comp. B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=5).value = sumataneb
    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=6).value = '[L]'

    # SUMACLAD 940 VERDE COMP. A

    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=3).value = '6300A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=4).value = 'Sumaclad 940 verde comp. A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=5).value = sumaclada
    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=6).value = '[L]'

    # SUMACLAD 940 VERDE COMP. B

    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=3).value = '6300B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=4).value = 'Sumaclad 940 verde comp. B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=5).value = sumacladb
    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=6).value = '[L]'

    # DILUENTE P/ TINTA DE ACABAMENT

    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=3).value = '7900'
    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=4).value = 'Diluente p/ tinta de acabamento'
    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=5).value = diluente
    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=6).value = '[L]'

    # PLACA IDENT. INOX 110X50X0,6MM

    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=3).value = 'RT12011003'
    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=4).value = 'Placa ident. inox 110x50x0,6mm'
    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=5).value = 1
    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=6).value = '[un]'

    # TELA  FV COM PROTEÇÃO UV

    areadatela = 2*3.1415*((diamint ** 2)/4)/1000000
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=3).value = 'RT23017043'
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=4).value = 'Tela FV com proteção UV'
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=5).value = areadatela
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=6).value = '[m^2]'

    # FITA CADARÇO TEXIFITA A5-20L

    fitaa520 = diamext*3.1415*8/1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=3).value = 'RT23017003'
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=4).value = 'Fita cadarço TEXFITA A5-20L'
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=5).value = fitaa520*(4/3)
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=6).value = '[m]'

    # FITA CADARÇO TEXIFITA B2-35

    fitab235 = (math.ceil(alturart/35)*(diamext*3.1415/2))/1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=3).value = 'RT23017002'
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=4).value = 'Fita cadarço TEXFITA B2-35'
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=5).value = fitab235*2.5
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=6).value = '[m]'

    # REBITE ALUMINIO 3,2 X 16 MM

    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=3).value = 'RT13023001'
    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=4).value = 'Rebite de alumínio 3,2 x 16mm'
    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=5).value = 2
    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=6).value = '[un]'

    # espaçadores

    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=3).value = 'RT23017040'
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=4).value = 'Espaçadores ' + str(diamespacador) + 'mm, comp: ' + str(int(alturart))
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=5).value = numtotalespacadores #nespacadores*hespaçador*0,0175
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=6).value = '[un]'
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=7).value = alturart*0.0175

    #custo cruzeta
    volumecruzeta = (numbracos*(float(cruzeta.split(' ')[2])*float(cruzeta.split(' ')[5])*(diamext/2))) + (((80-20) ** 2)/4)*3.1415*float(cruzeta.split(' ')[2])
    #print('volumecruzeta: ' + str(volumecruzeta) + ' mm^3')
    pesocruzeta = volumecruzeta*densidadealuminio/1000
    #print('pesocruzeta: ' + str(pesocruzeta) + ' kg')
    custocruzeta = pesocruzeta*11*2
    customocruzeta = (1.1 ** numbracos)*1.1*0.5*diamext
    #print('custocruzeta: R$' + str(custocruzeta) )
    #print('customocruzeta: R$' + str(customocruzeta) )
    #print(' ')

    # cruzeta

    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=4).value = 'Cruzeta ' + str(int(diamext)) + ' mm ' + str(cruzeta)
    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=5).value = 1
    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=6).value = '[cj]'
    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=7).value = custocruzeta + customocruzeta # arrumar

    #isolador recomendado

    #mc1a

    #dimensional do pedestal
    volumepedestal = 1
    #print('volumepedestal: ' + str(volumepedestal) + ' mm^3')
    pesopedestal = volumepedestal*densidadealuminio/1000
    #print('pesopedestal: ' + str(pesopedestal) + ' kg')
    custopedestal = pesopedestal*11*2
    #print('custopedestal: R$' + str(custopedestal) )

    # pedestal

    wb_lm['LISTA DE MATERIAIS'].cell(row=24+numcilindros,
                                     column=4).value = 'Pedestal'
    wb_lm['LISTA DE MATERIAIS'].cell(row=24+numcilindros,
                                     column=5).value = numbracos
    wb_lm['LISTA DE MATERIAIS'].cell(row=24+numcilindros,
                                     column=6).value = '[un]'

    #dimensional da sapata
    volumesapata = 1
    #print('volumesapata: ' + str(volumesapata) + ' mm^3')
    pesosapata = volumesapata*densidadealuminio/1000
    #print('pesosapata: ' + str(pesosapata) + ' kg')
    custosapata = pesosapata*11*2
    #print('custosapata: R$' + str(custosapata) )

    # sapata

    wb_lm['LISTA DE MATERIAIS'].cell(row=25+numcilindros,
                                     column=4).value = 'Sapata'
    wb_lm['LISTA DE MATERIAIS'].cell(row=25+numcilindros,
                                     column=5).value = numbracos
    wb_lm['LISTA DE MATERIAIS'].cell(row=25+numcilindros,
                                     column=6).value = '[un]'

    # embalagem

    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=3).value = 'RT42011XXX'
    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=4).value = 'Embalagem L: ' + str(math.ceil(int(diamext)/10)*10 + ajusteembalagem) + ' x C: ' + str(math.ceil(int(diamext)/10)*10 + ajusteembalagem) + ' x A: ' + str(math.ceil(int(alturart)/10)*10 + ajusteembalagem)
    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=5).value = 1
    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=6).value = '[un]'

    # pintura

    wb_lm['LISTA DE MATERIAIS'].cell(row=27+numcilindros,
                                     column=3).value = 'MORT39011XXX'
    wb_lm['LISTA DE MATERIAIS'].cell(row=27+numcilindros,
                                     column=4).value = 'MO Pintura: D ' + str(int(diamext)) + ' mm x A ' + str(int(alturart + 2*alturacruzeta)) + ' mm , Superfície: ' + str(round(superficiedoreator,2)) + ' m^2'
    wb_lm['LISTA DE MATERIAIS'].cell(row=27+numcilindros,
                                     column=5).value = superficiedoreator
    wb_lm['LISTA DE MATERIAIS'].cell(row=27+numcilindros,
                                     column=6).value = '[m^2]'
    wb_lm['LISTA DE MATERIAIS'].cell(row=27+numcilindros,
                                     column=7).value = 55

    # coroa reator hibrido

    if(hibrido == 1):
        pesocoroa = (float(espirashibrido[int(len(espirashibrido)-1)]))*perfilhibrido*23*3.1415*diamext*densidadealuminio/1000
        wb_lm['LISTA DE MATERIAIS'].cell(row=29+numcilindros,
                                     column=4).value = 'Materiais da coroa de perfil'

        #perfil    
                               
        if(perfilhibrido == 7.5):
            cod_perfil = '07'
        if(perfilhibrido == 9.5):
            cod_perfil = '09'
        if(perfilhibrido == 13):
            cod_perfil = '13'
        wb_lm['LISTA DE MATERIAIS'].cell(row=30+numcilindros,
                                     column=3).value = 'RT220130' + cod_perfil                                    
        wb_lm['LISTA DE MATERIAIS'].cell(row=30+numcilindros,
                                     column=4).value = 'Perfil de alumínio ' + str(perfilhibrido) + ' mm'
        wb_lm['LISTA DE MATERIAIS'].cell(row=30+numcilindros,
                                        column=5).value = pesocoroa
        wb_lm['LISTA DE MATERIAIS'].cell(row=30+numcilindros,
                                        column=6).value = '[kg]'   

        #calços        
        if(calco == 7.5):
            cod_calco = '7'
        if(calco == 5.5):
            cod_calco = '5'
        if(calco == 4.5):
            cod_calco = '4'
        wb_lm['LISTA DE MATERIAIS'].cell(row=31+numcilindros,
                                     column=3).value = 'RT2301700' + cod_calco                                 
        wb_lm['LISTA DE MATERIAIS'].cell(row=31+numcilindros,
                                     column=4).value = 'Calço ' + str(calco) + ' mm'
        wb_lm['LISTA DE MATERIAIS'].cell(row=31+numcilindros,
                                        column=5).value = int(maxespirashibrido)*numcolunashibrido
        wb_lm['LISTA DE MATERIAIS'].cell(row=31+numcilindros,
                                        column=6).value = '[un]'   

        #calco maior

        wb_lm['LISTA DE MATERIAIS'].cell(row=32+numcilindros,
                                     column=3).value = 'RT23017011'                                
        wb_lm['LISTA DE MATERIAIS'].cell(row=32+numcilindros,
                                     column=4).value = 'Calço fibra de vidro 23 X 40 X 30 MM-B'
        wb_lm['LISTA DE MATERIAIS'].cell(row=32+numcilindros,
                                        column=5).value = numbracos - 1
        wb_lm['LISTA DE MATERIAIS'].cell(row=32+numcilindros,
                                        column=6).value = '[un]' 
        wb_lm['LISTA DE MATERIAIS'].cell(row=32+numcilindros,
                                        column=7).value = 5

        #corda 6

        wb_lm['LISTA DE MATERIAIS'].cell(row=33+numcilindros,
                                     column=3).value = 'RT23017006'                                
        wb_lm['LISTA DE MATERIAIS'].cell(row=33+numcilindros,
                                     column=4).value = 'Corda de fibra de vidro Ø6 mm'
        wb_lm['LISTA DE MATERIAIS'].cell(row=33+numcilindros,
                                        column=5).value = alturacoroa*numcolunashibrido*2*1.5*numcilindros/1000
        wb_lm['LISTA DE MATERIAIS'].cell(row=33+numcilindros,
                                        column=6).value = '[m]' 

        #corda 3

        wb_lm['LISTA DE MATERIAIS'].cell(row=34+numcilindros,
                                     column=3).value = 'RT23017033'                                
        wb_lm['LISTA DE MATERIAIS'].cell(row=34+numcilindros,
                                     column=4).value = 'Corda de fibra de vidro Ø3 mm'
        wb_lm['LISTA DE MATERIAIS'].cell(row=34+numcilindros,
                                        column=5).value = alturacoroa*numcolunashibrido*4*1.5*numcilindros/1000
        wb_lm['LISTA DE MATERIAIS'].cell(row=34+numcilindros,
                                        column=6).value = '[m]' 
    
        

    # Encontrar codigo do fio

    custo, numlinhascusto, numcolunascusto = coletardadosdecusto(
        caminhodadoscusto, 0)

    for cilindro in range(numcilindros):
        customediomaximo = 0
        for l in range(2, 41):
            for c in range(28, 29):

                if ((str(wb_lm['LISTA DE MATERIAIS'].cell(
                        row=l, column=c).value).split())[6] == '4M036'):
                    isolamento = 'MYLAR'
                else:
                    isolamento = 'TEONEX'

                descricaofio = str(str(round(float((str(wb_lm['LISTA DE MATERIAIS'].cell(
                    row=l, column=c).value).split())[3].replace(',', '.')), 1)) + ' ' + (str(wb_lm['LISTA DE MATERIAIS'].cell(
                        row=l, column=c).value).split())[4] + ' ' + (str(wb_lm['LISTA DE MATERIAIS'].cell(
                            row=l, column=c).value).split())[5] + ' ' + str(isolamento))

                if (isolamento == '130'):
                    isolamento = 'MYLAR'
                else:
                    isolamento = 'TEONEX'

                encontrarcodfio = str(str(
                    round(float(fiocamada[cilindro+1]), 1)) + ' AWG - ' + str(isolamento))

                if (encontrarcodfio == descricaofio):
                    wb_lm['LISTA DE MATERIAIS'].cell(
                        row=5+cilindro, column=3).value = str(wb_lm['LISTA DE MATERIAIS'].cell(row=l, column=27).value)  # codigo do fio do cilindro
                    wb_lm['LISTA DE MATERIAIS'].cell(
                        row=5+cilindro, column=7).value = round(int(wb_lm['LISTA DE MATERIAIS'].cell(row=l, column=29).value),2)  # custo med

    for cod in range(5+numcilindros, 26+numcilindros):
        customediomaximo = 0
        for linhacusto in range(numlinhascusto):
            #print(str(wb_lm['LISTA DE MATERIAIS'].cell(row=cod, column=3).value))
            if (str(wb_lm['LISTA DE MATERIAIS'].cell(row=cod, column=3).value) == custo[0][0][linhacusto]):
                # print(str(wb_lm['LISTA DE MATERIAIS'].cell(
                #    row=cod, column=3).value) + ' : ' + str(custo[0][1][linhacusto]))
                if (float(customediomaximo) < float(custo[0][1][linhacusto])):
                    customediomaximo = custo[0][1][linhacusto]
                    wb_lm['LISTA DE MATERIAIS'].cell(row=cod,
                                                     column=7).value = float(customediomaximo)

    if(hibrido == 1):
        for cod in range(34, 39+numcilindros):
            customediomaximo = 0
            for linhacusto in range(numlinhascusto):
                #print(str(wb_lm['LISTA DE MATERIAIS'].cell(row=cod, column=3).value))
                if (str(wb_lm['LISTA DE MATERIAIS'].cell(row=cod, column=3).value) == custo[0][0][linhacusto]):
                    # print(str(wb_lm['LISTA DE MATERIAIS'].cell(
                    #    row=cod, column=3).value) + ' : ' + str(custo[0][1][linhacusto]))
                    if (float(customediomaximo) < float(custo[0][1][linhacusto])):
                        customediomaximo = custo[0][1][linhacusto]
                        wb_lm['LISTA DE MATERIAIS'].cell(row=cod,
                                                        column=7).value = float(customediomaximo)
    
    return numtotalespacadores, superficiedoreator, pesoroving, pesoresina, pesotecido, pesoacelerador

def criarroteirodeproducao(wb_lm):
    numtotalespiras = 0
    numtotalfios = 0
    numtotalcamadas = 0
    acumulador = 0
    isolamento = reator[1][1][1]
    comprimentodafita55 = 0
    numtotalespacadores = 0 
    for fio in range(len(numespiras)-1):
        numtotalespiras = numtotalespiras + float(numespiras[fio+1])   
    for cilindro in range(numcilindros):
        numtotalfios = numtotalfios + float(numfios[cilindro+1])
        numtotalcamadas = numtotalcamadas + float(numcamadas[cilindro+1])
    if(int(alturart) > 700 or int(diamext) > 700):
        dificuldademovimentacao = 1
    else:
        dificuldademovimentacao = 0
    
    """wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=6, column=7).value = float(numaneis)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=6, column=8).value = float(diamint)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=7, column=7).value = float(numaneis)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=8, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=9, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=10, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=10, column=8).value = float(numaneis)   

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=7).value = float(numtotalfios)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=8).value = float(diamext)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=9).value = float(numtotalespiras)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=10).value = float(numtotalespacadores)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=13, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=13, column=8).value = float(numaneis)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=14, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=15, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=15, column=8).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=16, column=7).value = float(superficiedoreator)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=17, column=7).value = float(superficiedoreator)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=18, column=7).value = float(superficiedoreator)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=18, column=8).value = float(numbracos)   

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=19, column=8).value = float(dificuldademovimentacao)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=20, column=7).value = float(dificuldademovimentacao)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=21, column=7).value = float(alturart)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=21, column=8).value = float(diamext)"""

    constantert = (numtotalcamadas*alturart*diamint + numcilindros*alturart*diamint)
    constantert2 = (numtotalcamadas*alturart*diamint + numcilindros*alturart*diamint)/int(numtotalfios)
    #print("Constante RT: " + str(constantert))
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=4, column=6).value = float(constantert)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=4, column=7).value = float(constantert2)

def criarcodigostotvs(wb_lm):
    numtotalespiras = 0
    numtotalfios = 0
    numtotalcamadas = 0
    acumulador = 0
    isolamento = reator[1][1][1]
    comprimentodafita55 = 0
    numtotalespacadores = 0 
    if(int(item) >= 10):
        auxitem = item
    else:
        auxitem = '0' + str(item)

    ultimocodutilizado = 'XXX'

    wb_lm['CÓDIGOS TOTVS'].cell(row=2, column=2).value = str(pi) + auxitem + 'RT'
    wb_lm['CÓDIGOS TOTVS'].cell(row=2, column=3).value = 'Reator - ' + tipo  

    wb_lm['CÓDIGOS TOTVS'].cell(row=3, column=2).value = 'RTPD' + str(pi) + auxitem
    wb_lm['CÓDIGOS TOTVS'].cell(row=3, column=3).value = 'Placa de dados reator 110 x 50 x 1.2 mm'

    wb_lm['CÓDIGOS TOTVS'].cell(row=5, column=2).value = 'RT39010' + ultimocodutilizado
    wb_lm['CÓDIGOS TOTVS'].cell(row=5, column=3).value = 'Reator pai - ' + tipo
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=6, column=2).value = 'RT42011' + ultimocodutilizado
    wb_lm['CÓDIGOS TOTVS'].cell(row=6, column=3).value = 'Embalagem L: ' + str(math.ceil(int(diamext)/10)*10 + ajusteembalagem) + ' x C: ' + str(math.ceil(int(diamext)/10)*10 + ajusteembalagem) + ' x A: ' + str(math.ceil(int(alturart)/10)*10 + ajusteembalagem)
    

    wb_lm['CÓDIGOS TOTVS'].cell(row=8, column=2).value = 'RTSA0' + ultimocodutilizado
    wb_lm['CÓDIGOS TOTVS'].cell(row=8, column=3).value = 'Reator - semi acabado ' + tipo
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=9, column=2).value = '6200A' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=9, column=3).value = 'Sumatane HB S/B comp. A' 
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=10, column=2).value = '6200B' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=10, column=3).value = 'Sumatane HB S/B comp. B'
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=11, column=2).value = '6300A'
    wb_lm['CÓDIGOS TOTVS'].cell(row=11, column=3).value = 'Sumaclad 940 verde comp. A'
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=12, column=2).value = '6300B' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=12, column=3).value = 'Sumaclad 940 verde comp. B'
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=13, column=2).value = '7900' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=13, column=3).value = 'Diluente p/ tinta de acabamento'
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=14, column=2).value = 'MORT39010' + ultimocodutilizado
    wb_lm['CÓDIGOS TOTVS'].cell(row=14, column=3).value = 'MO Pintura: D ' + str(int(diamext)) + ' mm x A ' + str(int(alturart + 2*alturacruzeta)) + ' mm , Superfície: ' + str(round(superficiedoreator,2)) + ' m^2'
  
    wb_lm['CÓDIGOS TOTVS'].cell(row=15, column=2).value = '-' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=15, column=3).value = 'Pedestal'


    wb_lm['CÓDIGOS TOTVS'].cell(row=17, column=2).value = 'RTSR0' + ultimocodutilizado
    wb_lm['CÓDIGOS TOTVS'].cell(row=17, column=3).value = 'Reator - seção reativa ' + tipo
    #fios    
    wb_lm['CÓDIGOS TOTVS'].cell(row=18, column=2).value = 'RT22013007'
    wb_lm['CÓDIGOS TOTVS'].cell(row=18, column=3).value = 'Perfil de alumínio 7,5 x 23 mm - EE223' 
        
    wb_lm['CÓDIGOS TOTVS'].cell(row=19, column=2).value = 'RT23017001'
    wb_lm['CÓDIGOS TOTVS'].cell(row=19, column=3).value = 'Tecido WR-600/3 0,20m' 
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=20, column=2).value = 'RT23017002' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=20, column=3).value = 'Fita cadarço TEXFITA B2-35' 
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=21, column=2).value = 'RT23017038' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=21, column=3).value = 'Roving contínuo 4400TEX' 
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=22, column=2).value = 'RT23017040'
    wb_lm['CÓDIGOS TOTVS'].cell(row=22, column=3).value = 'Espaçadores 19.05mm, comp: ' + str(alturart + alturacruzeta)
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=23, column=2).value = 'RT24015001'
    wb_lm['CÓDIGOS TOTVS'].cell(row=23, column=3).value = 'Fita adesiva poliester 19mm x 66m - BRA'
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=24, column=2).value = 'RT23019005' + ultimocodutilizado
    wb_lm['CÓDIGOS TOTVS'].cell(row=24, column=3).value = 'Fita adesiva poliester 50mm x 66m - BRA'
    
    wb_lm['CÓDIGOS TOTVS'].cell(row=25, column=2).value = '-' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=25, column=3).value = 'Sapata'

    wb_lm['CÓDIGOS TOTVS'].cell(row=26, column=2).value = 'RT25010001' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=26, column=3).value = 'Resina epoxi araldite MY750 BR'

    wb_lm['CÓDIGOS TOTVS'].cell(row=27, column=2).value = 'RT25020002' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=27, column=3).value = 'Acelerador DY 9577'

    wb_lm['CÓDIGOS TOTVS'].cell(row=27, column=2).value = '-' 
    wb_lm['CÓDIGOS TOTVS'].cell(row=27, column=3).value = 'Cruzeta' 

    if(hibrido == 1):
        wb_lm['CÓDIGOS TOTVS'].cell(row=29, column=2).value = 'RTCP0' + ultimocodutilizado
        wb_lm['CÓDIGOS TOTVS'].cell(row=29, column=3).value = 'Reator - corpo ' + tipo 
        
        wb_lm['CÓDIGOS TOTVS'].cell(row=30, column=2).value = 'RT23017006' 
        wb_lm['CÓDIGOS TOTVS'].cell(row=30, column=3).value = 'Corda de fibra de vidro Ø6 mm' 
        
        wb_lm['CÓDIGOS TOTVS'].cell(row=31, column=2).value = 'RT23017033' 
        wb_lm['CÓDIGOS TOTVS'].cell(row=31, column=3).value = 'Corda de fibra de vidro Ø3 mm' 
        #calcoespecial
        wb_lm['CÓDIGOS TOTVS'].cell(row=32, column=2).value = 'RT23017011' 
        wb_lm['CÓDIGOS TOTVS'].cell(row=32, column=3).value = 'Calço fibra de vidro 23 X 40 X 30 MM-B' 
        #calco        
        wb_lm['CÓDIGOS TOTVS'].cell(row=33, column=2).value = '-' 
        wb_lm['CÓDIGOS TOTVS'].cell(row=33, column=3).value = 'Calço ' + str(calco) + ' mm'


        wb_lm['CÓDIGOS TOTVS'].cell(row=35, column=2).value = 'RTPC0' + ultimocodutilizado
        wb_lm['CÓDIGOS TOTVS'].cell(row=35, column=3).value = 'Reator - perfil conformado ' + tipo 

        #perfil
        wb_lm['CÓDIGOS TOTVS'].cell(row=36, column=2).value = '-' 
        wb_lm['CÓDIGOS TOTVS'].cell(row=36, column=3).value = 'Calço ' + str(calco) + ' mm'

####################################################################################################################################
# PROPOSTA
####################################################################################################################################

def criarproposta(wb_pp):
    # PROPOSTA PT-BR

    wb_pp['PROPOSTA (PT-BR)'].cell(row=5, column=3).value = 'Reator(es), tipo '

    wb_pp['PROPOSTA (PT-BR)'].cell(row=9, column=4).value = indutancianominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=10, column=4).value = impedancianominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=11, column=4).value = tensao
    wb_pp['PROPOSTA (PT-BR)'].cell(row=12, column=4).value = nbi
    wb_pp['PROPOSTA (PT-BR)'].cell(row=13, column=4).value = frequencia
    wb_pp['PROPOSTA (PT-BR)'].cell(row=14, column=4).value = '' #frequencia de sintonia
    wb_pp['PROPOSTA (PT-BR)'].cell(row=15, column=4).value = correntenominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=16, column=4).value = str(correntecc) + ' / ' + str(duracaocc)
    wb_pp['PROPOSTA (PT-BR)'].cell(row=17, column=4).value = correnteccd
    wb_pp['PROPOSTA (PT-BR)'].cell(row=18, column=4).value = float(perdasdc) + float(perdasac)
    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=4).value = fatorq
    wb_pp['PROPOSTA (PT-BR)'].cell(row=20, column=5).value = '' # fator q a frequencia de sintonia
    wb_pp['PROPOSTA (PT-BR)'].cell(row=21, column=4).value = potencianominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=22, column=4).value = 'A.N.'
    wb_pp['PROPOSTA (PT-BR)'].cell(row=23, column=4).value = isolamento

    wb_pp['PROPOSTA (PT-BR)'].cell(row=28, column=4).value = alturart
    wb_pp['PROPOSTA (PT-BR)'].cell(row=29, column=4).value = diamext
    wb_pp['PROPOSTA (PT-BR)'].cell(row=30, column=4).value = '' #diametro fundacao
    wb_pp['PROPOSTA (PT-BR)'].cell(row=31, column=4).value = '' #peso por modulo
    wb_pp['PROPOSTA (PT-BR)'].cell(row=32, column=4).value = '' #peso total

    wb_pp['PROPOSTA (PT-BR)'].cell(row=9, column=8).value = 'ABNT NBR 5356-06'

    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=10).value = de
    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=21).value = mc1a
    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=23).value = mc1r

    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = altitude
    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = temperaturaambiente
    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = velocidadevento

    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = str(int(diamext)) + ' x ' + str(int(diamext)) + ' x ' + str(int(alturart)) 

    # PROPOSTA EN-US

    wb_pp['PROPOSTA (EN-US)'].cell(row=5, column=3).value = 'Air Coil(s), type '

    wb_pp['PROPOSTA (EN-US)'].cell(row=9, column=4).value = indutancianominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=10, column=4).value = impedancianominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=11, column=4).value = tensao
    wb_pp['PROPOSTA (EN-US)'].cell(row=12, column=4).value = nbi
    wb_pp['PROPOSTA (EN-US)'].cell(row=13, column=4).value = frequencia
    wb_pp['PROPOSTA (EN-US)'].cell(row=14, column=4).value = '' #frequencia de sintonia
    wb_pp['PROPOSTA (EN-US)'].cell(row=15, column=4).value = correntenominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=16, column=4).value = str(correntecc) + ' / ' + str(duracaocc)
    wb_pp['PROPOSTA (EN-US)'].cell(row=17, column=4).value = correnteccd
    wb_pp['PROPOSTA (EN-US)'].cell(row=18, column=4).value = float(perdasdc) + float(perdasac)
    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=4).value = fatorq
    wb_pp['PROPOSTA (EN-US)'].cell(row=20, column=5).value = '' # fator q a frequencia de sintonia
    wb_pp['PROPOSTA (EN-US)'].cell(row=21, column=4).value = potencianominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=22, column=4).value = 'A.N.'
    wb_pp['PROPOSTA (EN-US)'].cell(row=23, column=4).value = isolamento

    wb_pp['PROPOSTA (EN-US)'].cell(row=28, column=4).value = alturart
    wb_pp['PROPOSTA (EN-US)'].cell(row=29, column=4).value = diamext
    wb_pp['PROPOSTA (EN-US)'].cell(row=30, column=4).value = '' #diametro fundacao
    wb_pp['PROPOSTA (EN-US)'].cell(row=31, column=4).value = '' #peso por modulo
    wb_pp['PROPOSTA (EN-US)'].cell(row=32, column=4).value = '' #peso total

    wb_pp['PROPOSTA (EN-US)'].cell(row=9, column=8).value = 'ABNT NBR 5356-06'

    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=10).value = de
    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=21).value = mc1a
    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=23).value = mc1r

    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = altitude
    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = temperaturaambiente
    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = velocidadevento

    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = str(int(diamext)) + ' x ' + str(int(diamext)) + ' x ' + str(int(alturart)) 

    # PROPOSTA ES-ES

    wb_pp['PROPOSTA (ES-ES)'].cell(row=5, column=3).value = 'Reator(es), tipo '

    wb_pp['PROPOSTA (ES-ES)'].cell(row=9, column=4).value = indutancianominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=10, column=4).value = impedancianominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=11, column=4).value = tensao
    wb_pp['PROPOSTA (ES-ES)'].cell(row=12, column=4).value = nbi
    wb_pp['PROPOSTA (ES-ES)'].cell(row=13, column=4).value = frequencia
    wb_pp['PROPOSTA (ES-ES)'].cell(row=14, column=4).value = '' #frequencia de sintonia
    wb_pp['PROPOSTA (ES-ES)'].cell(row=15, column=4).value = correntenominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=16, column=4).value = str(correntecc) + ' / ' + str(duracaocc)
    wb_pp['PROPOSTA (ES-ES)'].cell(row=17, column=4).value = correnteccd
    wb_pp['PROPOSTA (ES-ES)'].cell(row=18, column=4).value = float(perdasdc) + float(perdasac)
    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=4).value = fatorq
    wb_pp['PROPOSTA (ES-ES)'].cell(row=20, column=5).value = '' # fator q a frequencia de sintonia
    wb_pp['PROPOSTA (ES-ES)'].cell(row=21, column=4).value = potencianominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=22, column=4).value = 'A.N.'
    wb_pp['PROPOSTA (ES-ES)'].cell(row=23, column=4).value = isolamento

    wb_pp['PROPOSTA (ES-ES)'].cell(row=28, column=4).value = alturart
    wb_pp['PROPOSTA (ES-ES)'].cell(row=29, column=4).value = diamext
    wb_pp['PROPOSTA (ES-ES)'].cell(row=30, column=4).value = '' #diametro fundacao
    wb_pp['PROPOSTA (ES-ES)'].cell(row=31, column=4).value = '' #peso por modulo
    wb_pp['PROPOSTA (ES-ES)'].cell(row=32, column=4).value = '' #peso total

    wb_pp['PROPOSTA (ES-ES)'].cell(row=9, column=8).value = 'ABNT NBR 5356-06'

    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=10).value = de
    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=21).value = mc1a
    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=23).value = mc1r

    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = altitude
    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = temperaturaambiente
    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = velocidadevento

    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = str(int(diamext)) + ' x ' + str(int(diamext)) + ' x ' + str(int(alturart)) 

####################################################################################################################################
# PLANO DE INSPEÇÃO E CONTROLE DE QUALIDADE
####################################################################################################################################

def criarpicq(wb_picq):
    #PICQ PT-BR

    wb_picq['PICQ PT-BR'].cell(row=5, column=3).value = ''
    wb_picq['PICQ PT-BR'].cell(row=6, column=3).value = ''
    wb_picq['PICQ PT-BR'].cell(row=7, column=3).value = ''
    wb_picq['PICQ PT-BR'].cell(row=5, column=6).value = ''
    wb_picq['PICQ PT-BR'].cell(row=6, column=6).value = ''
    wb_picq['PICQ PT-BR'].cell(row=7, column=6).value = ''
    wb_picq['PICQ PT-BR'].cell(row=5, column=9).value = ''
    wb_picq['PICQ PT-BR'].cell(row=6, column=9).value = ''

    wb_picq['PICQ PT-BR'].cell(row=11, column=4).value = indutancianominal
    wb_picq['PICQ PT-BR'].cell(row=12, column=4).value = correntenominal
    wb_picq['PICQ PT-BR'].cell(row=13, column=4).value = tensao
    wb_picq['PICQ PT-BR'].cell(row=14, column=4).value = frequencia
    wb_picq['PICQ PT-BR'].cell(row=15, column=4).value = ''
    wb_picq['PICQ PT-BR'].cell(row=16, column=4).value = correntecc
    wb_picq['PICQ PT-BR'].cell(row=11, column=8).value = impedancianominal
    wb_picq['PICQ PT-BR'].cell(row=12, column=8).value = round((float(perdasac) + float(perdasdc))/1000,2)
    wb_picq['PICQ PT-BR'].cell(row=13, column=8).value = nbi
    wb_picq['PICQ PT-BR'].cell(row=16, column=8).value = correnteccd

    wb_picq['PICQ PT-BR'].cell(row=24, column=7).value = ''

    #PICQ EN-US

    wb_picq['PICQ EN-US'].cell(row=5, column=3).value = ''
    wb_picq['PICQ EN-US'].cell(row=6, column=3).value = ''
    wb_picq['PICQ EN-US'].cell(row=7, column=3).value = ''
    wb_picq['PICQ EN-US'].cell(row=5, column=6).value = ''
    wb_picq['PICQ EN-US'].cell(row=6, column=6).value = ''
    wb_picq['PICQ EN-US'].cell(row=7, column=6).value = ''
    wb_picq['PICQ EN-US'].cell(row=5, column=9).value = ''
    wb_picq['PICQ EN-US'].cell(row=6, column=9).value = ''

    wb_picq['PICQ EN-US'].cell(row=11, column=4).value = indutancianominal
    wb_picq['PICQ EN-US'].cell(row=12, column=4).value = correntenominal
    wb_picq['PICQ EN-US'].cell(row=13, column=4).value = tensao
    wb_picq['PICQ EN-US'].cell(row=14, column=4).value = frequencia
    wb_picq['PICQ EN-US'].cell(row=15, column=4).value = ''
    wb_picq['PICQ EN-US'].cell(row=16, column=4).value = correntecc
    wb_picq['PICQ EN-US'].cell(row=11, column=8).value = impedancianominal
    wb_picq['PICQ EN-US'].cell(row=12, column=8).value = round((float(perdasac) + float(perdasdc))/1000,2)
    wb_picq['PICQ EN-US'].cell(row=13, column=8).value = nbi
    wb_picq['PICQ EN-US'].cell(row=16, column=8).value = correnteccd

    wb_picq['PICQ EN-US'].cell(row=24, column=7).value = ''

    #PICQ ES-ES

    wb_picq['PICQ ES-ES'].cell(row=5, column=3).value = ''
    wb_picq['PICQ ES-ES'].cell(row=6, column=3).value = ''
    wb_picq['PICQ ES-ES'].cell(row=7, column=3).value = ''
    wb_picq['PICQ ES-ES'].cell(row=5, column=6).value = ''
    wb_picq['PICQ ES-ES'].cell(row=6, column=6).value = ''
    wb_picq['PICQ ES-ES'].cell(row=7, column=6).value = ''
    wb_picq['PICQ ES-ES'].cell(row=5, column=9).value = ''
    wb_picq['PICQ ES-ES'].cell(row=6, column=9).value = ''

    wb_picq['PICQ ES-ES'].cell(row=11, column=4).value = indutancianominal
    wb_picq['PICQ ES-ES'].cell(row=12, column=4).value = correntenominal
    wb_picq['PICQ ES-ES'].cell(row=13, column=4).value = tensao
    wb_picq['PICQ ES-ES'].cell(row=14, column=4).value = frequencia
    wb_picq['PICQ ES-ES'].cell(row=15, column=4).value = ''
    wb_picq['PICQ ES-ES'].cell(row=16, column=4).value = correntecc
    wb_picq['PICQ ES-ES'].cell(row=11, column=8).value = impedancianominal
    wb_picq['PICQ ES-ES'].cell(row=12, column=8).value = round((float(perdasac) + float(perdasdc))/1000,2)
    wb_picq['PICQ ES-ES'].cell(row=13, column=8).value = nbi
    wb_picq['PICQ ES-ES'].cell(row=16, column=8).value = correnteccd

    wb_picq['PICQ ES-ES'].cell(row=24, column=7).value = ''

####################################################################################################################################
# FUNÇÕES GLOBAIS
####################################################################################################################################

def criarof(numtotalespacadores,path,arquivo, pesoroving, pesoresina, pesotecido, pesoacelerador):
    print('Criando Ordem de Fabricação, Roteiro de Bobinagem e Controle de Bobinagem...\n')    
    caminhoof = 'Reator Fio\Modelos\Ordem de fabricação.xlsm'
    wb_of = load_workbook(caminhoof)    
    pesofioaluminio = criarbobinagem(wb_of)
    criarofrfe(wb_of,numtotalespacadores, pesofioaluminio, pesoroving, pesoresina, pesotecido, pesoacelerador)
    criarcontrole( wb_of)
    wb_of.save(path + "OF " + arquivo + " (Saida).xlsx")

def criarlm(path,arquivo):
    print('\nCriando Lista de Materiais e Roteiro de Produção...\n')    
    caminholm = 'Reator Fio\Modelos\Lista de materiais.xlsm'
    wb_lm = load_workbook(caminholm)
    numtotalespacadores,superficiedoreator, pesoroving, pesoresina, pesotecido, pesoacelerador = criarlistademateriais(wb_lm, caminhodadoscusto)
    hh = criarroteirodeproducao(wb_lm)
    criarcodigostotvs(wb_lm)        
    wb_lm.save(path + "LM " + arquivo + " (Saida).xlsx")
    #print('pesoroving: ' + str(pesoroving))
    #print('pesoresina: ' + str(pesoresina))
    #print('pesotecido: ' + str(pesotecido))
    #print('pesoacelerador: ' + str(pesoacelerador))
    return numtotalespacadores,superficiedoreator, pesoroving, pesoresina, pesotecido, pesoacelerador

def criarpp(path, arquivo):
    print('Criando Desenho e Documento de Proposta...\n')    
    caminhopp = 'Reator Fio\Modelos\Proposta técnica.xlsm'
    wb_pp = load_workbook(caminhopp)
    criarproposta(wb_pp)
    wb_pp.save(path + "PT " + arquivo + " (Saida).xlsx")

def criarpit(path, arquivo):
    print('Criando Plano de inspeção e controle de qualidade...\n')    
    caminhopicq = 'Reator Fio\Modelos\Plano de inspeção e controle de qualidade.xlsm'
    wb_picq = load_workbook(caminhopicq)
    criarpicq(wb_picq)    
    wb_picq.save(path + "PICQ " + arquivo + " (Saida).xlsx")

def criarcaminho(path,arquivo,caminhodados):
    reator = coletardados(caminhodados, 0)
    item = reator[5][6][1]
    if(arquivo != ''):
        #print("Nome não nulo")
        if(os.path.exists(path + "/Item " + item)):
            print("Diretório já existe.")
            prosseguir = 0
            try:
                os.remove(path)
            except:
                print("O usuário não tem permissão para sobrepor a pasta.")
                pass
        else:
            os.mkdir(path)
            os.mkdir(path + "/Item " + item)
            os.mkdir(path + "/Item " + item + "/Documentos/")
            os.mkdir(path + "/Item " + item + "/Desenhos/")
            os.mkdir(path + "/Item " + item + "/Relatórios/")
            prosseguir = 1        
    else:
        prosseguir = 0
    return prosseguir       

#import threading
#lock = threading.Lock()
#condition_var = threading.Condition(lock=lock)   
#condition_var.acquire(blocking=True)      
#condition_var.wait(timeout=10)    

def criardocs():    
    reator = coletardados(caminhodados, 0)
    item = reator[5][6][1]
    pi = reator[5][2][1]
    cliente = reator[5][0][1]      
    caminhodadoscusto = 'Reator Fio\Base de dados\CustoMaterial.xlsx'      
    path = "C:/Users/felipe.franchi/Documents/calculo_bree/bree_reactor_capacitors-Double-Spin-Box/Reator Fio/Saida/"
    prosseguir = criarcaminho(path + str(pi) + '-' + str(cliente),arquivo,caminhodados)  
    if(prosseguir == 1):  
        #numtotalespacadores,superficiedoreator, pesoroving, pesoresina, pesotecido, pesoacelerador = criarlm(caminhodados, caminhodadoscusto, path + str(pi) + '-' + str(cliente) + "/Item " + item + '/Documentos/', arquivo)        

        #Tentativa procedural: tempo: 9,41      
        #criarof(caminhodados,numtotalespacadores,superficiedoreator, path + str(pi) + '-' + str(cliente) +  "/Item " + item + '/Documentos/', arquivo, pesoroving, pesoresina, pesotecido, pesoacelerador)       
        #criarpp(caminhodados, path + str(pi) + '-' + str(cliente) + "/Item " + item + '/Documentos/', arquivo)        
        #criarpit(caminhodados, path + str(pi) + '-' + str(cliente) +  "/Item " + item + '/Documentos/', arquivo)

        #Tentativa multiprocessing tempo: 23,76
        #pcriarof = Process(target=criarof, args=(caminhodados,numtotalespacadores,superficiedoreator, path + str(pi) + '-' + str(cliente) +  "/Item " + item + '/Documentos/', arquivo, pesoroving, pesoresina, pesotecido, pesoacelerador))
        #pcriarpp = Process(target=criarpp, args=(caminhodados, path + str(pi) + '-' + str(cliente) + "/Item " + item + '/Documentos/', arquivo))
        #pcriarpit = Process(target=criarpit, args=(caminhodados, path + str(pi) + '-' + str(cliente) +  "/Item " + item + '/Documentos/', arquivo))
        #pcriarof.start()
        #pcriarpp.start()
        #pcriarpit.start()
        #pcriarof.join()
        #pcriarpp.join()
        #pcriarpit.join()

        #Tentativa multithreading tempo: 8.75360369682312         
        #pcriarlm = threading.Thread(target=criarlm, args=(path + str(pi) + '-' + str(cliente) + "/Item " + item + '/Documentos/', arquivo))     
        numtotalespacadores,superficiedoreator, pesoroving, pesoresina, pesotecido, pesoacelerador = criarlm(path + str(pi) + '-' + str(cliente) + "/Item " + item + '/Documentos/', arquivo)
        pcriarof = threading.Thread(target=criarof, args=(numtotalespacadores, path + str(pi) + '-' + str(cliente) +  "/Item " + item + '/Documentos/', arquivo, pesoroving, pesoresina, pesotecido, pesoacelerador))
        pcriarpp = threading.Thread(target=criarpp, args=(path + str(pi) + '-' + str(cliente) + "/Item " + item + '/Documentos/', arquivo))
        pcriarpit = threading.Thread(target=criarpit, args=( path + str(pi) + '-' + str(cliente) +  "/Item " + item + '/Documentos/', arquivo))
                
        #pcriarlm.start()
        pcriarof.start()
        pcriarpp.start()
        pcriarpit.start()  
        
        #pcriarlm.join()          
        pcriarof.join()
        pcriarpp.join()
        pcriarpit.join()

        print('Finalizado o Processo de Emissão dos Documentos.\n')

def criarpd():
    caminhodados = 'Resultados\GENERICO1010MH25A.xlsx'
    reator = pd.read_excel(caminhodados,)
    print(reator)
    print('\n')
    print(reator['Frequência (Hz)'])

####################################################################################################################################
# CALCULO DE ESFORCOS
####################################################################################################################################

def calculoesforcos(caminholm):
    forcahoop = float(reator[2][29][1])
    forcacompressao = float(reator[2][28][1])

    if(forcahoop > forcacompressao):
        forcacc = forcahoop
    else:
        forcacc = forcacompressao

    nbi = 950
    print(nbi)

    #distancia entre eixos
    dee1 = (nbi/0.45+diamext)*0.03937007874 #polegadas
    dee2 = (diamext*1.67)*0.03937007874 #polegadas
    if(dee1 > dee2):
        dee = dee1
    else:
        dee = dee2
    
    raiomedio = ((diamint + ((diamext-diamint)/2))/2)*0.03937007874 #polegadas
    espessurareator = ((diamext-diamint)/2)*0.03937007874 #polegadas

    print('\nForças Horizontais\n')

    #forcas horizontais
    forcalinepullhorizontal = 500
    print('forca linepull horizontal: ' + str(forcalinepullhorizontal))
    constanteventoreator = 1*0.613*(float(velocidadevento)/3.6)**2
    forcaventohorizontal = constanteventoreator*alturart*diamext/1000000
    print('forca vento horizontal: ' + str(forcaventohorizontal))  
    constantesfcc = [0.166666667,0.006944444,3.75,0.583333333,0.054166667,0.001041667,13.671875,1.133333333,0.281666667,0.014642857,0.000168155,50.24414063,1.845238095,0.859920635,0.115008503,0.003771613,2.90828E-05]
    forca1ampereespira = (((3*(3.1415 ** 2))*(raiomedio ** 4))/((4.45*(10 ** 7))*(dee ** 4)))*((1+constantesfcc[0]*(espessurareator ** 2)/(raiomedio ** 2)+constantesfcc[1]*(espessurareator ** 4)/(raiomedio ** 4))+(constantesfcc[2]*((raiomedio ** 2)/(dee ** 2))*(1+constantesfcc[3]*(espessurareator ** 2)/(raiomedio ** 2)+constantesfcc[4]*(espessurareator ** 4)/(raiomedio ** 4)+constantesfcc[5]*(espessurareator ** 6)/(raiomedio ** 6)))+(constantesfcc[6]*((raiomedio ** 4)/(dee ** 4))*(1+constantesfcc[7]*(espessurareator ** 2)/(raiomedio ** 2)+constantesfcc[8]*(espessurareator ** 4)/(raiomedio ** 4)+constantesfcc[9]*(espessurareator ** 6)/(raiomedio ** 6)+constantesfcc[10]*(espessurareator ** 8)/(raiomedio ** 8)))+(constantesfcc[11]*((raiomedio ** 6)/(dee ** 6))*(1+constantesfcc[12]*(espessurareator ** 2)/(raiomedio ** 2)+constantesfcc[13]*(espessurareator ** 4)/(raiomedio ** 4)+constantesfcc[14]*(espessurareator ** 6)/(raiomedio ** 6)+constantesfcc[15]*(espessurareator ** 8)/(raiomedio ** 8)+constantesfcc[16]*(espessurareator ** 10)/(raiomedio ** 10))))
    forcacurtocircuitototal = (forca1ampereespira*(((float(correntecc)*1000) ** 2)*((float(numespiras[1])*float(numfios[1])) ** 2)))*0.5
    forcacurtocircuitohorizontal = forcacurtocircuitototal*4.448
    print('forca curto circuito horizontal: ' + str(forcacurtocircuitohorizontal))  
    forcahorizontalporbraco = (forcacurtocircuitohorizontal+forcaventohorizontal+forcalinepullhorizontal)/(numbracos)
    print('forca horizontal por braco: ' + str(forcahorizontalporbraco))

    print('\nForças Verticais\n')

    #forcas verticais
    forcaventovertical = (((forcaventohorizontal*((alturart+mc1a)/1000)+(((diamext-espessurareator/2000))))/(diamext/1000*(math.sin(((2*3.1415)/numbracos)))))/(mc1a/1000))/2
    print('forcaventovertical: ' + str(forcaventovertical))
    forcacurtocircuitovertical = ((forcacurtocircuitohorizontal*((alturart/1000)/2+(mc1a/1000)))/((diamext/1000)*(math.sin(((2*3.1415)/numbracos)))))/(mc1a/1000)
    print('forcacurtocircuitovertical: ' + str(forcacurtocircuitovertical))
    forcapesovertical = ((667)*9.81)
    print('forcapesovertical: ' + str(forcapesovertical))
    forcalinepullvertical = forcalinepullhorizontal*((alturart+mc1a)/1000)/((diamext/1000)*(math.sin(((2*3.1415)/numbracos))))/(mc1a/500)
    print('forcalinepullvertical: ' + str(forcalinepullvertical))
    forcaverticalporbraco = (forcacurtocircuitovertical-2*forcapesovertical+forcalinepullvertical+forcaventovertical)/(numbracos)
    print('forca vertical por braco: ' + str(forcaverticalporbraco))

    print('\n')

    dadosisoladores = pd.read_excel(caminholm, sheet_name='DADOS ISOLADORES')
    #print(dadosisoladores['TIPO'][1])
    #print('nbi: ' + str(nbi))
    for niveldetensao in range(len(dadosisoladores['NBI (kV)'])):
        if(float(nbi) <= float(dadosisoladores['NBI (kV)'][niveldetensao])):            
            print(str(dadosisoladores['TIPO'][niveldetensao]) + ', flexão: ' + str(dadosisoladores['FLEXÃO (N)'][niveldetensao]) + ', tração: ' + str(dadosisoladores['TRAÇÃO (N)'][niveldetensao]) + ', compressão: ' + str(dadosisoladores['COMP. (N)'][niveldetensao]))

"""# CALCULO ESFORÇOS
arquivo = 'CHESF2500UH400A2'   
caminhodados = 'Resultados/' + arquivo + '.xlsx'
caminholm = 'Reator Fio\Modelos\Lista de materiais.xlsm'
wb_lm = load_workbook(caminholm)
calculoesforcos(caminhodados, wb_lm, caminholm)"""

criardocs()

seconds_final = time.time()
print("Result =", seconds_final - seconds_inicial)	