from platform import java_ver
from openpyxl import Workbook, load_workbook
import numpy as np
import math

# Coletar dados do reator


def coletardados(caminhodados, printar):
    wb_dados = load_workbook(caminhodados)
    dados = {}
    vetortitulo = []
    cont = 1
    for sheet in wb_dados:
        vetortitulo.append(sheet.title)
    reator = []
    for i in range(len(vetortitulo)):
        # print(vetortitulo[i])
        sheet_obj = wb_dados[vetortitulo[i]]
        max_row = sheet_obj.max_row
        max_column = sheet_obj.max_column
        reator.append([])
        for j in range(max_column):
            reator[i].append([])
            for k in range(max_row):
                reator[i][j].append(
                    str(sheet_obj.cell(row=k+1, column=j+1).value))
        if (printar == 1):
            print(reator[i])
            print('')
    return reator

# Coletar dados de custo


numcolunascusto = 0
numlinhascusto = 0


def coletardadosdecusto(caminhodadoscusto, printar):
    wb_dados = load_workbook(caminhodadoscusto)
    vetortitulo = []
    for sheet in wb_dados:
        vetortitulo.append(sheet.title)
    custo = []
    for i in range(len(vetortitulo)):
        # print(vetortitulo[i])
        sheet_obj = wb_dados[vetortitulo[i]]
        max_row = sheet_obj.max_row
        numlinhascusto = max_row
        max_column = sheet_obj.max_column
        numcolunascusto = max_column
        custo.append([])
        for j in range(max_column):
            custo[i].append([])
            for k in range(max_row):
                custo[i][j].append(
                    str(sheet_obj.cell(row=k+1, column=j+1).value))
        if (printar == 1):
            print(custo[i])
            print('')
    return custo, numlinhascusto, numcolunascusto

####################################################################################################################################
# ORDEM DE FABRICAÇÃO
####################################################################################################################################

# Criar OF RFE


def criarofrfe(caminhodados, wb_of):
    # CABECALHO
    # DADOS PROJETO
    # ASPECTOS CONSTRUTIVOS
    reator = coletardados(caminhodados, 0)
    # diametro externo
    wb_of['OF RFE'].cell(row=14, column=3).value = reator[1][5][1]
    # diametro interno
    wb_of['OF RFE'].cell(row=14, column=5).value = reator[1][4][1]
    wb_of['OF RFE'].cell(row=17, column=5).value = str(
        reator[1][8][1])[0:1]  # numero de bracos
    wb_of['OF RFE'].cell(row=17, column=13).value = len(
        reator[2][0])-1  # numero de cilindros

    # GABARITO
    wb_of['OF RFE'].cell(row=34, column=3).value = str(
        reator[1][7][1])[9:14]  # altura cruzeta
    # EMBALAGEM
    # OBSERVACOES

# Criar Bobinagem


def criarbobinagem(caminhodados, wb_of):
    reator = coletardados(caminhodados, 1)

    numcilindros = len(reator[2][0])-1
    numcamadas = reator[2][3]
    numfios = reator[2][2]
    fiocamada = reator[2][1]
    diamisol = reator[2][10]
    diamfionu = reator[2][11]
    isolamento = reator[1][1][1]
    alturaanel = reator[1][2][1]
    fibrainterna = reator[2][4]
    diamint = reator[1][4][1]
    ladomadeira = 30
    pesoespira = reator[3][12]
    numespiras = reator[3][3]
    cilindrodacamada = reator[3][0]
    numfio = len(reator[3][0])
    alturadacamada = reator[3][6]
    circunferenciaexterna = reator[3][8]
    numbracos = int(str(reator[1][7][1])[0:1])
    vetoraux = []
    contadorbraco = 0

    # print('numcilindros: ' + str(numcilindros) + ', numcamadas: ' +
    #      str(numcamadas) + ', numfios: ' + str(numfios) + ', pesoespira: ' + str(pesoespira) + ', cilindrodacamada: ' + str(cilindrodacamada))

    # Criar paginas bobinagem
    for cilindro in range(numcilindros-1):
        target = wb_of.copy_worksheet(wb_of['BOBINAGEM C1'])
    for cilindro in range(numcilindros-1):
        if (cilindro > 0):
            wb_of['BOBINAGEM C1' + ' Copy' +
                  str(cilindro)].title = 'BOBINAGEM C' + str(cilindro+2)
        else:
            wb_of['BOBINAGEM C1 Copy'].title = 'BOBINAGEM C' + str(cilindro+2)

    for cilindro in range(numcilindros):
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=4,
                                                    column=4).value = 'Cilindro ' + str(cilindro+1)  # numero do cilindro
        contadorfiosbracoinf = []
        contadorfiosbracosup = []
        for lado in range(numbracos):
            contadorfiosbracoinf.append([0]*numbracos)
            contadorfiosbracosup.append([0]*numbracos)

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=3).value = fiocamada[cilindro+1]  # fio utilizado na camada
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=5).value = diamisol[cilindro+1]  # diametro do fio isolado na camada
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=7).value = numfios[cilindro+1]  # num de fios na camada

        if (isolamento == '130'):
            isolamento = 'Classe B: Mylar'
        else:
            isolamento = 'Classe F: Teonex'
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=9).value = isolamento  # isolamento do reator

        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=12,
                                                    column=13).value = alturaanel  # altura do anel de fibra

        if (cilindro == 0):
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=15,
                                                        column=3).value = float(diamint) - float(fibrainterna[cilindro+1]) - float(ladomadeira)  # diametro do gabarito
        else:
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=14,
                                                        column=3).value = 'Diâmetro interno do cilindro'  # diametro interno do cilindro
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=15,
                                                        column=3).value = float(diamint) + float(fibrainterna[cilindro+1])  # diametro interno do cilindro

        contadorauxfios = 0
        vetoraux.append([])
        for aux in range(len(cilindrodacamada)-1):
            if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                vetoraux[cilindro].append(aux)

        cont = 0
        pesofio = 0
        for fio in range(numfio-1):
            if (int(cilindrodacamada[fio+1]) == int(cilindro)):
                areaaluminio = (
                    (float(diamfionu[cilindro+1])/1000) ** 2)*3.1415/4
                volumeespiraaluminio = areaaluminio * \
                    (float(circunferenciaexterna[fio+1])/1000)
                volumealuminio = volumeespiraaluminio*float(numespiras[fio+1])
                pesoaluminio = volumealuminio*2800
                pesofio = pesofio + pesoaluminio

                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=24+11*cont,
                                                            column=6).value = float(alturadacamada[fio+1])  # altura da camada
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=26+11*cont,
                                                            column=6).value = float(circunferenciaexterna[fio+1])  # perimetro externo da camada
                wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=20+11*cont,
                                                            column=10).value = numespiras[fio + 1]  # numero de espiras
                cont = cont + 1
        wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=7,
                                                    column=13).value = pesofio  # peso de fio

        for braco in range(numbracos):
            wb_of['BOBINAGEM C' +
                  str(cilindro+1)].cell(row=23+braco, column=13).value = braco  # enumerar bracos

        for camada in range(int(numcamadas[cilindro+1])):
            for aux in range(len(cilindrodacamada)-1):
                for fio in range(int(numfios[cilindro+1])):
                    if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                        if (vetoraux[cilindro][camada] == aux):
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=8).value = fio+1  # numero do fio axial
                            fioinf = contadorbraco
                            if (fioinf >= numbracos):
                                fioinf = 0
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=9).value = fioinf  # braco inferior bobinagem
                            fiosup = fioinf + \
                                math.floor(
                                    (float(numespiras[aux+1]) - int(float(numespiras[aux+1])))/(1/numbracos))
                            if (fiosup >= numbracos):
                                fiosup = fiosup - numbracos
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=11).value = fiosup  # braco superior bobinagem
                            contadorbraco = contadorbraco + 1
                            if (contadorbraco >= numbracos):
                                contadorbraco = 0
                            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=21+11*camada+fio,
                                                                        column=10).value = int(float(numespiras[aux + 1]))  # numero de espiras arredondado

                            # contador de fios nos bracos
                            for braco in range(numbracos):
                                if (fioinf == braco):
                                    contadorfiosbracoinf[cilindro][braco] = contadorfiosbracoinf[cilindro][braco] + 1
                                if (fiosup == braco):
                                    contadorfiosbracosup[cilindro][braco] = contadorfiosbracosup[cilindro][braco] + 1

        for braco in range(numbracos):
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=23+braco,
                                                        column=12).value = contadorfiosbracoinf[cilindro][braco]
            wb_of['BOBINAGEM C' + str(cilindro+1)].cell(row=23+braco,
                                                        column=14).value = contadorfiosbracosup[cilindro][braco]

# Criar controle de bobinagem


def criarcontrole(caminhodados, wb_of):
    reator = coletardados(caminhodados, 0)

    numcilindros = len(reator[2][0])-1
    numcamadas = reator[2][3]
    numfios = reator[2][2]
    fiocamada = reator[2][1]
    diamisol = reator[2][10]
    isolamento = reator[1][1][1]
    alturaanel = reator[1][2][1]
    fibrainterna = reator[2][4]
    diamint = reator[1][4][1]
    diamfionu = reator[2][11]
    ladomadeira = 30
    pesoespira = reator[3][12]
    numespiras = reator[3][3]
    cilindrodacamada = reator[3][0]
    numfio = len(reator[3][0])
    alturadacamada = reator[3][6]
    circunferenciaexterna = reator[3][8]
    numbracos = int(str(reator[1][7][1])[0:1])
    vetoraux = []
    contadorbraco = 0

    # Criar paginas controle
    for cilindro in range(numcilindros-1):
        target = wb_of.copy_worksheet(wb_of['CONTROLE C1'])
    for cilindro in range(numcilindros-1):
        if (cilindro > 0):
            wb_of['CONTROLE C1' + ' Copy' +
                  str(cilindro)].title = 'CONTROLE C' + str(cilindro+2)
        else:
            wb_of['CONTROLE C1 Copy'].title = 'CONTROLE C' + \
                str(cilindro+2)
        #print('CONTROLE C' + str(cilindro+2))

    for cilindro in range(numcilindros):
        vetoraux.append([])
        for aux in range(len(cilindrodacamada)-1):
            if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                vetoraux[cilindro].append(aux)

        wb_of['CONTROLE C' + str(cilindro+1)].cell(
            row=4, column=4).value = 'CILINDRO ' + str(cilindro+1)  # numero do cilindro

        for camada in range(len(vetoraux[cilindro])):
            if (camada == 1):
                wb_of['CONTROLE C' + str(cilindro+1)].cell(row=6,
                                                           column=5).value = float(diamint) - float(fibrainterna[cilindro+1])  # diametro interno da camada de fibra
            wb_of['CONTROLE C' + str(cilindro+1)].cell(row=7+3*camada,
                                                       column=5).value = float(alturadacamada[int(float(vetoraux[cilindro][camada]))+1])  # altura da camada
            wb_of['CONTROLE C' + str(cilindro+1)].cell(row=8+3*camada,
                                                       column=5).value = float(circunferenciaexterna[int(float(vetoraux[cilindro][camada]))+1])  # perimetro da camada

####################################################################################################################################
# LISTA DE MATERIAIS E ROTEIRO DE PRODUÇÃO
####################################################################################################################################


def criarlistademateriais(caminhodados, wb_lm, caminhodadoscusto):
    reator = coletardados(caminhodados, 0)

    numcilindros = len(reator[2][0])-1
    numcamadas = reator[2][3]
    numfios = reator[2][2]
    fiocamada = reator[2][1]
    diamisol = reator[2][10]
    diamfionu = reator[2][11]
    isolamento = reator[1][1][1]
    alturaanel = reator[1][2][1]
    fibrainterna = reator[2][4]
    fibraexterna = reator[2][5]
    diamint = float(reator[1][4][1])
    diamext = float(reator[1][5][1])
    ladomadeira = 30
    pesoespira = reator[3][12]
    numespiras = reator[3][3]
    cilindrodacamada = reator[3][0]
    numfio = len(reator[3][0])
    alturadacamada = reator[3][6]
    circunferenciaexterna = reator[3][8]
    numbracos = int(str(reator[1][7][1])[0:1])
    vetoraux = []
    contadorbraco = 0
    alturart = float(reator[1][26][1])
    acumulador = 0
    percentualroving = 0.6
    percentualtecido = 0.25*percentualroving
    percentualresina = 0.35*(percentualroving+percentualtecido)
    percentualacelerador = 0.03*percentualresina
    densidadealuminio = 2700
    comprimentodafita55 = 0
    numtotalespacadores = 0
    distanciaentreespacadores = 150
    diamespacador = float(reator[1][6][1])
    cruzeta = reator[1][7][1]
    ajusteembalagem = 40

    for cilindro in range(numcilindros):
        if (isolamento == '130'):
            isolamento = '4M036'
        else:
            isolamento = '2M036-2T0'
        wb_lm['LISTA DE MATERIAIS'].cell(
            row=5+cilindro, column=4).value = 'Fio alumínio isolado ' + str(fiocamada[cilindro+1]) + ' AWG - ' + str(isolamento)  # fio do cilindro

        cont = 0
        pesofio = 0
        for fio in range(numfio-1):
            if (int(cilindrodacamada[fio+1]) == int(cilindro)):
                areaaluminio = (
                    (float(diamfionu[cilindro+1])/1000) ** 2)*3.1415/4
                volumeespiraaluminio = areaaluminio * \
                    (float(circunferenciaexterna[fio+1])/1000)
                volumealuminio = volumeespiraaluminio*float(numespiras[fio+1])
                pesoaluminio = volumealuminio*2800
                pesofio = pesofio + pesoaluminio
                cont = cont + 1
        wb_lm['LISTA DE MATERIAIS'].cell(row=5+cilindro,
                                         column=5).value = pesofio  # peso de fio
        wb_lm['LISTA DE MATERIAIS'].cell(row=5+cilindro,
                                         column=6).value = '[kg]'

        vetoraux.append([])
        for aux in range(len(cilindrodacamada)-1):
            if (int(cilindrodacamada[aux+1]) == int(cilindro)):
                vetoraux[cilindro].append(aux)

        # VOLUME DA CAMADA DE FIBRA
        for camada in range(len(vetoraux[cilindro])):
            if (camada == 0):
                raiointcilindro = ((float(circunferenciaexterna[int(float(vetoraux[cilindro][camada]))+1])/(
                    2*3.1415)) - float(diamisol[cilindro+1]) - float(fibrainterna[cilindro+1]))
                raioextcilindro = ((float(circunferenciaexterna[int(
                    float(vetoraux[cilindro][camada]))+1])/(2*3.1415)) - float(diamisol[cilindro+1]))
                volumefibra = (3.1415*((float(raioextcilindro)/1000) ** 2) -
                               3.1415*((float(raiointcilindro)/1000) ** 2))*(float(alturart)/1000)
                acumulador = acumulador + volumefibra

            if (camada == len(vetoraux[cilindro])-1):
                raioextcilindro = ((float(circunferenciaexterna[int(float(vetoraux[cilindro][len(
                    vetoraux[cilindro])-1]))+1])/(2*3.1415)) + float(fibraexterna[cilindro+1]))
                raiointcilindro = ((float(circunferenciaexterna[int(
                    float(vetoraux[cilindro][len(vetoraux[cilindro])-1]))+1])/(2*3.1415)))
                volumefibra = (3.1415*((float(raioextcilindro)/1000) ** 2) -
                               3.1415*((float(raiointcilindro)/1000) ** 2))*(float(alturart)/1000)
                acumulador = acumulador + volumefibra

                # 'FITA ADESIVA POLIESTER 50MM X 66 M - BRA'
                numerodevoltasfita = math.floor(
                    float(alturadacamada[vetoraux[cilindro][camada]+1])/150)
                comprimentodafita55 = (
                    comprimentodafita55 + numerodevoltasfita*raioextcilindro)

                if (cilindro < numcilindros - 1):
                    numtotalespacadores = numtotalespacadores + math.floor((float(circunferenciaexterna[int(
                        float(vetoraux[cilindro][len(vetoraux[cilindro])-1]))+1])/distanciaentreespacadores))

    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=3).value = 'RT23017038'   # roving
    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=4).value = 'Roving contínuo 4400TEX'
    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=5).value = percentualroving*acumulador*1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=5+numcilindros,
                                     column=6).value = '[L]'

    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=3).value = 'RT23017001'   # tecido
    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=4).value = 'Tecido WR-600/3 0,20m'
    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=5).value = percentualtecido*acumulador*1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=6+numcilindros,
                                     column=6).value = '[L]'

    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=3).value = 'RT25010001'   # resina
    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=4).value = 'Resina epoxi araldite MY750 BR'
    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=5).value = percentualresina*acumulador*1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=7+numcilindros,
                                     column=6).value = '[L]'

    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=3).value = 'RT25020002'   # acelerador
    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=4).value = 'Acelerador DY 9577'
    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=5).value = percentualacelerador*acumulador*1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=8+numcilindros,
                                     column=6).value = '[L]'

    numerodevoltasperfil = math.floor(float(alturart)/300)
    comprimentodeperfil = numerodevoltasperfil*2*3.1415*((float(diamint)+23)/2)
    volumedeperfil = comprimentodeperfil*7.5*23/(1000 ** 3)
    pesodeperfil = volumedeperfil*densidadealuminio

    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=3).value = 'RT22013007'
    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=4).value = 'Perfil de alumínio 7,5 x 23 mm - EE223'
    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=5).value = pesodeperfil
    wb_lm['LISTA DE MATERIAIS'].cell(row=9+numcilindros,
                                     column=6).value = '[kg]'

    #'FITA ADESIVA POLIESTER 19MM X 66 M - BRA'

    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=3).value = 'RT24015001'
    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=4).value = 'Fita adesiva poliester 19mm x 66m - BRA'
    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=5).value = (comprimentodafita55/(1000*66))*0.15
    wb_lm['LISTA DE MATERIAIS'].cell(row=10+numcilindros,
                                     column=6).value = '[un]'

    #'FITA ADESIVA POLIESTER 50MM X 66 M - BRA'

    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=3).value = 'RT23019005'
    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=4).value = 'Fita adesiva poliester 50mm x 66m - BRA'
    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=5).value = comprimentodafita55/(1000*66)
    wb_lm['LISTA DE MATERIAIS'].cell(row=11+numcilindros,
                                     column=6).value = '[un]'

    superficiedoreator = (alturart*(diamint + diamext) +
                          ((diamext ** 2)/4)*3.1415 - ((diamint ** 2)/4)*3.1415)/1000000
    sumatanea = superficiedoreator/6.2
    sumataneb = sumatanea*0.47/3.13
    sumaclada = superficiedoreator/15.5
    sumacladb = sumaclada
    diluente = (sumatanea + sumataneb)*0.2

    # SUMATANE HB S/B COMP. A

    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=3).value = '6200A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=4).value = 'Sumatane HB S/B comp. A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=5).value = sumatanea
    wb_lm['LISTA DE MATERIAIS'].cell(row=12+numcilindros,
                                     column=6).value = '[L]'

    # SUMATANE HB S/B COMP. B

    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=3).value = '6200B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=4).value = 'Sumatane HB S/B comp. B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=5).value = sumataneb
    wb_lm['LISTA DE MATERIAIS'].cell(row=13+numcilindros,
                                     column=6).value = '[L]'

    # SUMACLAD 940 VERDE COMP. A

    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=3).value = '6300A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=4).value = 'Sumaclad 940 verde comp. A'
    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=5).value = sumaclada
    wb_lm['LISTA DE MATERIAIS'].cell(row=14+numcilindros,
                                     column=6).value = '[L]'

    # SUMACLAD 940 VERDE COMP. B

    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=3).value = '6300B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=4).value = 'Sumaclad 940 verde comp. B'
    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=5).value = sumacladb
    wb_lm['LISTA DE MATERIAIS'].cell(row=15+numcilindros,
                                     column=6).value = '[L]'

    # DILUENTE P/ TINTA DE ACABAMENT

    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=3).value = '7900'
    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=4).value = 'Diluente p/ tinta de acabamento'
    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=5).value = diluente
    wb_lm['LISTA DE MATERIAIS'].cell(row=16+numcilindros,
                                     column=6).value = '[L]'

    # PLACA IDENT. INOX 110X50X0,6MM

    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=3).value = 'RT12011003'
    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=4).value = 'Placa ident. inox 110x50x0,6mm'
    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=5).value = 1
    wb_lm['LISTA DE MATERIAIS'].cell(row=17+numcilindros,
                                     column=6).value = '[un]'

    # TELA  FV COM PROTEÇÃO UV

    areadatela = 2*3.1415*((diamint ** 2)/4)/1000000
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=3).value = 'RT23017043'
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=4).value = 'Tela FV com proteção UV'
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=5).value = areadatela
    wb_lm['LISTA DE MATERIAIS'].cell(row=18+numcilindros,
                                     column=6).value = '[m^2]'

    # FITA CADARÇO TEXIFITA A5-20L

    fitaa520 = diamext*3.1415*8/1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=3).value = 'RT23017003'
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=4).value = 'Fita cadarço TEXFITA A5-20L'
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=5).value = fitaa520
    wb_lm['LISTA DE MATERIAIS'].cell(row=19+numcilindros,
                                     column=6).value = '[m]'

    # FITA CADARÇO TEXIFITA B2-35

    fitab235 = (math.ceil(alturart/35)*(diamext*3.1415/2))/1000
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=3).value = 'RT23017002'
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=4).value = 'Fita cadarço TEXFITA B2-35'
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=5).value = fitab235
    wb_lm['LISTA DE MATERIAIS'].cell(row=20+numcilindros,
                                     column=6).value = '[m]'

    # REBITE ALUMINIO 3,2 X 16 MM

    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=3).value = 'RT13023001'
    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=4).value = 'Rebite de alumínio 3,2 x 16mm'
    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=5).value = 2
    wb_lm['LISTA DE MATERIAIS'].cell(row=21+numcilindros,
                                     column=6).value = '[un]'

    # espaçadores

    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=3).value = 'RT23017040'
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=4).value = 'Espaçadores ' + str(diamespacador) + 'mm, comp: ' + str(int(alturart))
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=5).value = numtotalespacadores
    wb_lm['LISTA DE MATERIAIS'].cell(row=22+numcilindros,
                                     column=6).value = '[un]'

    # cruzeta

    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=4).value = 'Cruzeta ' + str(int(diamext)) + 'mm ' + str(cruzeta)
    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=5).value = 1
    wb_lm['LISTA DE MATERIAIS'].cell(row=23+numcilindros,
                                     column=6).value = '[cj]'

    # pedestal

    wb_lm['LISTA DE MATERIAIS'].cell(row=24+numcilindros,
                                     column=4).value = 'Pedestal'
    wb_lm['LISTA DE MATERIAIS'].cell(row=24+numcilindros,
                                     column=5).value = numbracos
    wb_lm['LISTA DE MATERIAIS'].cell(row=24+numcilindros,
                                     column=6).value = '[un]'

    # sapata

    wb_lm['LISTA DE MATERIAIS'].cell(row=25+numcilindros,
                                     column=4).value = 'Sapata'
    wb_lm['LISTA DE MATERIAIS'].cell(row=25+numcilindros,
                                     column=5).value = numbracos
    wb_lm['LISTA DE MATERIAIS'].cell(row=25+numcilindros,
                                     column=6).value = '[un]'

    # embalagem

    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=3).value = 'RT42011XXX'
    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=4).value = 'Embalagem L: ' + str(math.ceil(int(diamext)/10)*10 + ajusteembalagem) + ' x C: ' + str(math.ceil(int(diamext)/10)*10 + ajusteembalagem) + ' x A: ' + str(math.ceil(int(alturart)/10)*10 + ajusteembalagem)
    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=5).value = 1
    wb_lm['LISTA DE MATERIAIS'].cell(row=26+numcilindros,
                                     column=6).value = '[un]'

    # Encontrar codigo do fio

    custo, numlinhascusto, numcolunascusto = coletardadosdecusto(
        caminhodadoscusto, 0)

    for cilindro in range(numcilindros):
        customediomaximo = 0
        for l in range(2, 41):
            for c in range(28, 29):

                if ((str(wb_lm['LISTA DE MATERIAIS'].cell(
                        row=l, column=c).value).split())[6] == '4M036'):
                    isolamento = 'MYLAR'
                else:
                    isolamento = 'TEONEX'

                descricaofio = str(str(round(float((str(wb_lm['LISTA DE MATERIAIS'].cell(
                    row=l, column=c).value).split())[3].replace(',', '.')), 1)) + ' ' + (str(wb_lm['LISTA DE MATERIAIS'].cell(
                        row=l, column=c).value).split())[4] + ' ' + (str(wb_lm['LISTA DE MATERIAIS'].cell(
                            row=l, column=c).value).split())[5] + ' ' + str(isolamento))

                if (isolamento == '130'):
                    isolamento = 'MYLAR'
                else:
                    isolamento = 'TEONEX'

                encontrarcodfio = str(str(
                    round(float(fiocamada[cilindro+1]), 1)) + ' AWG - ' + str(isolamento))

                if (encontrarcodfio == descricaofio):
                    wb_lm['LISTA DE MATERIAIS'].cell(
                        row=5+cilindro, column=3).value = str(wb_lm['LISTA DE MATERIAIS'].cell(row=l, column=27).value)  # codigo do fio do cilindro

    for cod in range(5, 26+numcilindros):
        customediomaximo = 0
        for linhacusto in range(numlinhascusto):
            #print(str(wb_lm['LISTA DE MATERIAIS'].cell(row=cod, column=3).value))
            if (str(wb_lm['LISTA DE MATERIAIS'].cell(row=cod, column=3).value) == custo[0][0][linhacusto]):
                # print(str(wb_lm['LISTA DE MATERIAIS'].cell(
                #    row=cod, column=3).value) + ' : ' + str(custo[0][1][linhacusto]))
                if (float(customediomaximo) < float(custo[0][1][linhacusto])):
                    customediomaximo = custo[0][1][linhacusto]
                    wb_lm['LISTA DE MATERIAIS'].cell(row=cod,
                                                     column=7).value = float(customediomaximo)
    
    return numtotalespacadores, superficiedoreator

def criarroteirodeproducao(caminhodados, wb_lm, caminhodadoscusto,numtotalespacadores, superficiedoreator):
    reator = coletardados(caminhodados, 0)
    numcilindros = len(reator[2][0])-1
    numcamadas = reator[2][3]
    numfios = reator[2][2]
    fiocamada = reator[2][1]
    diamisol = reator[2][10]
    diamfionu = reator[2][11]
    isolamento = reator[1][1][1]
    alturaanel = reator[1][2][1]
    fibrainterna = reator[2][4]
    fibraexterna = reator[2][5]
    diamint = float(reator[1][4][1])
    diamext = float(reator[1][5][1])
    ladomadeira = 30
    pesoespira = reator[3][12]
    numespiras = reator[3][3]
    cilindrodacamada = reator[3][0]
    numfio = len(reator[3][0])
    alturadacamada = reator[3][6]
    circunferenciaexterna = reator[3][8]
    numbracos = int(str(reator[1][7][1])[0:1])
    vetoraux = []
    contadorbraco = 0
    alturart = float(reator[1][26][1])
    acumulador = 0
    percentualroving = 0.6
    percentualtecido = 0.25*percentualroving
    percentualresina = 0.35*(percentualroving+percentualtecido)
    percentualacelerador = 0.03*percentualresina
    densidadealuminio = 2700
    comprimentodafita55 = 0
    distanciaentreespacadores = 150
    diamespacador = float(reator[1][6][1])
    cruzeta = reator[1][7][1]
    ajusteembalagem = 40
    numtotalespiras = 0
    numtotalfios = 0

    numaneis = int(alturart/300)
    for fio in range(len(numespiras)-1):
        numtotalespiras = numtotalespiras + float(numespiras[fio+1])   
    
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=6, column=7).value = float(numaneis)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=6, column=8).value = float(diamint)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=7, column=7).value = float(numaneis)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=8, column=7).value = float(numbracos)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=9, column=7).value = float(numbracos)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=10, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=10, column=8).value = float(numaneis)

    for cilindro in range(numcilindros):
        numtotalfios = numtotalfios + float(numfios[cilindro+1])

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=7).value = float(numtotalfios)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=8).value = float(diamext)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=9).value = float(numtotalespiras)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=11, column=10).value = float(numtotalespacadores)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=13, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=13, column=8).value = float(numaneis)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=14, column=7).value = float(numbracos)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=15, column=7).value = float(numbracos)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=15, column=8).value = float(numbracos)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=16, column=7).value = float(superficiedoreator)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=17, column=7).value = float(superficiedoreator)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=18, column=7).value = float(superficiedoreator)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=18, column=8).value = float(numbracos)

    if(int(alturart) > 700 or int(diamext) > 700):
        dificuldademovimentacao = 1
    else:
        dificuldademovimentacao = 0

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=19, column=8).value = float(dificuldademovimentacao)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=20, column=7).value = float(dificuldademovimentacao)

    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=21, column=7).value = float(alturart)
    wb_lm['ROTEIRO DE PRODUÇÃO'].cell(row=21, column=8).value = float(diamext)

####################################################################################################################################
# PROPOSTA
####################################################################################################################################

def criarproposta(wb_pp, caminhodados):
    reator = coletardados(caminhodados, 0)
    numcilindros = len(reator[2][0])-1
    numcamadas = reator[2][3]
    numfios = reator[2][2]
    fiocamada = reator[2][1]
    diamisol = reator[2][10]
    diamfionu = reator[2][11]
    isolamento = reator[1][1][1]
    alturaanel = reator[1][2][1]
    fibrainterna = reator[2][4]
    fibraexterna = reator[2][5]
    diamint = float(reator[1][4][1])
    diamext = float(reator[1][5][1])
    ladomadeira = 30
    pesoespira = reator[3][12]
    numespiras = reator[3][3]
    cilindrodacamada = reator[3][0]
    numfio = len(reator[3][0])
    alturadacamada = reator[3][6]
    circunferenciaexterna = reator[3][8]
    numbracos = int(str(reator[1][7][1])[0:1])
    vetoraux = []
    contadorbraco = 0
    alturart = float(reator[1][26][1])
    acumulador = 0
    percentualroving = 0.6
    percentualtecido = 0.25*percentualroving
    percentualresina = 0.35*(percentualroving+percentualtecido)
    percentualacelerador = 0.03*percentualresina
    densidadealuminio = 2700
    comprimentodafita55 = 0
    distanciaentreespacadores = 150
    diamespacador = float(reator[1][6][1])
    cruzeta = reator[1][7][1]
    ajusteembalagem = 40
    numtotalespiras = 0
    numtotalfios = 0
    indutancianominal = float(reator[0][11][1])
    frequencia = reator[0][0][1]
    impedancianominal = 2*3.1415*int(frequencia)*float(indutancianominal)/1000
    tensao = reator[0][1][1]
    nbi = reator[0][2][1]
    correntenominal = reator[0][13][1]
    correntecc = reator[0][6][1]
    duracaocc = reator[0][7][1]
    correnteccd = reator[0][8][1]
    perdasdc = reator[1][18][1]
    perdasac = reator[1][19][1]
    fatorq = reator[1][20][1]
    potencianominal = float(impedancianominal)*((float(correntenominal)) ** 2)/1000
    #if (isolamento == '130'):
    #    isolamento = 'MYLAR'
    #else:
    #    isolamento = 'TEONEX'
    if(((float(nbi)/0.45)+diamext) > (1.67*diamext)):
        de = (float(nbi)/0.45)+diamext
    else:
        de = 1.67*diamext
    mc1a = 0.5*diamext
    mc1r = 1.1*diamext
    altitude = reator[0][4][1]
    temperaturaambiente = reator[0][9][1]
    velocidadevento =reator[0][5][1]

    # PROPOSTA PT-BR

    wb_pp['PROPOSTA (PT-BR)'].cell(row=5, column=3).value = 'Reator(es), tipo '

    wb_pp['PROPOSTA (PT-BR)'].cell(row=9, column=4).value = indutancianominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=10, column=4).value = impedancianominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=11, column=4).value = tensao
    wb_pp['PROPOSTA (PT-BR)'].cell(row=12, column=4).value = nbi
    wb_pp['PROPOSTA (PT-BR)'].cell(row=13, column=4).value = frequencia
    wb_pp['PROPOSTA (PT-BR)'].cell(row=14, column=4).value = '' #frequencia de sintonia
    wb_pp['PROPOSTA (PT-BR)'].cell(row=15, column=4).value = correntenominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=16, column=4).value = str(correntecc) + ' / ' + str(duracaocc)
    wb_pp['PROPOSTA (PT-BR)'].cell(row=17, column=4).value = correnteccd
    wb_pp['PROPOSTA (PT-BR)'].cell(row=18, column=4).value = float(perdasdc) + float(perdasac)
    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=4).value = fatorq
    wb_pp['PROPOSTA (PT-BR)'].cell(row=20, column=5).value = '' # fator q a frequencia de sintonia
    wb_pp['PROPOSTA (PT-BR)'].cell(row=21, column=4).value = potencianominal
    wb_pp['PROPOSTA (PT-BR)'].cell(row=22, column=4).value = 'A.N.'
    wb_pp['PROPOSTA (PT-BR)'].cell(row=23, column=4).value = isolamento

    wb_pp['PROPOSTA (PT-BR)'].cell(row=28, column=4).value = alturart
    wb_pp['PROPOSTA (PT-BR)'].cell(row=29, column=4).value = diamext
    wb_pp['PROPOSTA (PT-BR)'].cell(row=30, column=4).value = '' #diametro fundacao
    wb_pp['PROPOSTA (PT-BR)'].cell(row=31, column=4).value = '' #peso por modulo
    wb_pp['PROPOSTA (PT-BR)'].cell(row=32, column=4).value = '' #peso total

    wb_pp['PROPOSTA (PT-BR)'].cell(row=9, column=8).value = 'ABNT NBR 5356-06'

    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=10).value = de
    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=21).value = mc1a
    wb_pp['PROPOSTA (PT-BR)'].cell(row=19, column=23).value = mc1r

    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = altitude
    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = temperaturaambiente
    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = velocidadevento

    wb_pp['PROPOSTA (PT-BR)'].cell(row=36, column=9).value = str(int(diamext)) + ' x ' + str(int(diamext)) + ' x ' + str(int(alturart)) 

    # PROPOSTA EN-US

    wb_pp['PROPOSTA (EN-US)'].cell(row=5, column=3).value = 'Air Coil(s), type '

    wb_pp['PROPOSTA (EN-US)'].cell(row=9, column=4).value = indutancianominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=10, column=4).value = impedancianominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=11, column=4).value = tensao
    wb_pp['PROPOSTA (EN-US)'].cell(row=12, column=4).value = nbi
    wb_pp['PROPOSTA (EN-US)'].cell(row=13, column=4).value = frequencia
    wb_pp['PROPOSTA (EN-US)'].cell(row=14, column=4).value = '' #frequencia de sintonia
    wb_pp['PROPOSTA (EN-US)'].cell(row=15, column=4).value = correntenominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=16, column=4).value = str(correntecc) + ' / ' + str(duracaocc)
    wb_pp['PROPOSTA (EN-US)'].cell(row=17, column=4).value = correnteccd
    wb_pp['PROPOSTA (EN-US)'].cell(row=18, column=4).value = float(perdasdc) + float(perdasac)
    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=4).value = fatorq
    wb_pp['PROPOSTA (EN-US)'].cell(row=20, column=5).value = '' # fator q a frequencia de sintonia
    wb_pp['PROPOSTA (EN-US)'].cell(row=21, column=4).value = potencianominal
    wb_pp['PROPOSTA (EN-US)'].cell(row=22, column=4).value = 'A.N.'
    wb_pp['PROPOSTA (EN-US)'].cell(row=23, column=4).value = isolamento

    wb_pp['PROPOSTA (EN-US)'].cell(row=28, column=4).value = alturart
    wb_pp['PROPOSTA (EN-US)'].cell(row=29, column=4).value = diamext
    wb_pp['PROPOSTA (EN-US)'].cell(row=30, column=4).value = '' #diametro fundacao
    wb_pp['PROPOSTA (EN-US)'].cell(row=31, column=4).value = '' #peso por modulo
    wb_pp['PROPOSTA (EN-US)'].cell(row=32, column=4).value = '' #peso total

    wb_pp['PROPOSTA (EN-US)'].cell(row=9, column=8).value = 'ABNT NBR 5356-06'

    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=10).value = de
    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=21).value = mc1a
    wb_pp['PROPOSTA (EN-US)'].cell(row=19, column=23).value = mc1r

    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = altitude
    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = temperaturaambiente
    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = velocidadevento

    wb_pp['PROPOSTA (EN-US)'].cell(row=36, column=9).value = str(int(diamext)) + ' x ' + str(int(diamext)) + ' x ' + str(int(alturart)) 

    # PROPOSTA ES-ES

    wb_pp['PROPOSTA (ES-ES)'].cell(row=5, column=3).value = 'Reator(es), tipo '

    wb_pp['PROPOSTA (ES-ES)'].cell(row=9, column=4).value = indutancianominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=10, column=4).value = impedancianominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=11, column=4).value = tensao
    wb_pp['PROPOSTA (ES-ES)'].cell(row=12, column=4).value = nbi
    wb_pp['PROPOSTA (ES-ES)'].cell(row=13, column=4).value = frequencia
    wb_pp['PROPOSTA (ES-ES)'].cell(row=14, column=4).value = '' #frequencia de sintonia
    wb_pp['PROPOSTA (ES-ES)'].cell(row=15, column=4).value = correntenominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=16, column=4).value = str(correntecc) + ' / ' + str(duracaocc)
    wb_pp['PROPOSTA (ES-ES)'].cell(row=17, column=4).value = correnteccd
    wb_pp['PROPOSTA (ES-ES)'].cell(row=18, column=4).value = float(perdasdc) + float(perdasac)
    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=4).value = fatorq
    wb_pp['PROPOSTA (ES-ES)'].cell(row=20, column=5).value = '' # fator q a frequencia de sintonia
    wb_pp['PROPOSTA (ES-ES)'].cell(row=21, column=4).value = potencianominal
    wb_pp['PROPOSTA (ES-ES)'].cell(row=22, column=4).value = 'A.N.'
    wb_pp['PROPOSTA (ES-ES)'].cell(row=23, column=4).value = isolamento

    wb_pp['PROPOSTA (ES-ES)'].cell(row=28, column=4).value = alturart
    wb_pp['PROPOSTA (ES-ES)'].cell(row=29, column=4).value = diamext
    wb_pp['PROPOSTA (ES-ES)'].cell(row=30, column=4).value = '' #diametro fundacao
    wb_pp['PROPOSTA (ES-ES)'].cell(row=31, column=4).value = '' #peso por modulo
    wb_pp['PROPOSTA (ES-ES)'].cell(row=32, column=4).value = '' #peso total

    wb_pp['PROPOSTA (ES-ES)'].cell(row=9, column=8).value = 'ABNT NBR 5356-06'

    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=10).value = de
    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=21).value = mc1a
    wb_pp['PROPOSTA (ES-ES)'].cell(row=19, column=23).value = mc1r

    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = altitude
    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = temperaturaambiente
    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = velocidadevento

    wb_pp['PROPOSTA (ES-ES)'].cell(row=36, column=9).value = str(int(diamext)) + ' x ' + str(int(diamext)) + ' x ' + str(int(alturart)) 

####################################################################################################################################

# Criar Ordem de Fabricação


def criarof(caminhodados):
    caminhoof = 'Reator Fio\Modelos\Ordem de fabricação.xlsm'
    wb_of = load_workbook(caminhoof)
    criarofrfe(caminhodados, wb_of)
    criarbobinagem(caminhodados, wb_of)
    criarcontrole(caminhodados, wb_of)
    wb_of.save("Reator Fio\Saida\Ordem de fabricação (Saida).xlsx")

# Criar Lista de Materiais


def criarlm(caminhodados, caminhodadoscusto):
    caminholm = 'Reator Fio\Modelos\Lista de materiais.xlsm'
    wb_lm = load_workbook(caminholm)
    numtotalespacadores,superficiedoreator = criarlistademateriais(caminhodados, wb_lm, caminhodadoscusto)
    criarroteirodeproducao(caminhodados, wb_lm, caminhodadoscusto,numtotalespacadores,superficiedoreator)
    wb_lm.save("Reator Fio\Saida\Lista de materiais (Saida).xlsx")

# Criar Proposta

def criarpp(caminhodados):
    caminhopp = 'Reator Fio\Modelos\Proposta técnica.xlsm'
    wb_pp = load_workbook(caminhopp)
    criarproposta(wb_pp, caminhodados)
    wb_pp.save("Reator Fio\Saida\Proposta técnica (Saida).xlsx")

# Criar Documentos


def criardocs(caminhodados = 'Resultados\magnesita.xlsx'):    
    caminhodadoscusto = 'Reator Fio\Base de dados\CustoMaterial.xlsx'
    print('\nCriando Ordem de Fabricação, Roteiro de Bobinagem e Controle de Bobinagem...\n')
    criarof(caminhodados)
    print('Criando Lista de Materiais e Roteiro de Produção...\n')
    criarlm(caminhodados, caminhodadoscusto)
    print('Criando Desenho e Documento de Proposta...\n')
    criarpp(caminhodados)
    print('Finalizado o Processo de Emissão dos Documentos.\n')


#criardocs()
