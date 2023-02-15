from platform import java_ver
from openpyxl import Workbook, load_workbook
import numpy as np
import math

wb_dados = load_workbook('Resultados\saida_teste.xlsx')
dados = {}
vetortitulo = []
cont = 1

for sheet in wb_dados:
    vetortitulo.append(sheet.title)
reator = []
for i in range(len(vetortitulo)):
    print(vetortitulo[i])
    sheet_obj = wb_dados[vetortitulo[i]]
    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column
    reator.append([])
    for j in range(max_column):
        reator[i].append([])
        for k in range(max_row):
            reator[i][j].append(str(sheet_obj.cell(row=k+1, column=j+1).value))
    print(reator[i])
    print('')

wb_of = load_workbook('Reator Fio\Ordem de fabricação.xlsm')

###########################################################################################
# ORDEM DE FABRICACAO
###########################################################################################
# CABECALHO


# DADOS PROJETO

# ASPECTOSCONSTRUTIVOS
wb_of['OF RFE'].cell(row=14, column=3).value = reator[1][6][1]
wb_of['OF RFE'].cell(row=14, column=5).value = reator[1][5][1]
wb_of['OF RFE'].cell(row=17, column=5).value = str(reator[1][9][1])[0:1]
wb_of['OF RFE'].cell(row=17, column=13).value = len(reator[3][1])-2

# GABARITO
wb_of['OF RFE'].cell(row=34, column=3).value = str(reator[1][9][1])[9:14]

# EMBALAGEM

# OBSERVACOES

###########################################################################################
# BOBINAGEM
###########################################################################################
for i in range(len(reator[2][1])-2):
    target = wb_of.copy_worksheet(wb_of['BOBINAGEM C1'])

for i in range(len(reator[2][1])-2):
    l = i
    if (i > 0):
        l = i
        wb_of['BOBINAGEM C1' + ' Copy' +
              str(l)].title = 'BOBINAGEM C' + str(l+2)
    else:
        wb_of['BOBINAGEM C1 Copy'].title = 'BOBINAGEM C' + str(l+2)
    print('BOBINAGEM C' + str(l+2))

fioinf = 0
lastj = 0
contaux = 0
contfiosbraco = []
for i in range(len(reator[2][1])-1):
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=4,
                                         column=4).value = 'Cilindro ' + str(i+1)
    pesodefio = 0
    vetor = []
    for j in range(len(reator[3][1])):
        if (str(reator[3][0][j]) == str(i)):
            pesodefio = pesodefio + \
                float(reator[3][3][j])*float(reator[3][13][j])/10
            if (i >= 1):
                wb_of['BOBINAGEM C' + str(i+1)].cell(row=14,
                                                     column=3).value = 'Diametro interno do cilindro ' + str(i+1)
                wb_of['BOBINAGEM C' + str(i+1)].cell(row=15,
                                                     column=3).value = float(reator[3][8][j])/(3.1415) + float(reator[1][6][1]) + float(reator[2][4][i+1])
            vetor.append(j)

    #contadorbraco = [0]
    #numbracos = int(str(reator[1][9][1])[0:1])
    # for cil in range(12):
    #    for passo in range(74):
    #        for braco in range(numbracos):
    #            if (sheet_obj.cell(row=passo+1, column=9).value == braco):
    #                contadorbraco[braco] = contadorbraco[braco] + 1
    #                print()
    # print(contadorbraco)

    # numero de espiras
    for k in range(len(vetor)):
        wb_of['BOBINAGEM C' + str(i+1)].cell(row=20+11*k,
                                             column=10).value = float(reator[3][3][vetor[k]])
        # numero de fios axiais
        for l in range(int(reator[2][2][i+1])):
            wb_of['BOBINAGEM C' + str(i+1)].cell(row=21+l+11*k,
                                                 column=8).value = l+1
            wb_of['BOBINAGEM C' + str(i+1)].cell(row=23+l,
                                                 column=13).value = l
            # distribuicao dos fios na cruzeta
            if (str(reator[3][0][j]) == str(i)):
                # if (j != lastj):
                numcamadas = reator[2][3][i+1]
                wb_of['BOBINAGEM C' + str(i+1)].cell(row=21+l+11*contaux,
                                                     column=9).value = fioinf

                numbracos = int(str(reator[1][9][1])[0:1])
                # for braco in range(numbracos):
                #    contfiosbraco.append([])
                #    if (fioinf == braco):
                #        contfiosbraco[braco] = contfiosbraco[braco] + 1
                #        print(
                #            'contfiosbraco[' + str(contaux) + ']: ' + contfiosbraco[contaux])

                numespiras = float(reator[3][3][k+1])

                fiosup = fioinf + math.floor(
                    (numespiras - int(numespiras))/(1/numbracos)) + 1
                if (fiosup >= numbracos + 1):
                    fiosup = fiosup - numbracos - 1
                fioinf = fioinf + 1
                wb_of['BOBINAGEM C' + str(i+1)].cell(row=21+l+11*contaux,
                                                     column=11).value = fiosup
                wb_of['BOBINAGEM C' + str(i+1)].cell(row=21+l+11*contaux,
                                                     column=10).value = int(numespiras)
                if (fioinf > int(str(reator[1][9][1])[0:1])):
                    fioinf = 0

                if (l < int(reator[2][2][i+1]) - 2):
                    wb_of['BOBINAGEM C' + str(i+1)].cell(row=24 + 11*l,
                                                         column=6).value = float(reator[3][6][l+1])
                if (l < int(reator[2][2][i+1]) - 2):
                    wb_of['BOBINAGEM C' + str(i+1)].cell(row=26 + 11*l,
                                                         column=6).value = float(reator[3][8][l+1])

        contaux = contaux + 1
    #lastj = j
    contaux = 0
    #fioinf = fioinf + 1
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=7, column=13).value = pesodefio
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=12,
                                         column=3).value = float(reator[2][1][i+1])
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=12,
                                         column=5).value = float(reator[2][10][i+1])
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=12,
                                         column=7).value = float(reator[2][2][i+1])*float(reator[2][2][i+1])
    if ((reator[1][1][1]) == '130'):
        isolamento = 'Classe B: Mylar'
    else:
        isolamento = 'Classe F: Teonex'
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=12,
                                         column=9).value = isolamento
    wb_of['BOBINAGEM C' + str(i+1)].cell(row=12,
                                         column=13).value = float(reator[1][2][1])
    wb_of['BOBINAGEM C1'].cell(row=15,
                               column=3).value = float(reator[1][4][1])-float(reator[2][4][1])

###########################################################################################
# CONTROLE
###########################################################################################
for i in range(len(reator[2][1])-2):
    target = wb_of.copy_worksheet(wb_of['CONTROLE C1'])

for i in range(len(reator[2][1])-2):
    l = i
    if (i > 0):
        l = i
        wb_of['CONTROLE C1' + ' Copy' +
              str(l)].title = 'CONTROLE C' + str(l+2)
    else:
        wb_of['CONTROLE C1 Copy'].title = 'CONTROLE C' + str(l+2)

wb_of.save("Reator Fio\sample.xlsx")
