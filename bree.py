from modulos import bree_reactor_core_1 as reactor
from modulos import bree_capacitor_core_1 as calc_cap
from modulos import criar_documentos as criar
from ast import Pass
from numpy import typename
import pandas as pd
# import xlwings as xl
from statistics import mean
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
import formulas
#from styleframe import StyleFrame, Styler, utils
#from pandasql import sqldf
import duckdb
# import dataframe_sql #import register_temp_table, query
import sys
import os
from pathlib import Path
from PySide2 import QtWidgets
from PySide2 import QtCore
from PySide2 import QtGui
from PySide2.QtWidgets import QApplication, QDialog, QMainWindow, QFileDialog, QTableView, QMessageBox, QPushButton
from PySide2.QtCore import Qt, Signal, QObject, QModelIndex, QAbstractTableModel

from template.main_system_window import Ui_MainWindow
from template.capacitor_Potencia import Ui_win_CpEn
from template.capacitor_Frequencia import Ui_Win_CF
from template.fusein import Ui_Win_FuseIn
from template.resistor import Ui_Win_Resistor
# Ui_reactorReactorMainWindow
from template.main_reactor_window import Ui_BreeReactorMainWindow
from template.SysDataDlg_SpinBox import Ui_DialogSys
from template.ConstrDlg_SpinBox import Ui_DialogConstr
from template.DiagHarm import Ui_DialogHarm
from template.DialogCrudFibra import Ui_DialogCrudFibra
from template.Force_Side_by_Side import Ui_DialogForceIcc
from template.reator_aberto_modulo_hibrido4 import Ui_DialogHibrido
from template.obsDlg import Ui_DialogObs
import Cad.breecad as BreeCad
# sys.path.append('OneDrive/Documentos/reactor')
sys.path.append('OneDrive/Documentos/GitHub/reactor_reactor')
sys.path.append('Documents/GitHub')

#from pathlib import Path
path_name = os.getcwd()
path_name = path_name = path_name.replace("\\", "/")
path_name = path_name = path_name.replace("C:", "")
path_name
path_bree = 'H:/4-Engenharias/4.8-Engenharia de Produto/Software de Cálculo de Capacitores e Reatores'

#pysqldf = lambda q: sqldf(q, globals())
con = duckdb.connect()

# import reactor_reactor_core_1 as reactor

# def df_from_excel(path):
#     app = xl.App(visible=False)
#     book = app.books.open(path)
#     book.save(reactor.path_name+'\\Engenharia Reversa\\modelo1.xlsx')
#     app.kill()
#     return pd.read_excel(path)

class DataFrameModel(QtCore.QAbstractTableModel):
    DtypeRole = QtCore.Qt.UserRole + 1000
    ValueRole = QtCore.Qt.UserRole + 1001

    def __init__(self, df=pd.DataFrame(), parent=None):
        super(DataFrameModel, self).__init__(parent)
        self._dataframe = df

    def setDataFrame(self, dataframe):
        self.beginResetModel()
        self._dataframe = dataframe.copy()
        self.endResetModel()

    def dataFrame(self):
        return self._dataframe

    #dataFrame = QtCore.pyqtProperty(pd.DataFrame, fget=dataFrame, fset=setDataFrame)
    dataFrame = QtCore.Property(
        pd.DataFrame, fget=dataFrame, fset=setDataFrame)

    @QtCore.Slot(int, QtCore.Qt.Orientation, result=str)
    def headerData(self, section: int, orientation: QtCore.Qt.Orientation, role: int = QtCore.Qt.DisplayRole):
        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                return self._dataframe.columns[section]
            else:
                return str(self._dataframe.index[section])
        return QtCore.QVariant()

    def rowCount(self, parent=QtCore.QModelIndex()):
        if parent.isValid():
            return 0
        return len(self._dataframe.index)

    def columnCount(self, parent=QtCore.QModelIndex()):
        if parent.isValid():
            return 0
        return self._dataframe.columns.size

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            if role == Qt.DisplayRole or role == Qt.EditRole:
                value = self._dataframe.iloc[index.row(), index.column()]
                if isinstance(value, float):
                    # Render float to 2 dp
                    return "%.5f" % value
                else:
                    return str(value)
            if role == Qt.TextAlignmentRole:
                value = self._dataframe.iloc[index.row(), index.column()]
                if isinstance(value, int) or isinstance(value, float):
                    return int(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

    def roleNames(self):
        roles = {
            QtCore.Qt.DisplayRole: b'display',
            DataFrameModel.DtypeRole: b'dtype',
            DataFrameModel.ValueRole: b'value'
        }
        return roles

    def setData(self, index, value, role):
        if role == Qt.EditRole:
            self._dataframe.iloc[index.row(), index.column()] = value
            return True
        return False

    def flags(self, index):
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable

    def insertRows(self, position, rows, QModelIndex, parent):
        self.beginInsertRows(QModelIndex, position, position+rows-1)
        #default_row = [[None] for _ in range(self._dataframe.shape[1])]
        default_row = [[self._dataframe.iloc[position, col]]
                       for col in range(self._dataframe.shape[1])]
        new_df = pd.DataFrame(
            dict(zip(list(self._dataframe.columns), default_row)))
        self._dataframe = pd.concat([self._dataframe, new_df])
        self._dataframe = self._dataframe.reset_index(drop=True)
        self.endInsertRows()
        self.layoutChanged.emit()
        return True

    def removeRows(self, position, rows, QModelIndex):
        start, end = position, rows
        # if remove these 02 lines, it works
        self.beginRemoveRows(QModelIndex, start, end)
        self._dataframe.drop(position, inplace=True)
        self._dataframe.reset_index(drop=True, inplace=True)
        self.endRemoveRows()  # if remove these 02 lines, it works
        self.layoutChanged.emit()
        return True


class Resistor(QDialog, Ui_Win_Resistor):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        self.setupUi(self)
        self.table = QtWidgets.QTableView()
        self.table.setParent = self.frame_DF
        self.bt_calc.clicked.connect(self.calc_resistor)

    @QtCore.Slot()
    def calc_resistor(self):
        self.tabWidget.setCurrentIndex(1)
        conex = self.cb_Conex.currentIndex() == 0
        tensao = float(self.ed_Vn.text())
        fr, kvar = float(self.ed_Fr.text()), float(self.ed_KVAR.text())
        maxW_R = float(self.ed_max_W_R.text())
        tempo, vf = float(self.ed_tempo.text()), float(self.ed_Vend.text())
        file_name = self.ed_file_name.text()
        if conex:
            self.df = calc_cap.rs(tensao, fr, kvar, maxW_R, tempo, vf)
        else:
            self.df = calc_cap.rsp(tensao, fr, kvar, maxW_R, tempo, vf)
        df1 = self.df.T
        tipo = '\\resultados\\'
        path_name = os.getcwd()
        path_name = path_name+tipo+file_name+'.xlsx'
        with pd.ExcelWriter(path_name) as writer:
            self.df.to_excel(writer, sheet_name="Resistores de descarga")

        model = DataFrameModel(self.df.reset_index())
        self.tableView_1.setModel(model)


class Fusein(QDialog, Ui_Win_FuseIn):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        self.setupUi(self)
        self.table = QtWidgets.QTableView()
        self.table.setParent = self.frame_DF
        self.hiddenConex()
        self.bt_calc.clicked.connect(self.calc_fusein)
        self.cb_Conex.currentTextChanged.connect(
            self.conex_change)
        self.bt_read_file.clicked.connect(self.open_fusein)
        self.bt_save_file.clicked.connect(self.save_fusein)

    @QtCore.Slot()
    def save_fusein(self):
        pass

    @QtCore.Slot()
    def open_fusein(self):
        self.openFileNameDialog()

    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(
            self, "Abrir arquivo de projeto", "", "All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            print(fileName)

    def hiddenConex(self, invisivel=True):
        self.label_S1.setHidden(invisivel)
        self.label_S2.setHidden(invisivel)
        self.label_P1.setHidden(invisivel)
        self.label_P2.setHidden(invisivel)
        self.label_P1.setHidden(invisivel)
        self.ed_S1.setHidden(invisivel)
        self.ed_S2.setHidden(invisivel)
        self.ed_P1.setHidden(invisivel)
        self.ed_P2.setHidden(invisivel)

    @QtCore.Slot()
    def conex_change(self):
        conex = self.cb_Conex.currentText()
        if conex == 'Y' or conex == 'I' or conex == 'D' or conex == 'Y0':
            self.hiddenConex()
        elif conex == 'H' or conex == 'DH' or conex == 'YH' or conex == 'Y0H':
            self.hiddenConex(False)
        elif conex == 'HI' or conex == 'DHI' or conex == 'Y0HI' or conex == 'YHI':
            self.hiddenConex(False)
            self.ed_P0.setHidden(True)
            self.label_P0.setHidden(True)
        elif conex == 'YY':
            self.hiddenConex(False)
            self.ed_P0.setHidden(True)
            self.label_P0.setHidden(True)
            self.ed_S1.setHidden(True)
            self.ed_S2.setHidden(True)
            self.label_S1.setHidden(True)
            self.label_S2.setHidden(True)

    @QtCore.Slot()
    def calc_fusein(self):
        self.tabWidget.setCurrentIndex(1)
        num_elos = self.spinBoxNumElos.value()
        tensao = float(self.ed_Vn.text())
        fr, kvar = float(self.ed_Fr.text()), float(self.ed_KVAR.text())
        cap = float(self.ed_Cap.text())
        size = int(self.ed_size_un.text())
        series = int(self.ed_G_SI.text())
        paralelo = int(self.ed_GPI.text())
        size = int(self.ed_size_un.text())
        conexao = self.cb_Conex.currentText()
        s0 = int(self.ed_serie.text())
        p0 = int(self.ed_P0.text())
        s1 = int(self.ed_S1.text())
        s2 = int(self.ed_S2.text())
        p1 = int(self.ed_P1.text())
        p2 = int(self.ed_P2.text())
        pst = float(self.ed_PST.text())
        kurl = float(self.ed_KURL.text())
        kmin = float(self.ed_KMIN.text())
        file_name = self.ed_file_name.text()
        self.df1, df2, self.df3 = calc_cap.fusein(
            tensao, kvar, cap, fr, size, series, paralelo, conexao, s0, p0, s1, s2, p1, p2, pst, kurl, kmin, num_elos=num_elos)
        self.df = df2.T
        self.ed_GPI.setText(str(df2['Grupos paralelos'][0]))
        tipo = '\\resultados\\'
        path_name = os.getcwd()
        path_name = path_name+tipo+file_name+'.xlsx'
        with pd.ExcelWriter(path_name) as writer:
            self.df.to_excel(writer, sheet_name="Fusível interno")
            self.df1.to_excel(writer, sheet_name="Aprovados")
            self.df3.to_excel(writer, sheet_name="Tabela")

        model1 = DataFrameModel(self.df.reset_index())
        model2 = DataFrameModel(self.df1)
        model3 = DataFrameModel(self.df3)
        self.tableView_1.setModel(model1)
        self.tableView_2.setModel(model2)
        self.tableView_3.setModel(model3)
        #print(typename, type(typename))


class CF(QDialog, Ui_Win_CF):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        self.setupUi(self)
        self.table = QtWidgets.QTableView()
        self.table.setParent = self.frame_DF
        self.bt_calc.clicked.connect(self.calc_cf)

    @QtCore.Slot()
    def calc_cf(self):
        self.tabWidget.setCurrentIndex(1)
        eixo = self.cb_Eixos.currentIndex() == 0
        papelBk, filmeDk, oleoDK, tensao = float(self.ed_PapelBK.text()), float(
            self.ed_FilmeDK.text()), float(self.ed_OleoDK.text()), float(self.ed_Vn.text())
        fr, kvar = float(self.ed_Fr.text()), float(self.ed_KVAR.text())
        mils_paper, mils_film, mils_al = float(self.Ed_MilsPaper.text()), float(
            self.ed_milsFilm.text()), float(self.ed_mils_Al.text())
        #larg_max, larg_min = float(self.ed_larg_max.text()), float(self.ed_larg_min.text())
        larg_ativa = float(self.cb_larg_ativa.currentText())
        esp_coil, lar_coil = float(self.ed_Esp_coil.text()), float(
            self.ed_Larg_Coil.text())
        tol_c, fator_espc = float(self.ed_tol.text()), float(
            self.ed_FatorEsp.text())
        rho_papel, custo_papel = float(
            self.ed_rhoPaper.text()), float(self.ed_custoPaper.text())
        custo_filme = float(self.ed_custoFilm.text())
        refugo = float(self.ed_Refugo.text())
        file_name = self.ed_file_name.text()
        custo_al = float(self.ed_custo_al.text())

        if eixo:
            df = calc_cap.c_hf(papelBk, filmeDk, oleoDK, 47800, tensao, fr, kvar, esp_coil, lar_coil, mils_al,
                               0, 0, tol_c, mils_paper, custo_papel, mils_film, custo_filme, larg_ativa, fator_espc)
        else:
            df = calc_cap.c_hf(papelBk, filmeDk, oleoDK, 47800, tensao, fr, kvar, esp_coil, lar_coil, mils_al,
                               0, 0, tol_c, mils_paper, custo_papel, mils_film, custo_filme, larg_ativa, fator_espc)
            '''df = calc_cap.c_hf(papelBk,filmeDk,oleoDK,47800,tensao,fr,kvar,esp_coil,lar_coil,mils_al,larg_max,
            larg_min,tol_c,mils_paper,custo_papel,mils_film,custo_filme,11,fator_espc)'''
        if type(df) is str:
            button = QMessageBox.critical(
                self,
                "Oh dear!",
                "Something went very wrong.",
                buttons=QMessageBox.Discard | QMessageBox.NoToAll | QMessageBox.Ignore,
                defaultButton=QMessageBox.Discard,
            )
        else:
            df1 = df .T
            tipo = '\\resultados\\'
            path_name = os.getcwd()
            path_name = path_name+tipo+file_name+'.xlsx'
            with pd.ExcelWriter(path_name) as writer:
                df1.to_excel(writer, sheet_name="Capacitor Alta Frequência")

            #print(df1.iloc[1:, :])
            model = DataFrameModel(df1.reset_index())
            self.tableView.setModel(model)


class CP(QDialog, Ui_win_CpEn):
    # dicionário de alertas de Volts/bobina e Stress em VPM de acordo com o modelo do capacitor de potência
    alertas = {'MI': [2100, 1900], 'MH': [2000, 1800], 'MG': [1800, 1600], 'MF': [
        1600, 1440], 'ME': [1380, 1270], 'ST': [1500, 1000], 'VDC': [4000, 3600]}
    filmes = ['24', '40', '47', '50', '55', '60']
    codigos_filmes = ["B03018AAD", "B03018AED",
                      "B03018AFD", "B03018AGD", "B03018AHD", "B03018AID"]                      

    d_br = ['5.0', '3.65', '3.35', '3.12',
            '3.04', '2.96', '2.88', '2.8', '2.72',
            '2.56', '2.48', '2.4', '2.32', '2.24',
            '2.18', '2.11', '2.04', '1.96', '1.89',
            '1.8', '1.72', '1.71', '1.69', '1.63',
            '1.61', '1.55', '1.0']
    d_en = ['7.01', '6.93', '6.85', '6.77', '6.69',
            '6.61', '6.53', '6.45', '6.37', '6.29',
            '6.21', '6.13', '6.05', '5.97', '5.89',
            '5.81', '5.73', '5.65', '5.57', '5.49',
            '5.41', '5.33', '5.25', '5.17', '5.09',
            '5.01', '4.93', '4.85', '4.77', '4.69',
            '4.61', '4.53', '4.45', '4.37', '4.29',
            '4.21', '4.13', '4.05', '3.97', '3.81',
            '3.65', '3.17', '3.09', '3.01', '2.93',
            '2.85', '2.77', '2.61', '2.53', '2.45',
            '2.37', '2.29', '2.23', '2.16', '2.09',
            '2.01', '1.93', '1.85', '1.77', '1.69',
            '1.61', '1.53', '1.45', '1.29', '1.20',
            '1']
    d_bree = ['4.29', '4.25', '3.65', '3.35', '3.12',
              '3.04', '3.01', '2.96', '2.88', '2.80',
              '2.72', '2.56', '2.53', '2.48', '2.46',
              '2.40', '2.32', '2.29', '2.24', '2.23',
              '2.18', '2.16', '2.11', '2.04', '1.96',
              '1.89', '1.85', '1.80', '1.75', '1.72',
              '1.71', '1.69', '1.63', '1.61', '1.56',
              '1.55', '1.375']

    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        self.setupUi(self)
        self.tabWidgetCP.setCurrentIndex(0)
        self.bt_calc.clicked.connect(self.calc_cp)
        self.ck_eixo.clicked.connect(self.load_eixos)
        self.cb_Eixos.activated.connect(self.load_eixos)
        self.label_filme3.setHidden(True)
        self.label_CodFilm3.setHidden(True)
        self.ed_codFilm3.setHidden(True)
        self.label_paper.setHidden(True)
        self.Ed_MilsPaper.setHidden(True)
        self.label_47.setHidden(True)
        self.label_48.setHidden(True)
        self.lb_ctSec.setHidden(True)
        self.lb_cPapSec.setHidden(True)
        self.lb_cFilSec.setHidden(True)
        self.lb_cFoSec.setHidden(True)
        self.lb_ctUn.setHidden(True)
        self.lb_cPapUn.setHidden(True)
        self.lb_cFilUn.setHidden(True)
        self.lb_cFoUn.setHidden(True)
        self.label_37.setHidden(True)
        self.inputCodLateral.setHidden(True)
        self.labelDescr.setHidden(True)
        self.inputDescr.setHidden(True)
        self.label_40.setHidden(True)
        self.ed_codFilm1.setHidden(True)
        self.ed_codFilm2.setHidden(True)
        self.ed_codFilm3.setHidden(True)
        self.label_42.setHidden(True)
        self.label_CodFilm3.setHidden(True)
        self.cb_filme1.addItems(self.filmes)
        self.cb_filme2.addItems(self.filmes)
        self.cb_filme3.addItems(self.filmes)
        self.cb_filme3.setHidden(True)
        self.tabWidgetCP.setTabVisible(6,False)
        self.cb_filme1.setCurrentText('55')
        self.cb_filme2.setCurrentText('60')
        self.ed_codFilm1.setText('B03018AHD')
        self.ed_codFilm2.setText('B03018AID')
        self.ed_codFilm3.setText('B03018AAD')
        self.spinBoxNumFilmes.textChanged.connect(self.numfilmeschanged)
        self.cb_filme1.currentTextChanged.connect(self.numfilmeschanged)
        self.cb_filme2.currentTextChanged.connect(self.numfilmeschanged)
        self.cb_filme3.currentTextChanged.connect(self.numfilmeschanged)
        self.cb_modelo.currentTextChanged.connect(self.model_volt_coil_changed)
        self.ed_Vcoil.textChanged.connect(self.model_volt_coil_changed)
        self.placa_lateral = calc_cap.placa_lateral1(altura=0,largura=0)
        self.cb_PlacaLateral.addItems(self.placa_lateral['Descrição'])
        self.indice_mandril = 0
        self.cb_PlacaLateral.setCurrentText('330')
        self.inputCodLateral.setText('330')
        self.inputDescr.setPlainText('PLACA LATERAL 320X330X0.8MM  ')
        self.cb_PlacaLateral.currentTextChanged.connect(
            self.alt_change)
        self.dSPCalco.valueChanged.connect(self.alt_change)
        self.dSpinFolga.valueChanged.connect(self.alt_change)
        self.win_rsp = Resistor()
        self.win_fusein = Fusein()
        self.model_volt_coil_changed()
        self.listaTecnicapushButton.clicked.connect(self.lista_tecnica)
        # caixa = pd.read_excel(path_name+'\\db\\MH 400kvar (2).xlsx', sheet_name=8,header=1)
        # qc = """Select *  From caixa"""                   
        # caixa = con.execute(qc).df()
        # caixa['Descrição'] = caixa['P'].astype(str)+'x'+caixa['H'].astype(str)
        # self.comboBoxCaixas.addItems(caixa['Descrição'])
        self.comboBoxCaixas.currentTextChanged.connect(self.caixa_change)
        self.alt_change()
        self.cartao = {105:"4430521H03",115:"4430521H18",130:"4430521H01",140:"4430521H06",153:"4430521H06",182:"4430521H04",203:"4430521H08" } 
           

    @QtCore.Slot()
    def lista_tecnica(self):        
        self.df['Código Capacitor'] = [self.codSecCapacitivaLineEdit.text()]                 
        self.df['Altura da caixa'][0] = self.dSpinHBox.value()
        self.df['Largura da caixa'][0] = self.dSpinPBox.value()
        self.df['Corrente'][0] = self.df['29- Potência'][0]*1000/self.df['30- Tensão'][0]
        solda_simples = self.df['8- Esp. bobina (mm)'][0]*self.df['6- Total de bobinas'][0]*0.000843
        if self.df['Corrente'][0]<20:
            self.df['Caminhos de solda'] = [solda_simples]
        elif self.df['Corrente'][0]<40:
            self.df['Caminhos de solda'] = [solda_simples*2]
        elif self.df['13- Largura (mm)'][0]>=182:
            self.df['Caminhos de solda'] = [solda_simples*3]
        else:
            self.df['Caminhos de solda'] = [solda_simples*2]   
        # if self.df['Corrente'][0]<20:
        #     self.df['Caminhos de solda'] = [1]
        # elif self.df['Corrente'][0]<40:
        #     self.df['Caminhos de solda'] = [2]
        # elif self.df['13- Largura (mm)'][0]>=182:
        #     self.df['Caminhos de solda'] = [3]
        # if self.df['30- Tensão'][0]>14400:       
        #     self.df['Voltas de papel Bobina Envoltória'] = [17]
        # elif self.df['30- Tensão'][0]>9960:
        #     self.df['Voltas de papel Bobina Envoltória'] = [13]
        # elif self.df['30- Tensão'][0]>2400:
        #     self.df['Voltas de papel Bobina Envoltória'] = [9]
        # elif self.df['30- Tensão'][0]>600:
        #     self.df['Voltas de papel Bobina Envoltória'] = [7]
        # else:
        #     self.df['Voltas de papel Bobina Envoltória'] = [6]
        if self.df['NBI'][0]>=200:       
            self.df['Voltas de papel Bobina Envoltória'] = [20]
        elif self.df['NBI'][0]>=150:
            self.df['Voltas de papel Bobina Envoltória'] = [17]
        elif self.df['NBI'][0]>=110:
            if self.df['30- Tensão'][0]<=9960:
                self.df['Voltas de papel Bobina Envoltória'] = [9]
            else:
                self.df['Voltas de papel Bobina Envoltória'] = [13]
        elif  self.df['NBI'][0]>=15:
            if self.df['30- Tensão'][0]<=600:
                self.df['Voltas de papel Bobina Envoltória'] = [5]
            else:
                self.df['Voltas de papel Bobina Envoltória'] = [7] 


        largura = float(self.df['Largura da caixa'][0])
        altura = float(self.df['Altura da caixa'][0])
        nbi = self.df['NBI'][0]
        # caixa_pintura_df = calc_cap.caixa_pintura(altura=altura,largura=largura)
        # print(caixa_pintura_df)
        tampa_df = calc_cap.tampa(nbi=nbi,largura=largura,nbuchas=self.spN_Buchas.value())        
        self.df['Número de Buchas'] = tampa_df['n.Buchas']
        items = []
        codigos = []
        descricoes = []
        quantidades = []
        unidades = []
        dict = {'Item':items,'Código':codigos,'Descrição':descricoes,'Quantidade':quantidades,'Unidade':unidades}
        
        items.append('1.0')
        codigos.append(str(self.df['Código Capacitor']))
        descricoes.append('Capacitor')
        quantidades.append(1)
        unidades.append('pç')

        items.append('1.1')
        codigos.append('B14008')
        descricoes.append('SC Óleo')
        quantidades.append(altura*largura*343.0/1000000.0*0.3597*0.988)
        unidades.append('kg')        

        items.append('1.2')
        codigos.append(self.caixa['Pintura Cod. SC'][self.comboBoxCaixas.currentIndex()])
        descricoes.append('Pintura Cod. SC')
        quantidades.append(1)
        unidades.append('pç')

        items.append('2.0')
        codigos.append(2)
        descricoes.append('SG Grupo de montagem')
        quantidades.append(1)
        unidades.append('pç')       

        items.append('2.1')
        codigos.append(self.caixa['Pintura Cod. SC'][self.comboBoxCaixas.currentIndex()])
        descricoes.append('SC Caixa Capacitor')
        quantidades.append(1)
        unidades.append('pç')

        items.append('2.2')
        codigos.append(tampa_df['cod'][0])
        descricoes.append('SC Tampa Capacitor')
        quantidades.append(1)
        unidades.append('pç')

        resis = self.win_rsp.df.iloc[self.resitoresAprovadosComboBox.currentIndex()]
        if self.cb_Conexao.currentText()=='Série':
            res_df = calc_cap.resistores(resistor=str(resis['Resistor usado']),serie=resis['Número de resitores em série'],paralelo=1)
        else:
            res_df = calc_cap.resistores(resistor=str(resis['Resistor usado']),serie=resis['Número de resitores em série'],paralelo=resis['Número de resitores em paralelo'])    

        # if self.cb_Conexao.currentText()=='Série':
        #     res_df = calc_cap.resistores(resistor=str(resis['Resistor usado'])+'MO',serie=resis['Número de resitores em série'],paralelo=1)
        # else:
        #     res_df = calc_cap.resistores(resistor=str(resis['Resistor usado'])+'MO',serie=resis['Número de resitores em série'],paralelo=resis['Número de resitores em paralelo'])    

        items.append('2.3')
        if len(res_df)>0:
            codigos.append(res_df['Código'][0])
            descricoes.append('Conjunto de resistores de descarga '+self.resitoresAprovadosComboBox.currentText())
        else:
            codigos.append('Código nâo encontrado')
            descricoes.append('Conjunto de resistores de descarga '+self.resitoresAprovadosComboBox.currentText())    
        quantidades.append(1)
        unidades.append('pç')

        items.append('2.4')
        codigos.append(self.cartao[int(self.dSpinPBox.value())])
        descricoes.append('Cartão Duplo')
        quantidades.append(1)
        unidades.append('pç')

        items.append('2.5')
        codigos.append('B02023A')
        descricoes.append('Fio de Solda 60Sn x 40Pb')
        quantidades.append(0.01)
        unidades.append('kg')

        items.append('2.6')
        lcaixa = str(self.dSpinPBox.value())
        if self.df['Voltas de papel Bobina Envoltória'][0]<10:
            codigos.append('BE0'+str(self.df['Voltas de papel Bobina Envoltória'][0])+lcaixa)
        else:
            codigos.append('BE'+str(self.df['Voltas de papel Bobina Envoltória'][0])+lcaixa)    
        descricoes.append('Bobina Envoltória')
        quantidades.append(((self.df['13- Largura (mm)'][0]*1.5)+self.df['Altura da placa lateral'][0])/1000)
        unidades.append('kg')

        items.append('2.7')
        voltas_papel = self.df['Voltas de papel Bobina Envoltória'][0]
        if voltas_papel==9:
            codigos.append("E00172")
        else:
            codigos.append("Usar Placa Superior de Isolação") 
        descricoes.append('Cantoneira de Isolação ')
        quantidades.append(2)
        unidades.append('pç')

        items.append('2.8')
        lbox = self.dSpinPBox.value()
        cod = {105:'4430424H01',115:"4430424H04",130:"4430424H02",140:"4430424H8",153:"4430424H02",182:"4430424H10",203:"4430424H07"}
        if voltas_papel>=13:
            codigos.append(cod[lbox])
        else:
            codigos.append(False)    
        descricoes.append('Placa Superior de Isolação')
        quantidades.append(1)
        unidades.append('pç')

        items.append('2.9')
        codigos.append('SCSCNM')
        descricoes.append('SC Solda Cordoalha no Núcleo')
        quantidades.append(2)
        unidades.append('pç')

        items.append('2.10')
        codigos.append('TU70250')
        descricoes.append('Tubo de Papel')
        quantidades.append(2)
        unidades.append('pç')

        items.append('2.11')
        codigos.append('SCIC9')
        descricoes.append('SC Isolação da Cordoalha')
        pl = self.placa_lateral['Altura'][self.cb_PlacaLateral.currentIndex()]
        if self.df['32- Grupos série'][0]%2==0:
            quant = (pl**2+332**2)**.5+(self.dSpinPBox.value()-pl)+100
        else:
            quant = pl+(self.dSpinPBox.value()-pl)+100
        quantidades.append(quant)
        unidades.append('m')

        items.append('2.12')
        if self.df['Corrente'][0]>150:
            codigos.append('Verificar')
        elif self.df['Corrente'][0]>=116:
            codigos.append('B02061H')
        elif self.df['Corrente'][0]>80:
            codigos.append('B02061G') 
        elif self.df['Corrente'][0]>48:
            codigos.append('B02061D')
        else:
            codigos.append('B02061C')
        descricoes.append('Cordoalha')
        if self.df['32- Grupos série'][0]%2==0:
            quant = (pl**2+332**2)**.5+(self.dSpinPBox.value()-pl)+(self.dSpinPBox.value()+pl)+100
        else:
            quant = pl+(self.dSpinPBox.value()-pl)+100
        quantidades.append(quant)
        unidades.append('m')

        if self.cb_modelo.currentText()=='FI - 1 elo' or self.cb_modelo.currentText()=='FI - 2 elos':
            items.append('3.0')
            codigos.append('N'+str(self.df['Código Capacitor']))
            descricoes.append('SC Pacote com Caixa')
            quantidades.append(1)
            unidades.append('pç')
        else:
            items.append('3.0')
            codigos.append('PC')
            descricoes.append('SC Pacote com Caixa')
            quantidades.append(1)
            unidades.append('pç')

        calco_df = calc_cap.calco_lateral(largura=self.df['13- Largura (mm)'][0],espessura=3)
        
        items.append('3.1')
        codigos.append(calco_df['Código'][0])
        descricoes.append('Calço Lateral '+str(calco_df['Espessura (mm)'][0])+' x '+str(calco_df['Largura (mm)'][0]))
        quantidades.append(2)
        unidades.append('pç')
        placa_df = calc_cap.placa_lateral1(altura=self.df['Altura da placa lateral'][0],largura=self.df['13- Largura (mm)'][0])

        items.append('3.2')
        codigos.append(placa_df['Código'][0])
        descricoes.append('Placa lateral '+str(placa_df['Altura'][0])+' x '+str(placa_df['Largura'][0]))
        unidades.append('pç')
        if self.cb_modelo.currentText()=='FI - 1 elo' or self.cb_modelo.currentText()=='FI - 2 elos':            
            quantidades.append(1)
            items.append('3.2.1')
            hplaca = placa_df['Altura'][self.cb_PlacaLateral.currentIndex()]
            placa_lateral = calc_cap.placa_lateral1(altura=hplaca,largura=140)
            codigos.append(placa_lateral['Código'][0])
            quantidades.append(2)
            descricoes.append('Placa lateral '+str(placa_lateral['Altura'][0])+' x 140')
            unidades.append('pç')
               
        else:
            quantidades.append(2)

        divisor_df = calc_cap.divisor(largura=self.df['13- Largura (mm)'][0]+40)
        #divisor_dobra_20 = {105:'3430005H31',115:'-',130:'3430005H16'}
        if self.cb_modelo.currentText()=='FI - 1 elo':
            pass
        if self.cb_modelo.currentText()=='FI - 2 elos':
            pass
        else:
            items.append('3.3')
            codigos.append(divisor_df['Cod'][0])
            descricoes.append('Divisor moldado '+str(divisor_df['Largura'][0])+' mm')
            quantidades.append((self.df['32- Grupos série'][0]-1)*4)
            unidades.append('pç')

        items.append('3.4')
        codigos.append('B02021A')
        descricoes.append('Solda em Barra 62Snx38Zn')
        #=SE(B42<20;B10*B8*0,000843;SE(E(B42>=20;B42<=40);(B10*B8*0,000843)*2;SE(B15>160;(B10*B8*0,000843)*3;(B10*B8*0,000843)*2)))
        if self.df['20- Peso foil seção'][0]<20:
            qut = self.df['8- Esp. bobina (mm)'][0]*self.df['6- Total de bobinas'][0]*0.000843
        elif  self.df['20- Peso foil seção'][0]<40:
            qut = self.df['8- Esp. bobina (mm)'][0]*self.df['6- Total de bobinas'][0]*0.000843*2
        elif self.df['13- Largura (mm)'][0]>160:
            qut = self.df['8- Esp. bobina (mm)'][0]*self.df['6- Total de bobinas'][0]*0.000843*3
        else:
            qut = self.df['8- Esp. bobina (mm)'][0]*self.df['6- Total de bobinas'][0]*0.000843*2
        quantidades.append(qut)
        unidades.append('kg')

        items.append('3.5')
        # =SE(E(B42>40;B15<160);(B45-(B48*2))*0,5*12*8,96/1000000;"Não")
        if self.df['20- Peso foil seção'][0]>40 and self.df['13- Largura (mm)'][0]>160:
            qut = (self.df['Placa lateral'][0]-(self.df['Calço Lateral'][0]*2))*0.5*12*8.96/1000000
            codigos.append('B02057A06')
        else:
            qut = 0
            codigos.append('Não Utiliza')
        descricoes.append('Fita de Cobre 0,5 x 12mm')
        quantidades.append(qut)
        unidades.append('kg')

        items.append('4.0')        
        codigos.append('SC'+str(self.df['Código Capacitor'][0]))
        descricoes.append('Seção Capacitiva')
        quantidades.append(self.df['6- Total de bobinas'][0])
        unidades.append('pç')

        items.append('4.1')
        codigos.append('B02007D08')
        descricoes.append('Folha de Aluminio')
        quantidades.append(self.df['20- Peso foil seção'][0])
        unidades.append('kg')

        if self.df['# Filmes'][0]==2:
            items.append('4.2')
            # =SE(B37=24;"B03018AAD";SE(B37=40;"B03018AED";SE(B37=47;"B03018AFD";SE(B37=50;"B03018AGD";SE(B37=55;"B03018AHD";SE(B37=60;"B03018AID"))))))
            dict_cod ={24:"B03018AAD",40:"B03018AED",47:"B03018AFD",50:"B03018AGD",55:"B03018AHD",60:"B03018AID"}    
            codigos.append(dict_cod[self.df['Filme 1'][0]])
            descricoes.append('Filme Polipropileno '+str(self.df['Filme 1'][0]))
            # =(B21*B37)/(B37+B38)
            pfilm = self.df['19- Peso filme seção'][0]*self.df['Filme 1'][0]/(self.df['Filme 1'][0]+self.df['Filme 2'][0])
            quantidades.append(pfilm)
            unidades.append('kg')

            items.append('4.3')
            # =SE(B37=24;"B03018AAD";SE(B37=40;"B03018AED";SE(B37=47;"B03018AFD";SE(B37=50;"B03018AGD";SE(B37=55;"B03018AHD";SE(B37=60;"B03018AID"))))))
            codigos.append(dict_cod[self.df['Filme 2'][0]])
            descricoes.append('Filme Polipropileno '+str(self.df['Filme 2'][0]))
            # =(B21*B38)/(B38+B37)
            pfilm = self.df['19- Peso filme seção'][0]*self.df['Filme 2'][0]/(self.df['Filme 1'][0]+self.df['Filme 2'][0])
            quantidades.append(pfilm)
            unidades.append('kg')
        else:
            items.append('4.2')
            # =SE(B37=24;"B03018AAD";SE(B37=40;"B03018AED";SE(B37=47;"B03018AFD";SE(B37=50;"B03018AGD";SE(B37=55;"B03018AHD";SE(B37=60;"B03018AID"))))))
            dict_cod ={24:"B03018AAD",40:"B03018AED",47:"B03018AFD",50:"B03018AGD",55:"B03018AHD",60:"B03018AID"}    
            codigos.append(dict_cod[self.df['Filme 1'][0]])
            descricoes.append('Filme Polipropileno '+str(self.df['Filme 1'][0]))
            # =(B21*B37)/(B37+B38)
            pfilm = self.df['19- Peso filme seção'][0]*self.df['Filme 1'][0]/(self.df['Filme 1'][0]+self.df['Filme 2'][0]+self.df['Filme 3'][0])
            quantidades.append(pfilm)
            unidades.append('kg')

            items.append('4.3')
            # =SE(B37=24;"B03018AAD";SE(B37=40;"B03018AED";SE(B37=47;"B03018AFD";SE(B37=50;"B03018AGD";SE(B37=55;"B03018AHD";SE(B37=60;"B03018AID"))))))
            codigos.append(dict_cod[self.df['Filme 2'][0]])
            descricoes.append('Filme Polipropileno '+self.df['Filme 2'][0])
            # =(B21*B38)/(B38+B37)
            pfilm = self.df['19- Peso filme seção'][0]*self.df['Filme 2'][0]/(self.df['Filme 1'][0]+self.df['Filme 2'][0]+self.df['Filme 3'][0])
            quantidades.append(pfilm)
            unidades.append('kg')
            
            items.append('4.4')
            # =SE(B37=24;"B03018AAD";SE(B37=40;"B03018AED";SE(B37=47;"B03018AFD";SE(B37=50;"B03018AGD";SE(B37=55;"B03018AHD";SE(B37=60;"B03018AID"))))))
            codigos.append(dict_cod[self.df['Filme 3'][0]])
            descricoes.append('Filme Polipropileno '+self.df['Filme 3'][0])
            # =(B21*B38)/(B38+B37)
            pfilm = self.df['19- Peso filme seção'][0]*self.df['Filme 3'][0]/(self.df['Filme 1'][0]+self.df['Filme 2'][0]+self.df['Filme 3'][0])
            quantidades.append(pfilm)
            unidades.append('kg')

        
        lt_df = pd.DataFrame(dict)
        model_lt = DataFrameModel(lt_df)
        self.tableView_bom.setModel(model_lt)

        tipo = '\\resultados\\'
        path_name = os.getcwd()
        file_name = self.ed_file_name.text()
        path_name = path_name+tipo+file_name+'.xlsx'
        t = self.df.T
        with pd.ExcelWriter(path_name) as writer:
            t.to_excel(writer, sheet_name="Capacitor de potência")
            self.win_rsp.df.to_excel(writer, sheet_name="Resistor")            
            if self.cb_modelo.currentText()=='FI - 1 elo' or self.cb_modelo.currentText()=='FI - 2 elos':
                self.win_fusein.df1.to_excel(writer, sheet_name="Fusível")
            lt_df.to_excel(writer, sheet_name="Lista técnica")
        self.tabWidgetCP.setCurrentIndex(7)

    @QtCore.Slot()
    def caixa_change(self):
        self.dSpinHBox.setValue(self.caixa['H'][self.comboBoxCaixas.currentIndex()])
        self.dSpinPBox.setValue(self.caixa['P'][self.comboBoxCaixas.currentIndex()])
        pass
    @QtCore.Slot()
    def alt_change(self):
        try:
            placa_df = self.placa_lateral            
            hplaca = placa_df['Altura'][self.cb_PlacaLateral.currentIndex()]
            qc = """SELECT * FROM placa_df \
                WHERE Altura = """ + str(hplaca) + """  \
                """
            placa_df = con.execute(qc).df()
            self.inputCodLateral.setText(placa_df['Código'][0])
            self.inputDescr.setPlainText(placa_df['Descrição'][0])
            hbox = self.dSpinFolga.value()+hplaca
            lbox = self.df['13- Largura (mm)'][0] 
            caixa = pd.read_excel(path_name+'\\db\\MH 400kvar (2).xlsx', sheet_name=8,header=1)
            qc = """Select *  From caixa \
                        Where H>="""+str(hbox)+""" and P>="""+str(lbox)+"""
                        order by H"""
            caixa = con.execute(qc).df()            
            caixa['Descrição'] = caixa['P'].astype(str) +' x '+ caixa['H'].astype(str)            
            self.caixa = caixa
            self.comboBoxCaixas.clear()
            self.comboBoxCaixas.addItems(caixa['Descrição'])
        except:    
            pass

    @QtCore.Slot()
    def model_volt_coil_changed(self):
        modelo = self.cb_modelo.currentText()
        if modelo in ('MI', 'MH', 'MG', 'MF', 'ME', 'ST', 'VDC'):
            limite = self.alertas.get(modelo)
            self.label_vCoil.setText('Volts/Bobina ≤'+str(limite[0])+'V')
            self.aproved(self.ed_Vcoil, limite[0])            
        if modelo in ('FI - 1 elo','FI - 2 elos'):
            self.tabWidgetCP.setTabVisible(6,True)
            self.fusVeisAprovadosLabel.setHidden(False)
            self.fusVeisAprovadosComboBox.setHidden(False)
        else:
            self.tabWidgetCP.setTabVisible(6,False)
            self.fusVeisAprovadosLabel.setHidden(True)
            self.fusVeisAprovadosComboBox.setHidden(True)

    def aproved(self, spinbox, limite):
        if spinbox.value() > limite:
            spinbox.setStyleSheet("QDoubleSpinBox"
                                  "{"
                                  "background : red;"
                                  "}")
            return False
        else:
            spinbox.setStyleSheet("QDoubleSpinBox"
                                  "{"
                                  "background : lightgreen;"
                                  "}")
            return True

    @QtCore.Slot()
    def numfilmeschanged(self):
        if self.spinBoxNumFilmes.value() == 2:
            self.cb_filme3.setHidden(True)
            self.label_filme3.setHidden(True)
            self.label_CodFilm3.setHidden(True)
            self.ed_codFilm3.setHidden(True)
            espfilm = int(self.cb_filme1.currentText()) + \
                int(self.cb_filme2.currentText())
        else:
            self.cb_filme3.setHidden(False)
            self.label_filme3.setHidden(False)
            # self.label_CodFilm3.setHidden(False)
            # self.ed_codFilm3.setHidden(False)
            espfilm = int(self.cb_filme1.currentText(
            ))+int(self.cb_filme2.currentText())+int(self.cb_filme3.currentText())
        self.ed_milsFilm.setText(str(espfilm/100))
        self.ed_codFilm1.setText(
            self.codigos_filmes[self.cb_filme1.currentIndex()])
        self.ed_codFilm2.setText(
            self.codigos_filmes[self.cb_filme2.currentIndex()])
        self.ed_codFilm3.setText(
            self.codigos_filmes[self.cb_filme3.currentIndex()])

    @QtCore.Slot()
    def change_br(self):
        eixo = self.cb_Eixos.currentIndex() == 0
        self.cb_mandris.clear()
        if self.ck_eixo.isChecked():
            if eixo:
                self.cb_mandris.addItems(self.d_en)
            else:
                self.cb_mandris.addItems(self.d_br)
        self.cb_mandris.setCurrentIndex(self.indice_mandril)

    @QtCore.Slot()
    def load_eixos(self):
        eixo = self.cb_Eixos.currentIndex() == 0
        self.cb_mandris.clear()
        if self.ck_eixo.isChecked():
            if eixo:
                self.cb_mandris.addItems(self.d_bree)
            else:
                self.cb_mandris.addItems(self.d_bree)
        self.cb_mandris.setCurrentIndex(self.indice_mandril)

    @QtCore.Slot()
    def calc_cp(self):
        self.listaTecnicapushButton.setStyleSheet('background-color: rgb(255, 0, 0)')
        self.resitoresAprovadosComboBox.clear()
        self.fusVeisAprovadosComboBox.clear()
        modelo = self.cb_modelo.currentText()
        if modelo=='FI - 1 elo':
            self.win_fusein.spinBoxNumElos.setValue(1)
            self.init_fusein()
        elif modelo=='FI - 2 elos':
            self.win_fusein.spinBoxNumElos.setValue(2)
            self.init_fusein()

        self.tabWidgetCP.setCurrentIndex(2)
        eixo = self.cb_Eixos.currentIndex() == 0
        papelBk, filmeDk, oleoDK, tensao = float(self.ed_PapelBK.text()), float(
            self.ed_FilmeDK.text()), float(self.ed_OleoDK.text()), float(self.ed_Vn.text())
        fr, kvar, n_coils, v_coil = float(self.ed_Fr.text()), float(
            self.ed_KVAR.text()), int(self.Ed_Ncoil.text()), self.ed_Vcoil.value()
        mils_paper, mils_film, gauge_al = float(self.Ed_MilsPaper.text()), float(
            self.ed_milsFilm.text()), float(self.ed_gauge_Al.text())
        h_coil, lamelas_foil, w_foil = float(self.ed_Hcoil.text()), int(
            self.ed_lamelas_foil.text()), float(self.ed_larg_foil.text())
        tol_c, fator_espc, sobre_voltas = float(self.ed_tol.text()), float(
            self.ed_FatorEsp.text()), float(self.ed_sobre_voltas.text())
        fator_coil, rho_papel, custo_papel = float(self.ed_fator_coil.text()), float(
            self.ed_rhoPaper.text()), float(self.ed_custoPaper.text())
        custo_filme, custo_foil, larg_foil = float(self.ed_custoFilm.text()), float(
            self.ed_custoFoil.text()), float(self.ed_largFoil.text())
        larg_diel, refugo = float(
            self.ed_largDie.text()), float(self.ed_Refugo.text())
        file_name = self.ed_file_name.text()
        if self.ck_eixo.isChecked():
            indice_eixo = self.cb_mandris.currentIndex()
        else:
            indice_eixo = 0
        auto_eixo = not (self.ck_eixo.isChecked())
        while 1:
            if eixo:
                self.df = calc_cap.cp_en(papelBk, filmeDk, oleoDK, tensao, fr, kvar, n_coils, v_coil, mils_paper, mils_film, gauge_al, h_coil, lamelas_foil,
                                    w_foil, tol_c, fator_espc, sobre_voltas, fator_coil, rho_papel, custo_papel, custo_filme, custo_foil, larg_foil, larg_diel, refugo,
                                    indice_eixo, auto_eixo)
            else:
                self.df = calc_cap.cp_br(papelBk, filmeDk, oleoDK, tensao, fr, kvar, n_coils, v_coil, mils_paper, mils_film, gauge_al, h_coil, lamelas_foil,
                                    w_foil, tol_c, fator_espc, sobre_voltas, fator_coil, rho_papel, custo_papel, custo_filme, custo_foil, larg_foil, larg_diel, refugo,
                                    indice_eixo, auto_eixo)
            if  not self.checkBoxAutoNCoils.isChecked():
                break
            if self.df['8- Esp. bobina (mm)'][0]>=26:
                n_coils = self.df['32- Grupos série'][0]*(self.df['33- Grupos paralelo'][0]+1)
            else:
                break        
        h_placa = n_coils*self.df['8- Esp. bobina (mm)'][0]+2*self.dSPCalco.value()
        self.df['6- Total de bobinas'][0] = self.df['32- Grupos série'][0]*self.df['33- Grupos paralelo'][0]
        self.cb_PlacaLateral.clear()
        self.placa_lateral = calc_cap.placa_lateral1(altura=h_placa,largura=self.df['13- Largura (mm)'][0])
        self.cb_PlacaLateral.addItems(self.placa_lateral['Descrição'])
        self.cb_PlacaLateral.setCurrentIndex(0)
        self.dSpinHBox.setValue(self.placa_lateral['Altura'][0]+self.dSpinFolga.value())
        # self.dSpinPBox.setValue()
        self.df['Modelo'] = [self.cb_modelo.currentText()]
        self.df['NBI'] = [float(self.cb_NBI.currentText())]
        self.df['# Filmes'] = [self.spinBoxNumFilmes.value()]
        self.df['Filme 1'] = [int(self.cb_filme1.currentText())]
        self.df['Filme 2'] = [int(self.cb_filme2.currentText())]
        self.df['Codigo filme 1'] = self.ed_codFilm1.text()
        self.df['Codigo filme 2'] = self.ed_codFilm2.text()
        if self.spinBoxNumFilmes.value() == 3:
            self.df['Filme 3'] = [int(self.cb_filme3.currentText())]
            self.df['Codigo filme 3'] = self.ed_codFilm3.text()           
        self.df['Altura da placa lateral'] = self.placa_lateral['Altura'][0]
        self.df['Codigo da placa lateral'] = self.inputCodLateral.text()
        self.df['Descrição da placa lateral'] = self.inputDescr.toPlainText()
        self.df['Altura dos calços'] = self.dSPCalco.value()
        self.df['Folga da caixa'] = self.dSpinFolga.value()
        self.df['Altura da caixa'] = self.dSpinHBox.value()
        self.df['Largura da caixa'] = self.dSpinPBox.value()
        self.df['Corrente'] = self.df['29- Potência']*1000/self.df['30- Tensão'] 
        self.df['Gauge Al'] = [gauge_al] 
        t = self.df.T        
        model = DataFrameModel(t.reset_index())
        self.tableView.setModel(model)
        dict = t[0]
        self.indice_mandril = self.df['28- Índice mandril'][0]
        self.lb_95PF.setValue(dict[15])
        self.lb_cCoil.setValue(dict[10])
        # self.lb_cFilSec.setText(str(dict[28]))
        # self.lb_cFilUn.setText(str(dict[32]))
        # self.lb_cFoSec.setText(str(dict[29]))
        # self.lb_cFilSec.setText(str(dict[28]))
        # self.lb_ctUn.setText(str(dict[30]))
        # self.lb_ctSec.setText(str(dict[26]))
        self.lb_epBob.setValue(dict[8])
        self.lb_fat_Bob.setValue(dict[3])
        self.lb_fatorEsp.setValue(dict[0])
        self.lb_larg.setValue(dict[13])
        # self.lb_cFoUn.setText(str(dict[33]))
        # self.lb_cPapSec.setText(str(dict[27]))
        self.lb_cUn.setValue(dict[9])
        self.lb_pFilSec.setText(str(dict[19]))
        self.lb_pFilUn.setText(str(dict[23]))
        self.lb_pFoil.setValue(dict[16])
        self.lb_pfPrin.setValue(dict[14])
        self.lb_pTlUn.setText(str(dict[21]))
        self.lb_sobreVoltas.setValue(dict[5])
        self.lb_tCoils.setValue(dict[6])
        self.lb_pTSec.setText(str(dict[17]))
        self.lb_vCoil.setValue(dict[12])
        self.spin_stressVPM.setValue(dict[34])
        self.spin_Stress.setValue(dict[35])
        modelo = self.cb_modelo.currentText()
        if modelo in ('MI', 'MH', 'MG', 'MF', 'ME', 'ST', 'VDC'):
            limite = self.alertas.get(modelo)
            self.aproved(self.lb_vCoil, limite[0])
            self.aproved(self.spin_stressVPM, limite[1])
            self.aproved(self.spin_Stress, limite[1]/25.4)
        self.lb_voltasFoil.setValue(dict[4])
        self.lb_pFoUn.setText(str(dict[24]))
        self.lb_pPapSec.setText(str(dict[22]))
        self.lb_vCoil.setValue(dict[12])
        self.lb_compFoil.setValue(dict[2])
        self.lbdsMan.setValue(dict[11])
        self.lb_pPapUn.setText(str(dict[22]))
        self.lb_pFoSec.setText(str(dict[20]))
        self.spinBoxGS.setValue(dict[32])
        self.spinBoxGP.setValue(dict[33])
        self.spinBoxGSGP.setValue(dict[6])
        self.Ed_Ncoil.setText(str(dict[6]))
        self.win_rsp.cb_Conex.setCurrentText(self.cb_Conexao.currentText())
        self.win_rsp.ed_Vn.setText(self.ed_Vn.text())
        self.win_rsp.ed_Fr.setText(self.ed_Fr.text())
        self.win_rsp.ed_KVAR.setText(self.ed_KVAR.text())
        self.win_rsp.ed_max_W_R.setText(self.ed_Watts.text())
        self.win_rsp.ed_tempo.setText(self.ed_Tempo.text())
        self.win_rsp.ed_Vn.setText(self.ed_Vn.text())
        self.win_rsp.calc_resistor()
        if self.cb_Conexao.currentText()=='Série':            
            self.win_rsp.df['Formação'] = self.win_rsp.df['Número de resitores em série'].astype('str')+' x '+ self.win_rsp.df['Resistor usado'].astype('str')+ 'MΩ'
        else:
            self.win_rsp.df['Formação'] = self.win_rsp.df['Número de resitores em série'].astype('str')+' x '+self.win_rsp.df['Número de resitores em paralelo'].astype('str')+' x '+ self.win_rsp.df['Resistor usado'].astype('str')+ 'MΩ'    
        self.resitoresAprovadosComboBox.addItems(self.win_rsp.df['Formação'])
        self.df['Conexão dos resistores'] = self.win_rsp.df['Formação'][0]         
        model_res = DataFrameModel(self.win_rsp.df)
        self.tableView_Resistor.setModel(model_res)
        if modelo=='FI - 1 elo' or modelo=='FI - 2 elos':
            self.fusVeisAprovadosComboBox.addItems(self.win_fusein.df1['Diam'].astype('str'))
            model_fusein = DataFrameModel(self.win_fusein.df1)
            self.tableView_fusein.setModel(model_fusein)
        tipo = '\\resultados\\'
        path_name = os.getcwd()
        path_name = path_name+tipo+file_name+'.xlsx'
        with pd.ExcelWriter(path_name) as writer:
            t.to_excel(writer, sheet_name="Capacitor de potência")
            self.win_rsp.df.to_excel(writer, sheet_name="Resistor")
            if modelo=='FI - 1 elo' or modelo=='FI - 2 elos':
                self.win_fusein.df1.to_excel(writer, sheet_name="Fusível")
        self.listaTecnicapushButton.setStyleSheet('background-color: rgb(0, 255, 0)')
          

    def init_fusein(self):
        self.win_fusein.spinBoxNumElos.setReadOnly(True)
        self.win_fusein.ed_Vn.setText(self.ed_Vn.text())
        self.win_fusein.ed_Vn.setReadOnly(True)
        self.win_fusein.ed_Fr.setText(self.ed_Fr.text())
        self.win_fusein.ed_Fr.setReadOnly(True)
        self.win_fusein.ed_KVAR.setText(self.ed_KVAR.text())
        self.win_fusein.ed_KVAR.setReadOnly(True)
        self.win_fusein.tabWidget.setCurrentIndex(0)
        self.win_fusein.exec_()
        self.Ed_Ncoil.setText(str(int(self.win_fusein.ed_G_SI.text())*int(self.win_fusein.ed_GPI.text())))
        self.spinBoxGS.setValue(int(self.win_fusein.ed_G_SI.text())) 
        self.spinBoxGP.setValue(int(self.win_fusein.ed_GPI.text()))  
        self.spinBoxGSGP.setValue(int(self.Ed_Ncoil.text()))     
        # self.lb_cPapUn.setText(str(dict[31]))

class Obs(QDialog, Ui_DialogObs):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI 
        self.setupUi(self)
    # def closeEvent(self, event):
    #     return self.plainTextEdit.toPlainText()

class Open_Hib(QDialog, Ui_DialogHibrido):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI 
        self.setupUi(self)

class Force_Side_by_Side(QDialog, Ui_DialogForceIcc):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI 
        self.setupUi(self)

class CrudFibra(QDialog, Ui_DialogCrudFibra):
    def __init__(self, dados, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI
        self.setupUi(self)


class SysData(QDialog, Ui_DialogSys):
    def __init__(self, dados, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI
        self.setupUi(self)
        self.sys_data = dados


class ConstrData(QDialog, Ui_DialogConstr):
    def __init__(self, dados, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI
        self.setupUi(self)
        self.const_data = dados


class HarmData(QDialog, Ui_DialogHarm):
    def __init__(self, dados, parent=None):
        super().__init__(parent)
        # Create an instance of the GUI
        # Run the .setupUi() method to show the GUI
        self.setupUi(self)
        self.harm_data = dados


class Signals(QtCore.QObject):
    signal_sys = QtCore.Signal(pd.DataFrame, name='signal_sys')
    signal_constr = QtCore.Signal(pd.DataFrame, name='signal_constr')


class MainReactor(QMainWindow, Ui_BreeReactorMainWindow):
    def __init__(self, *args, obj=None, **kwargs):
        super(MainReactor, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.actionCarregar_projeto.triggered.connect(self.openFileNameDialog)
        self.actionSalvar_projeto.triggered.connect(self.saveFileDialog)
        self.sys_data = pd.DataFrame
        self.constr_dados = pd.DataFrame
        self.cilindros_dados = pd.DataFrame
        self.camadas_dados = pd.DataFrame
        self.resultado_dados = pd.DataFrame
        self.harm_dados = pd.DataFrame
        self.dados_projeto = pd.DataFrame
        self.actionDados_do_sistema.triggered.connect(self.system_data)
        self.actionCorrentes_harm_nicas.triggered.connect(self.harmonicos)
        self.actionDados_construtivos.triggered.connect(self.system_contr)
        self.actionEspessuras_das_fibras.triggered.connect(self.crud_fibra)
        self.actionFor_as_de_curto_e_vento.triggered.connect(self.force_side_by_side)
        self.win_crud_fibra = CrudFibra(self)
        self.signals = Signals()
        self.signals.signal_sys.connect(self.receive_signal_sys)
        self.win_system = SysData(self.sys_data)
        self.win_system.btn_Ok.clicked.connect(self.sys_data_ok)
        self.win_system.inputAlt.textChanged.connect(self.alt_change)
        self.win_system.normaComboBox.currentTextChanged.connect(self.alt_change)
        self.win_system.inputICCSys.textChanged.connect(self.icc_sys_change)
        self.win_system.inputICC.textChanged.connect(self.icc_change)
        self.win_system.inputLn.textChanged.connect(self.Ln_change)
        self.win_system.inputInominal.textChanged.connect(self.iN_change)
        self.win_system.inputNBI.textChanged.connect(self.NBI_change)
        self.win_constr = ConstrData(self.constr_dados)
        self.set_win_constr()
        self.win_constr.comboBoxCruzInf.currentIndexChanged.connect(self.cruz_inf_change)
        self.win_constr.btn_Ok.clicked.connect(self.constr_ok)
        self.win_constr.btn_calc.clicked.connect(self.calcular)
        self.win_constr.bt_inicial.clicked.connect(self.iniciar)
        self.win_constr.comboBoxFibra.currentIndexChanged.connect(self.fibra_change)
        self.win_constr.comboBox_TempIcc.currentIndexChanged.connect(self.norma_temp_change)
        self.win_constr.comboBox_IccMax.currentIndexChanged.connect(self.norma_icc_change)
        self.actionNovo_projeto.triggered.connect(self.novo_projeto)
        self.win_harm = HarmData(self.harm_dados)
        self.win_harm.bt_Fill.clicked.connect(self.preencher_harm)
        self.win_system.inputLtarget.textChanged.connect(self.bt_calc_enable)
        self.win_system.inputPico.textChanged.connect(self.bt_calc_enable)
        self.win_system.inputITotal.textChanged.connect(self.bt_calc_enable)
        self.win_constr.bt_Del_Cil.clicked.connect(self.bt_del_cil)
        self.win_constr.bt_ins_cil.clicked.connect(self.bt_ins_cil)
        self.win_constr.btnFreeCad.clicked.connect(self.bt_FreeCad)
        self.win_constr.pbCriar.clicked.connect(self.bt_criar)
        self.win_constr.pbHibrido.clicked.connect(self.bt_hibrido)
        self.win_constr.pbObs.clicked.connect(self.bt_obs)
        self.win_constr.pushButtonCalc.clicked.connect(self.asCalcClicked)
        self.win_constr.pushButtonAsBuilt.clicked.connect(self.asBuiltClicked)
        self.force_lado_a_lado = Force_Side_by_Side()
        self.obs = Obs()
        #self.obs.closeEvent.connect(self.quit_obs)
        self.hibrido = Open_Hib()
        self.hibrido.pbCalc.clicked.connect(self.calcular_hibrido)
        self.hibrido.bt_AddTap.clicked.connect(self.add_tap)
        self.hibrido.bt_DelTap.clicked.connect(self.del_Tap)
        self.hibrido.bt_CAD.clicked.connect(self.bt_FreeCadTap)
        self.win_constr.comboBoxMontagem.currentIndexChanged.connect(self.montagem_change)
        self.set_win_force_side()
        self.tem_tap = False

    @QtCore.Slot()
    def montagem_change(self):
        pass 
    
    @QtCore.Slot()
    def asCalcClicked(self):
        self.win_constr.pushButtonCalc.setStyleSheet('background-color: rgb(255, 0, 0)')        
        cilindro, camada, disol, ax, rd, nesp, he, circ, he_b,circ_b, err_he, err_circ = [], [],[], [], [], [], [], [], [], [],[],[]
        cam = 1
        for index, cil in self.cilindros_dados.iterrows():
            frd = int(cil['Fios radiais'])
            for layer in range(frd):
                cilindro.append(index+1)
                camada.append(cam)
                disol.append(cil['dfio isol avg (mm)'])
                ax.append(cil['Fios axiais'])
                rd.append(cil['Fios radiais'])
                nesp.append(self.camadas_dados['# Espiras'][cam-1])
                he.append(self.camadas_dados['Altura do enrolamento'][cam-1])
                he_b.append(self.camadas_dados['Altura do enrolamento'][cam-1])
                circ.append(self.camadas_dados['Circunferência externa'][cam-1])
                circ_b.append(self.camadas_dados['Circunferência externa'][cam-1])
                #dm.append('=H'+str(cam+1)+'/D'+str(cam+1))
                #dm_b.append('=H'+str(cam+1)+'/D'+str(cam+1))
                err_he.append('=(1-H'+str(cam+1)+'/I'+str(cam+1)+')*100')
                err_circ.append('=(1-J'+str(cam+1)+'/K'+str(cam+1)+')*100')
                cam += 1
        dict = {'Cil':cilindro,'Cam':camada,'Disol(mm)':disol,'Axiais':ax,'Radiais':rd,'Nesp':nesp,'Hcalc(mm)':he,'Hmedido(mm)':he_b,'π.DextCalc':circ,'π.DextMed':circ_b,'Erro H %':err_he, 'Erro D %':err_circ }
        df_eng_rev = pd.DataFrame(dict)
        # df_eng_rev.to_csv(reactor.path_name+'\\Engenharia Reversa\\modelo.csv')
        df_eng_rev.to_excel(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx',engine="openpyxl")
        # Carregar dados para variável
        wb = load_workbook(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx')
        # Escolhe active sheet
        ws = wb.active
        col = ws.column_dimensions['I']
        col.font = Font(bold=True)
        col.fill = PatternFill("solid", fgColor="00FFFF00")
        col = ws.column_dimensions['K']
        col.font = Font(bold=True)
        col.fill = PatternFill("solid", fgColor="00FFFF00")
       
        wb.save(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx')
        self.win_constr.pushButtonCalc.setStyleSheet('background-color: rgb(0, 255, 0)')  


    
    
    @QtCore.Slot()
    def asBuiltClicked(self):
        self.win_constr.pushButtonAsBuilt.setStyleSheet('background-color: rgb(255, 0, 0)')  
        #xls = pd.ExcelFile(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx') 
        #df_eng_rev = xls.parse('Sheet1')
        xl_model = formulas.ExcelModel().loads(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx').finish()
        xl_model.calculate()
        os.remove(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx')
        xl_model.write(dirpath=reactor.path_name+'\\Engenharia Reversa') 
            
        # rev = load_workbook(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx', data_only=True)
        # rev.save(reactor.path_name+'\\Engenharia Reversa\\modelo1.xlsx')
        df_eng_rev = pd.read_excel(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx',keep_default_na=True,
                    usecols=['Disol(mm)','Nesp','Hmedido(mm)','π.DextMed'])#,dtype={'Disol(mm)':'float','Nesp':'float','He(mm)':'float','π.Dext':'float','Dm(mm)':'float'})
        # df_eng_rev = df_from_excel(reactor.path_name+'\\Engenharia Reversa\\modelo.xlsx')
        # df_eng_rev = pd.read_csv(reactor.path_name+'\\Engenharia Reversa\\modelo.csv',
        #              usecols=['Disol(mm)','Nesp','He(mm)','π.Dext','Dm(mm)'])
        self.camadas_dados['# Espiras'] = df_eng_rev['Nesp']
        self.camadas_dados['Altura do enrolamento'] = df_eng_rev['Hmedido(mm)']
        #self.camadas_dados['Diâmetro médio'] = df_eng_rev['π.Dext']/reactor.np.pi-df_eng_rev['Disol(mm)']#df_eng_rev['Dm(mm)']
        self.camadas_dados['Circunferência externa'] = df_eng_rev['π.DextMed']
        self.camadas_dados['Diâmetro médio'] = df_eng_rev['π.DextMed']-df_eng_rev['Disol(mm)']
        self.win_constr.checkBoxAutoTurn.setChecked(False)
        self.win_constr.checkBoxHe.setChecked(False)
        self.win_constr.pushButtonAsBuilt.setStyleSheet('background-color: rgb(0, 255, 0)') 
        
    
    @QtCore.Slot()
    def del_Tap(self):
        """Delete rows with currently selected cells and/or selected rows of the model"""
        #index = self.win_constr.tableViewCil.currentIndex()
        #self.model_cil.removeRows(index.row(), 1, index)
        indices = self.hibrido.tv_Taps.selectionModel().selectedRows()
        for index in sorted(indices, reverse=True):
            self.model_hibrido.removeRows(index.row(), 1, index)
        # for index, tap in self.model_cil._dataframe.iterrows():
        #     tap['tap'] = index        

    @QtCore.Slot()
    def add_tap(self):
        dbs_mod_tap,localiz, r_esp, rdc, peso_esp,he,hesp,zc,nesp, dm = self.extrair_main_data()
        tap = self.hibrido.inputTap.value()
        ind = self.hibrido.inputLn.value()*(1+tap/100)
        index = self.hibrido.tv_Taps.currentIndex()
        if index.row()==-1:
            dict = {'Tap': [tap],'Lalvo(mH)':[ind],'L1 (mH)':[self.hibrido.inputL1.value()],'Lcalc(mH)':[0.0],'Erro(%)':[100.0],'Ltap(mH)':[0.0],'M(mH)':[0.0], 'Z(mm)': zc, '# Espiras': nesp, 'Altura da espira': hesp, 
            'Altura do enrolamento': he, 'Diâmetro médio': dm, 'Rdc @ 20°C (mΩ)': rdc, 'Rdc Espira': r_esp,
            }
            self.open_style = pd.DataFrame(dict)
            self.model_hibrido = DataFrameModel(self.open_style)
            self.hibrido.tv_Taps.setModel(self.model_hibrido)
            index = self.hibrido.tv_Taps.currentIndex()
        else:            
            indices = self.hibrido.tv_Taps.selectionModel().selectedRows()   
            for index in sorted(indices, reverse=False):
                self.model_hibrido.insertRows(index.row(), 1, index, None)
                self.model_hibrido._dataframe['Tap'][index.row()] = tap
                self.model_hibrido._dataframe['Lalvo(mH)'][index.row()] = ind    
        self.model_hibrido._dataframe.sort_values(by='Tap', inplace=True)
        self.open_style = self.model_hibrido._dataframe
        #self.hibrido.tv_Taps.setModel(self.model_hibrido)
        #index = self.hibrido.tv_Taps.currentIndex()
        pass

    def dados_tap(self):
        dict = {'L1 (mH)':[self.hibrido.inputL1.value()],
            'Calço':[self.hibrido.inputCalco.value()],
            'Largura Perfil (mm)':[self.hibrido.inputLargPerf.value()],
            'Altura do Perfil (mm':[self.hibrido.inputEspPerfil.value()],
            '# Entradas':[self.hibrido.inputNumPerf.value()],
            'Material do perfil':[self.hibrido.cbMaterialTap.currentText()],
            'Instalação':[self.hibrido.cbPosicao.currentText()],
            'DBST':[self.hibrido.inputDBSb.value()],
            'Diâmetro interno (mm)':[self.hibrido.inputDi.value()],
            'Altura do enrolamento (mm)':[self.hibrido.inputHe.value()],
            'Número de espiras':[self.hibrido.inputNesp.value()]}
        self.dados_taps = pd.DataFrame(dict) 
        index = self.hibrido.tv_Taps.currentIndex()
        if index.row()==-1:
            dict = {'Tap': [],'Lalvo(mH)':[],'L1 (mH)':[],'Lcalc(mH)':[],'Erro(%)':[],'Ltap(mH)':[],'M(mH)':[], 'Z(mm)': [], '# Espiras': [], 'Altura da espira': [], 
            'Altura do enrolamento': [], 'Diâmetro médio': [], 'Rdc @ 20°C (mΩ)': [], 'Rdc Espira': [],
            }   

    @QtCore.Slot()
    def calcular_hibrido(self):
        self.hibrido.pbCalc.setStyleSheet('background-color: rgb(255, 0, 0)') 

        #dbs_mod_tap,localiz, r_esp, rdc, peso_esp,he,hesp,zc,nesp, dm = self.extrair_main_data()
        dbs_mod_tap =  self.hibrido.inputDBSb.value()
        localiz = self.hibrido.cbPosicao.currentText()
        try:
            M1 = reactor.np.copy(self.M)
            R1 = reactor.np.copy(self.R)
            if len(self.camadas_dados)!=len(self.R):
                self.calcular()
                M1 = reactor.np.copy(self.M)
                R1 = reactor.np.copy(self.R)
        except:
            self.calcular()
            M1 = reactor.np.copy(self.M)
            R1 = reactor.np.copy(self.R)    
        freq = self.win_system.inputFreq.value() 
        harms = self.harm_dados
        #harms['RMS (A)'] = harms['RMS (A)'].astype(float)
        harms.astype({'RMS (A)': 'float'}).dtypes 
        ind_calc,ltap,m,nesp, altura = reactor.equ_hibr(R=R1,M=M1,l1=self.win_constr.inputLcalc.value(),camadas=self.camadas_dados,
                        taps=self.open_style,dbs=self.win_constr.inputDBS.value(),dbs_mod_tap=dbs_mod_tap,altura1=self.win_constr.inputHs.value(),localiz=localiz,sinal=self.hibrido.inputSinal.value(),harms=harms,equilibrar=self.hibrido.checkBoxAutoTurn.isChecked())                             
        hotspot, tmed, dtmax = self.temperatura(self.win_constr.inputDBS.value(), 
                                self.win_constr.inputLargEsp.value(),harms['RMS (A)'][0])
        self.hibrido.inputLcalc.setValue(ind_calc)
        self.hibrido.inputNesp.setValue(nesp)
        self.hibrido.inputMutua.setValue(m)
        self.hibrido.inputLtap.setValue(ltap)
        self.hibrido.inputHe.setValue(altura)
        self.hibrido.ck_Tap.setChecked(True)
        self.model_hibrido = DataFrameModel(self.open_style)
        self.hibrido.tv_Taps.setModel(self.model_hibrido)
        self.model_cil = DataFrameModel(self.cilindros)
        self.hibrido.tv_Cil.setModel(self.model_cil)
        model_cam = DataFrameModel(self.camadas_dados)
        self.hibrido.tv_Camadas.setModel(model_cam)
        self.tem_tap = True
        self.hibrido.pbCalc.setStyleSheet('background-color: rgb(0, 255, 0)') 

    def extrair_main_data(self):
        rho20, dens = self.material_condutor(self.hibrido.cbMaterialTap.currentText())
        nesp = [self.hibrido.inputNesp.value()]
        fax = self.hibrido.inputNumPerf.value()
        # hesp = [fax*(self.hibrido.inputEspPerfil.value()+
        #         float(self.hibrido.cbAltCalco.currentText()))]
        hesp = [fax*(self.hibrido.inputEspPerfil.value()+
                 float(self.hibrido.inputCalco.value()))]
        he = [nesp[0]*hesp[0]]
        dbs_mod_tap =  self.hibrido.inputDBSb.value()
        localiz = self.hibrido.cbPosicao.currentText()       
        if localiz == 'Acima da cruzeta superior':
            zc = [self.win_constr.inputHs.value()/2+he[0]/2+dbs_mod_tap]
        elif localiz == 'Satetlite por baixo da cruzeta superior':
            zc = [self.win_constr.inputDBS.value()/2-he[0]/2-dbs_mod_tap]
        elif localiz == 'Satétile acima da cruzeta inferior':
            zc = [-self.win_constr.inputDBS.value()/2+he[0]/2+dbs_mod_tap]
        elif localiz == 'Abaixo da cruzeta inferior':
            zc = [-self.win_constr.inputHs.value()/2-he[0]/2-dbs_mod_tap]
        dm = [self.hibrido.inputDi.value()+self.hibrido.inputLargPerf.value()]
        i_alvo = [self.win_system.inputITotal.value()]
        circunferencia_externa = [dm[0]*reactor.np.pi]
        r_esp = [rho20*reactor.np.pi*dm[0]/(fax*self.hibrido.inputEspPerfil.value()*
                self.hibrido.inputLargPerf.value())]
        rdc = [r_esp[0]*nesp[0]]
        peso_esp = [dens*reactor.np.pi*dm[0]*fax*self.hibrido.inputEspPerfil.value()*self.hibrido.inputLargPerf.value()]
        return dbs_mod_tap,localiz, r_esp, rdc, peso_esp,he,hesp,zc,nesp,dm
        
    @QtCore.Slot()
    def bt_obs(self):         
        self.obs.exec_()
        if len(self.obs.plainTextEdit.toPlainText())!=0:
            self.win_constr.pbObs.setStyleSheet('background-color: rgb(255, 0, 0)')
            observs = [self.obs.plainTextEdit.toPlainText()]
            # for row in self.obs.plainTextEdit.toPlainText():
            #     if row !='\n':
            #         observs.append(row)            
            obs_dict = {'Observações':observs}   
            self.pd_obs = pd.DataFrame(obs_dict)
        else:
            self.win_constr.pbObs.setStyleSheet('background-color: rgb(0, 170, 0)')


    @QtCore.Slot()
    def bt_hibrido(self):
        if self.hibrido.inputDi.value()==0:
            self.hibrido.inputDi.setValue(self.win_constr.inputDs.value())  
        self.win_constr.checkBoxAutoTurn.setChecked(False)
        self.hibrido.inputLn.setValue(self.win_system.inputLn.value())      
        self.model_cil = DataFrameModel(self.cilindros)
        self.hibrido.tv_Cil.setModel(self.model_cil)
        model_cam = DataFrameModel(self.camadas_dados)
        self.hibrido.tv_Camadas.setModel(model_cam) 
        self.hibrido.exec_()
        pass 
        

    def set_win_force_side(self):
        isoladores = reactor.isoladores()  
        nbi = self.win_system.inputFatorNBI.value()*self.win_system.inputNBI.value()      
        q = """SELECT * FROM isoladores
               WHERE CAST(isoladores."NBI (kV)" AS DOUBLE)>=""" + str(nbi) +\
            """ ORDER BY isoladores."NBI (kV)"  """
        #cond = sqldf(q,locals())
        isoladores = con.execute(q).df()
        nbis = isoladores['NBI (kV)'].astype(str)
        self.force_lado_a_lado.cbIsolador.addItems(isoladores['TIPO'])

        pass

    

    @QtCore.Slot()
    def bt_criar(self):           
        self.saveFileDialog()            
        self.caminhodados = self.arquivo_rt + '.xlsx'   
        self.arquivo = self.arquivo_rt.rsplit('/', 1)[-1]
        print(self.arquivo)      
        print(self.caminhodados)    
        criar.criardocs(0,self.caminhodados,self.arquivo)
        pass

    @QtCore.Slot()
    def force_side_by_side(self):
        self.force_lado_a_lado.exec_()
        pass

    @QtCore.Slot()
    def bt_FreeCadTap(self):
        breecad = BreeCad.BreeCadReactor(51)
        # breecad.modelo(self.cilindros, self.camadas_dados, self.constr_dados)
        larguraperfil = self.hibrido.inputLargPerf.value()
        raio = (self.hibrido.inputDi.value()+larguraperfil)/2
        alturacalco = self.hibrido.inputCalco.value()
        numerovoltas = self.hibrido.inputNesp.value()
        numentradas = self.hibrido.inputNumPerf.value()
        DBST = self.hibrido.inputDBSb.value()
        posicao  = self.hibrido.cbPosicao.currentText()
        perfil = self.hibrido.inputEspPerfil.value()
        h1 = self.win_constr.inputHs.value()
        dbs1 = self.win_constr.inputDBS.value()
        breecad.clearAll()
        breecad.modelo(self.cilindros, self.camadas_dados, self.constr_dados,self.win_constr.cB_FC_model.isChecked())
        breecad.criarperfil(raio=raio,h1=h1,dbs1=dbs1,alturaperfil=perfil,larguraperfil=larguraperfil,
                            alturacalco=alturacalco, numerovoltas=numerovoltas, numentradas=numentradas,
                            DBST=DBST,posicao=posicao)
                            
        self.hibrido.close()
        self.win_constr.close()                    

    @QtCore.Slot()
    def bt_FreeCad(self):
        # freecad.desenharReator()
        breecad = BreeCad.BreeCadReactor(51)
        # breecad.desenha(self.cilindros)
        breecad.clearAll()
        breecad.modelo(self.cilindros, self.camadas_dados, self.constr_dados,self.win_constr.cB_FC_model.isChecked())
        self.win_constr.close()

    @QtCore.Slot()
    def bt_ins_cil(self):
        indices = self.win_constr.tableViewCil.selectionModel().selectedRows()
        for index in sorted(indices, reverse=False):
            self.model_cil.insertRows(index.row(), 1, index, None)
        self.model_cil._dataframe.sort_values(by='Cilindro', inplace=True)
        for index, cil in self.model_cil._dataframe.iterrows():
            cil['Cilindros'] = index
        self.cilindros = self.model_cil._dataframe

    @QtCore.Slot()
    def bt_del_cil(self):
        """Delete rows with currently selected cells and/or selected rows of the model"""
        #index = self.win_constr.tableViewCil.currentIndex()
        #self.model_cil.removeRows(index.row(), 1, index)
        indices = self.win_constr.tableViewCil.selectionModel().selectedRows()
        for index in sorted(indices, reverse=True):
            self.model_cil.removeRows(index.row(), 1, index)
        for index, cil in self.model_cil._dataframe.iterrows():
            cil['Cilindros'] = index

    @QtCore.Slot()
    def crud_fibra(self):
        model_fibras = DataFrameModel(reactor.fibras())
        self.win_crud_fibra.tableView.setModel(model_fibras)
        self.win_crud_fibra.exec_()

    @QtCore.Slot()
    def bt_calc_enable(self):
        self.win_constr.btn_calc.setEnabled(True)
        #self.win_constr.btn_calc.setEnabled(False)

    @QtCore.Slot()
    def Ln_change(self):
        Lalvo = 1.015*self.win_system.inputLn.value()
        self.win_system.inputLtarget.setValue(Lalvo)

    @QtCore.Slot()
    def NBI_change(self):
        NBI = self.win_system.inputNBI.value()
        fator_nbi = self.win_system.inputFatorNBI.value()
        self.win_system.inputNBI_sea_level.setValue(NBI*fator_nbi)

    @QtCore.Slot()
    def iN_change(self):
        iN = self.win_system.inputInominal.value()
        fator_corrente = self.win_system.InputFatorCorrente.value()
        self.win_system.inputITotal.setValue(iN*fator_corrente)

    @QtCore.Slot()
    def icc_sys_change(self):
        iccsys = self.win_system.inputICCSys.value()
        vsys = self.win_system.inputVn.value()
        zsys = vsys/iccsys
        ind = self.win_system.inputLn.value()
        zys_plus_xl = zsys + 2*reactor.np.pi*ind * \
            self.win_system.inputFreq.value()/1000
        icc = vsys*1.05/(zys_plus_xl)
        self.win_system.inputICC.setValue(icc)

    @QtCore.Slot()
    def icc_change(self):
        icc = self.win_system.inputICC.value()
        ipeak = 2.55*icc
        self.win_system.inputPico.setValue(ipeak)

    @QtCore.Slot()
    def preencher_harm(self):
        start = float(self.win_harm.inputF1.text())
        n_harm = int(self.win_harm.inputNharm.text())
        passo = float(self.win_harm.inputPasso.text())
        In = self.win_system.inputInominal.value()
        rms_s = []
        freqs = []
        for harm in range(n_harm):
            freqs.append(start)
            if harm == 0:
                rms_s.append(In)
            else:
                rms_s.append(1)
            start += passo
        # dictionary of lists
        dict = {'RMS (A)': rms_s, 'freq (Hz)': freqs}
        self.harm_dados = pd.DataFrame(dict)
        model_harm = DataFrameModel(self.harm_dados)
        self.win_harm.tableViewHarm.setModel(model_harm)
        pass

    def set_win_constr(self):
        condutores = reactor.condutores()
        q = reactor.string_cond(155)
        #cond = sqldf(q,locals())
        cond = con.execute(q).df()
        cruzetas = reactor.cruzetas()
        fibras = reactor.fibras()
        self.win_constr.comboBoxFibra.clear()
        self.win_constr.comboBoxFibra.addItems(fibras['Padrão'])
        self.win_constr.comboBoxCruzInf.clear()
        self.win_constr.comboBoxCruzSup.clear()
        self.win_constr.comboBoxAWG.clear()
        awg = cond['AWG'].astype(str)
        self.win_constr.comboBoxAWG.addItems(awg)
        self.win_constr.comboBoxAWG.setCurrentText('11.0')
        cruzes = cruzetas['Formação']
        # print(cruzes)
        self.win_constr.comboBoxCruzInf.addItems(cruzes)
        self.win_constr.comboBoxCruzSup.addItems(cruzes)
        if self.win_constr.inputDs.text() == '':
            self.win_constr.inputDs.setValue(self.win_constr.inputDi.value())

    def ini_harm(self):
        In = self.win_system.inputInominal.value()
        if self.harm_dados.isnull:
            rms_s = [In]
            freqs = [self.win_system.inputFreq.value()]
        # dictionary of lists
        dict = {'RMS (A)': rms_s, 'freq (Hz)': freqs}
        self.harm_dados = pd.DataFrame(dict)
        model_harm = DataFrameModel(self.harm_dados)
        self.win_harm.tableViewHarm.setModel(model_harm)
        pass

    @QtCore.Slot()
    def harmonicos(self):
        self.ini_harm()
        self.win_harm.exec_()
        pass

    @QtCore.Slot()
    def novo_projeto(self):
        self.win_system.exec_()
        self.ini_harm()
        self.win_harm.exec_()
        self.set_win_constr()
        self.win_constr.exec_()
        pass

    @QtCore.Slot()
    def iniciar(self):
        rho20, dens = self.material_condutor()
        condutores = reactor.condutores()
        awg = float(self.win_constr.comboBoxAWG.currentText())
        q = reactor.string_cil_cond(self.classe, awg)
        cond = con.execute(q).df()
        min_awg = cond['AWG'].min()
        In = self.win_system.inputITotal.value()
        ax = int(self.win_constr.inputAxial.value())
        rd = int(self.win_constr.inputRadial.value())
        qc = """Select * From cond \
               Where AWG="""+str(awg)+"""   """
        cil = con.execute(qc).df()
        fibras = reactor.fibras()
        q = """Select * From fibras \
               Where "Padrão" = '"""+self.win_constr.comboBoxFibra.currentText()+"""'   """
        fibra = con.execute(q).df()
        fi1 = fibra['Fi1'][0]
        fin = fibra['Fin'][0]
        fen = fibra['Fen'][0]
        feExt = fibra['FeExt'][0]
        awgs = [awg]
        axiais = [ax]
        radiais = [rd]
        fi = [fi1]
        fe = [fen]
        secao = cil['Seção condutora avg (mm2)'][0]*ax*rd
        idx_cil = 0
        cils = [idx_cil]
        rel = secao/In

        while rel < 0.5:
            if rel < 0.5:
                if min_awg < awg-0.5:
                    awg = awg-0.5
                awgs.append(awg)
                axiais.append(ax)
                if rd > 1:
                    radiais.append(rd-1)
                else:
                    radiais.append(rd)
                fi.append(fin)
                fe.append(fen)
                idx_cil += 1
                cils.append(idx_cil)
                qc = """Select * From cond \
                     Where AWG="""+str(awg)+"""   """
                cil = con.execute(qc).df()
                if rd > 1:
                    secao += cil['Seção condutora avg (mm2)'][0]*ax*(rd-1)
                else:
                    secao += cil['Seção condutora avg (mm2)'][0]*ax*rd
            rel = secao/In

        sec_rels = secao+cil['Seção condutora avg (mm2)'][0]*ax*rd
        rels = sec_rels/In
        last_item = len(radiais)-1
        radiais[last_item] = rd
        fe[last_item] = feExt
        espc = self.win_constr.inputLargEsp.value()
        axiais = reactor.axial_cil(awgs, axiais)
        radiais = reactor.radial_cil(awgs, radiais)
        r_m = reactor.rho_cil(awgs, rho20=rho20)
        if len(awgs) > 1:
            ducts = [espc]
        else:
            ducts = [0]
        ducts = reactor.duto_cil(awgs, ducts, espc)
        folga = self.win_constr.input_Folga.value()
        fator_folga = reactor.fator_folga(awgs, folga)
        r_esp = reactor.rho_cil(awgs, rho20=rho20)
        densidade = reactor.densidade(awgs, dens)
        dict = {'Cilindro': cils, 'condutor AWG': awgs, 'Fios axiais': axiais, 'Fios radiais': radiais, 'Fibra interna (mm)': fi,
                'Fibra externa (mm)': fe, 'Espaçadores (mm)': ducts, 'Fator de enrolamento': fator_folga, 'Resistividade': r_m, 'Densidade': densidade}
        cilindros = reactor.pd.DataFrame(dict)
        self.cilindros = cilindros
        #q = reactor.string_join_cil_cond(classe)
        #self.cil = con.execute(q).df()
        self.model_cil = DataFrameModel(self.cilindros)
        self.win_constr.tableViewCil.setModel(self.model_cil)        
        self.win_constr.btn_calc.setEnabled(True)
        self.win_constr.btn_calc.setStyleSheet('background-color: rgb(255, 0, 0)')

    def material_condutor(self,material='Al'):
        secao = 0
        materiais = reactor.materiais()
        self.classe = float(self.win_constr.comboBoxClasse.currentText())
        q = reactor.query_material(material=material)
        self.mat = con.execute(q).df()
        rho20 = self.mat['Resistividade (Ohm x mm2 /m)'][0]
        self.tref = self.win_system.inputTref.value()
        rho_Tref = (self.mat['Coeficiente K'][0] + self.tref)/(self.mat['Coeficiente K'][0]+20)*rho20
        dens = self.mat['Densidade (kg/m3)'][0]
        return rho20, dens

    @QtCore.Slot()
    def calcular(self):
        self.win_constr.comboBox_TempIcc.clear()
        self.win_constr.comboBox_IccMax.clear()
        #self.win_constr.comboBoxNormaIcc.clear()
        ax = int(self.win_constr.inputAxial.value())
        cruzetas = reactor.cruzetas()
        q = reactor.string_cruz(self.win_constr.comboBoxCruzSup.currentText())
        cruz_sup = con.execute(q).df()
        q = reactor.string_cruz(self.win_constr.comboBoxCruzInf.currentText())
        cruz_inf = con.execute(q).df()
        rho20, dens = self.material_condutor(self.win_constr.condutorComboBox.currentText())
        self.classe = float(self.win_constr.comboBoxClasse.currentText())
        condutores = reactor.condutores()
        cilindros = self.cilindros
        q = reactor.string_join_cil_cond(self.classe)
        self.cilindros = con.execute(q).df()
        rel = con.from_df(self.cilindros)
        rel.types
        col = rel.columns
        di = self.win_constr.inputDi.value()
        In = self.win_system.inputITotal.value()
        try:
            secao = sum(self.cilindros['area'])
        except:
            self.cilindros['area'] = self.cilindros['Seção condutora avg (mm2)']*self.cilindros['Fios axiais']*cilindros['Fios radiais']
            secao = sum(self.cilindros['area'])
                
        self.cilindros['Corrente Alvo'] = In*self.cilindros['area']/secao
        camadas, ds, secao = self.auto_turn_he(di, In)        
        espc = self.win_constr.inputLargEsp.value() 
        harms = self.harm_dados
        #harms['RMS (A)'] = harms['RMS (A)'].astype(float)
        harms.astype({'RMS (A)': 'float'}).dtypes
        self.harm_dados = harms
        anel_i = self.win_constr.inputAnelInf.value()
        anel_s = self.win_constr.inputAnelSup.value()
        Ltarget = self.win_system.inputLtarget.value()
        loop = 0        
        temp = True
        dtant = 1000
        while temp:            
            self.M, self.R, ind_calc, ind, flx, itarget = reactor.equilibrio2(
                camadas, Ltarget, nbr=max(
                    cruz_inf["Braços"][0], cruz_sup["Braços"][0]), equilibrar=self.win_constr.checkBoxAutoTurn.isChecked())
            rms_s = [In]
            freq = self.win_system.inputFreq.value()                        
            R, X, Z, V, I = reactor.mesh(
                ind=ind_calc, M=self.M, R=self.R, harms=harms, modulos=camadas)               
            camadas['Peso fio (kg)'] = camadas['Peso espira']*camadas['# Espiras'].astype('float')
            res = reactor.result_by_cil(camadas)
            res['Espessura'] = self.cilindros['dfio isol avg (mm)']*self.cilindros['Fios radiais'].astype(
                'int32')*reactor.np.cos(reactor.np.pi/6)
            self.cilindros['Espessura'] = res['Espessura']
            self.cilindros['# Espiras'] = res['# Espiras']
            self.cilindros['Altura do enrolamento'] = res['Altura do enrolamento']
            self.cilindros['Diâmetro médio'] = res['Diâmetro médio']            
            self.cilindros['Peso fio(kg)'] = res['Peso fio (kg)']
            ds_cil = res['Diâmetro médio']+res['Espessura']+2*self.cilindros['Fibra externa (mm)']
            self.cilindros['# Espaçadores'] = ds_cil
            self.cilindros['# Espaçadores'] = self.cilindros['# Espaçadores'].apply(
                                        lambda x: int(reactor.np.pi*x/70) 
                                        if x>600 else int(reactor.np.pi*x/(70 - (600 - x) / 10)))
            ncil = len(cilindros)
            self.cilindros['# Espaçadores'][ncil-1] = 0
            self.cilindros['Corrente (A)'] = reactor.np.abs(res['Corrente (A)'])
            corrT = (self.mat['Coeficiente K'][0]+self.tref)/(self.mat['Coeficiente K'][0]+20)
            # Perdas DC corrigidas de 20⁰C para temperatura de referência normalmente 75 ⁰C
            self.cilindros['Perdas (W)'] = res['Perdas (W)']*corrT
            self.cilindros['Z(mm)'] = res['Z(mm)']
            rho20 = self.mat['Resistividade (Ohm x mm2 /m)'][0]
            self.cilindros, total_pad = reactor.ind_losses(self.cilindros, frq=freq, ro=rho20)
            # Perdas adicionais corrigidas de 20⁰C para temperatura de referência normalmente 75 ⁰C
            self.cilindros['Pad (W)'] = self.cilindros['Pad (W)']/corrT
            # Perdas de corrente alternada somatório das perdas DC mais AD
            self.cilindros['Pac (W)'] = self.cilindros['Perdas (W)'] + self.cilindros['Pad (W)']             
            dbs = max(camadas['Altura do enrolamento'].astype('float'))+anel_s+anel_i 
            hotspot, tmed, dtmax = self.temperatura(dbs, espc,In) 
            if dtmax<1:
                temp = False              
            else:
                #
                if loop>10:
                    temp = False
                    break
                elif dtmax>dtant:
                    temp = False
                    break
                cam = 0
                for index, cil in self.cilindros.iterrows():
                    frd = int(cil['Fios radiais'])
                    for layer in range(frd):
                        camadas['Corrente alvo(A)'][cam] = cil['Corrente Alvo']/frd 
                        cam += 1
            if not self.win_constr.checkBoxAutoTurn.isChecked():
                temp = False
            dtant = dtmax
            loop += 1            
        jdc = float(harms['RMS (A)'][0])/secao                                
        jdc = In/secao
        self.win_constr.inputJdc.setValue(jdc)
        zs = V/I
        z = V[0]/sum(I)
        x = z.imag        
        self.win_constr.inputLcalc.setValue(ind_calc)
        Ln = self.win_system.inputLn.value()
        self.win_constr.inputErro.setValue(100*reactor.np.abs(1-ind_calc/Ln))
        model_cam = DataFrameModel(camadas)
        self.win_constr.tableViewCam.setModel(model_cam)
        self.win_constr.inputDs.setValue(ds)        
        ind_calc_mesh = x*1000/(2*reactor.np.pi*freq)
        self.win_constr.inputErro.setValue(100*reactor.np.abs(1-ind_calc_mesh/Ln))
        self.win_constr.inputLcalc.setValue(ind_calc_mesh)        
        rdc75 = 1/sum(1/camadas['Rdc @ 20°C (mΩ)'])*corrT
        pdc75 = rdc75*In**2/1000000
        self.win_constr.inputPerdas.setValue(pdc75)
        fatores = reactor.fator_nao_linearidade(camadas)
        NBI = self.win_system.inputNBI.value() * self.win_system.inputFatorNBI.value()
        camadas['kV/esp'] = fatores*NBI/(camadas['# Espiras'].astype('float')*min(ax, cruz_inf["Braços"][0], cruz_sup["Braços"][0]))
        self.win_constr.inputNBI_Esp.setValue(max(camadas['kV/esp'].astype('float')))
        self.aproved(self.win_constr.inputNBI_Esp, 10)
        camadas['kV/mm'] = fatores*NBI /camadas['Altura do enrolamento'].astype('float')
        self.win_constr.inputNBI_He.setValue(max(camadas['kV/mm'].astype('float')))
        self.aproved(self.win_constr.inputNBI_He, 0.5)
        kv_cil, kv_cam = reactor.kv_camada_cil(NBI, fatores, self.cilindros, camadas)
        camadas['kV entre camadas'] = kv_cam        
        model_cam = DataFrameModel(camadas)
        self.win_constr.tableViewCam.setModel(model_cam)        
        # Forças de compressão nos enrolamentos
        ipeak_in = self.win_system.inputPico.value()*1000/self.win_system.inputITotal.value()
        self.cilindros, compress = reactor.compress(self.cilindros, ipeak_in=ipeak_in)
        self.win_constr.inputCompress.setValue(compress)
        # Forças de hoop nos enrolamentos
        self.cilindros, hoop = reactor.hoop(self.cilindros, ipeak_in)
        self.win_constr.inputHoop.setValue(hoop)
        self.cilindros['kV entre cilindros'] = kv_cil
        self.win_constr.inputNBI_Cil.setValue(max(kv_cil))
        self.aproved(self.win_constr.inputNBI_Cil, 50)
        self.win_constr.inputNBI_Cam.setValue(max(kv_cam))
        self.aproved(self.win_constr.inputNBI_Cam, 10)        
        self.win_constr.inputDBS.setValue(dbs)
        hs = cruz_inf["Largura (mm)"][0]+cruz_sup["Largura (mm)"][0]+dbs
        self.win_constr.inputHs.setValue(hs)       
        self.win_constr.InputTmed.setValue(tmed)
        self.win_constr.inputHotSpot.setValue(hotspot)
        self.aproved(self.win_constr.inputHotSpot, self.classe)
        total_pad75 = total_pad/(corrT*1000)
        Pac75 = pdc75+total_pad75
        self.win_constr.inputPac.setValue(Pac75)
        qualidade = 2*reactor.np.pi*harms['freq (Hz)'][0]*ind_calc/1000 *harms['RMS (A)'][0]**2/(Pac75*1000)
        self.win_constr.inputQ.setValue(qualidade)
        self.model_cil = DataFrameModel(self.cilindros)
        self.win_constr.tableViewCil.setModel(self.model_cil)
        self.cilindros_dados = self.cilindros
        self.camadas_dados = camadas
        peso_fio = sum(self.cilindros['Peso fio(kg)'] )
        self.win_constr.pesoTotalDoCondutorKgDoubleSpinBox.setValue(peso_fio)
        incc = self.win_system.inputICC.value()*self.win_system.InputFatorCorrente.value()*self.cilindros['Corrente (A)']/In
        t = self.win_system.inputTempo.value()
        k = self.mat['Coeficiente K'][0]
        #self.icc_iec_max = sum(0.001*self.cilindros['area']*(1000*(190-hotspot)/((0.043500*(200+hotspot)/2 + 10.430000)*t))**.5)
        jcc = 1000*incc/self.cilindros['area']
        tmaxicc = 200
        self.icc_iec_max = sum(reactor.np.sqrt(45700*tmaxicc/(2*(hotspot+k)+tmaxicc)*t)*self.cilindros['area'])/1000
        
        # print(icc_iec_max)
        # self.temp_iec = hotspot+t*(0.043500*(200+hotspot)/2 + 10.430000)*(1000*incc/self.cilindros['area'])**2/1000
        
        self.temp_iec = hotspot+2*(hotspot+k)/(45700/(jcc**2*t)-1)
        # temp_ansi = tmed+t*(0.043500*(200+tmed)/2 + 10.430000)*(1000*incc/self.cilindros['area'])**2/1000
        # print(temp_iec)
        self.temp_iec = mean(self.temp_iec)
        #temp_ansi = max(temp_ansi)
        self.win_constr.comboBox_TempIcc.addItem(f'{self.temp_iec:.1f}')
        self.win_constr.comboBox_IccMax.addItem(f'{self.icc_iec_max:.1f}') 
        tamb = self.win_system.inputTamb.value()
        self.tempLimitIscAnsi = -0.00000571088815*(self.classe)**4 + 0.00384187291*(self.classe)**3 - 0.955677072*(self.classe)**2 + 105.535235*self.classe - 4093.14609        
        tstart = (hotspot-tamb)*(k+hotspot-tamb+30)/(k+hotspot)+30
        k60 = Pac75/pdc75
        m = 0.5*reactor.np.log(((k+self.tempLimitIscAnsi)**2+(k60-1)*(k+tstart)**2)/(k60*(k+tstart)**2))
        secao_isol = sum(dbs*self.cilindros['espessura']/self.cilindros['# Espiras'])-secao
        cx = 892.9+0.4409*tmed+793.7*(secao_isol-secao)/secao
        rac = Pac75*1000000/In**2
        self.icc_ansi_max = (m*cx*peso_fio*(k+tstart)/(t*rac/1000*(k+hotspot-tamb+30)/(k+hotspot)))**0.5/1000
        self.win_constr.comboBox_IccMax.addItem(f'{self.icc_ansi_max:.1f}')
        m = t*rac/1000*(sum(incc)*1000)**2/(cx*(k+tstart)*peso_fio)
        self.temp_ansi = (k+tstart)*((reactor.np.exp(2*m)+(k60-1)*(reactor.np.exp(2*m)-1))-1)+tstart
        self.win_constr.comboBox_TempIcc.addItem(f'{self.temp_ansi:.1f}')
        # print(temp_iec)
        self.constr_pd()
        self.win_constr.btn_calc.setStyleSheet('background-color: rgb(0, 255, 0)')

    def auto_turn_he(self, di, In):
        if self.win_constr.checkBoxAutoTurn.isChecked():
            camadas, ds, secao = reactor.camadas(di=di, cils=self.cilindros, x=0, y=0, z=0, corr=In)
            self.win_constr.inputDs.setValue(ds)
            self.win_constr.inputSecao.setValue(secao)
            self.win_constr.InputTmed.setValue(75)
            self.win_constr.inputHotSpot.setValue(75)
            self.win_constr.inputPerdas.setValue(0)
            self.win_constr.inputPac.setValue(0)
            self.win_constr.inputQ.setValue(0)
        else:
            q = reactor.string_camadas()
            camadas = self.camadas_dados.drop('Corrente (A)', axis='columns')
            camadas['# Espiras'] = camadas['# Espiras'].astype('float')
            camadas['Altura do enrolamento'] = camadas['Altura do enrolamento'].astype(
                'float')
            camadas = con.execute(q).df()
            if self.win_constr.checkBoxHe.isChecked():
                camadas['Altura do enrolamento'] = camadas['# Espiras'] *camadas['Altura da espira'].astype('float')
            else:
                camadas['Altura do enrolamento'] = camadas['Altura do enrolamento'].astype('float')    
                camadas['Diâmetro médio'] = camadas['Diâmetro médio'].astype('float')
            ds = self.win_constr.inputDs.value()
            secao = self.win_constr.inputSecao.value()
        return camadas, ds, secao

    def aproved(self, spinbox, limite):
        if spinbox.value() > limite:
            spinbox.setStyleSheet("QDoubleSpinBox"
                                  "{"
                                  "background : orange;"
                                  "}")
            return False
        else:
            spinbox.setStyleSheet("QDoubleSpinBox"
                                  "{"
                                  "background : lightgreen;"
                                  "}")
            return True

    def temperatura(self, dbs, espc, In):
        ncil = self.cilindros.shape[0]
        hc = reactor.heat_constant(ncil, dbs, espc, self.classe)
        self.cilindros['Heat constant'] = hc
        tamb = self.win_system.inputTamb.value()
        dt = self.tref-tamb
        tmed_ant = tamb
        Pac_Tmed = self.cilindros['Pac (W)']
        while dt > 1:
            if self.win_constr.comboBox_metodo_temp.currentIndex() == 0:  # Método Heat constant
                self.cilindros['Tmed'] = tamb+self.cilindros['Heat constant']*Pac_Tmed/(
                    self.cilindros['Diâmetro médio']/25.4*reactor.np.sqrt(self.cilindros['Altura do enrolamento']/25.4))
            else:  # Método Dubbel Vol1 pgs 469, 475
                Q_KCAL_por_Hora = Pac_Tmed*3600/4184
                area_calor = 2*reactor.np.pi * \
                    self.cilindros['Diâmetro médio']*dbs/1000000
                if dbs > 300:
                    self.cilindros['Tmed'] = tamb+(Q_KCAL_por_Hora*(
                        self.cilindros['Diâmetro médio']/1000)**0.25/(1.13*area_calor))**0.8
                else:
                    self.cilindros['Tmed'] = tamb+(Q_KCAL_por_Hora *
                                                   (dbs/1000)**0.25/(1.18*area_calor))**0.8
            hotspot = max(self.cilindros['Tmed'])-tamb
            tmed = self.cilindros['Tmed'].mean()
            corrT = (self.mat['Coeficiente K'][0]+tmed) / (self.mat['Coeficiente K'][0]+tmed_ant)
            # Perdas DC corrigidas de 75⁰C para temperatura de referência normalmente Tmed ⁰C
            Pdc_Tmed = self.cilindros['Perdas (W)']*corrT
            Pad_Tmed = self.cilindros['Pad (W)']/corrT
            # Perdas de corrente alternada somatório das perdas DC mais AD
            Pac_Tmed = Pdc_Tmed+Pad_Tmed
            dt = tmed-tmed_ant
            tmed_ant = tmed
        if dbs <= 1000:
            hotspot = hotspot*1.3+tamb
        elif dbs <= 2000:
            hotspot = hotspot*1.4+tamb
        else:
            hotspot = hotspot*1.45+tamb
        #print(self.cilindros['Corrente Alvo'])
        self.cilindros['Corrente Alvo'] = ((tmed-40)/(self.cilindros['Tmed']-40))**0.5*self.cilindros['Corrente Alvo'] 
        #print(sum(self.cilindros['Corrente Alvo']))
        self.cilindros['Corrente Alvo'] = In/sum(self.cilindros['Corrente Alvo'])*self.cilindros['Corrente Alvo']  
        #print(sum(self.cilindros['Corrente Alvo']))
        # print(self.cilindros['Corrente Alvo'])
        return hotspot, tmed, max(self.cilindros['Tmed'])-min(self.cilindros['Tmed'])    

    @QtCore.Slot()
    def norma_temp_change(self):
        if self.win_constr.comboBox_TempIcc.currentIndex()==0:
            self.win_constr.label_temp_norma.setText('Temp. Icc pela norma IEC<200 °C')
            if float(self.win_constr.comboBox_TempIcc.currentText())<200:
                self.win_constr.comboBox_TempIcc.setStyleSheet('background-color: rgb(0, 255, 0)')
            else:
                self.win_constr.comboBox_TempIcc.setStyleSheet('background-color: rgb(255, 0, 0)')
        else:
            self.win_constr.label_temp_norma.setText('Temp. Icc pela norma ANSI<'+f'{self.tempLimitIscAnsi:.1f}'+' °C')
            if float(self.win_constr.comboBox_TempIcc.currentText())<self.tempLimitIscAnsi:
                self.win_constr.comboBox_TempIcc.setStyleSheet('background-color: rgb(0, 255, 0)')
            else:
                self.win_constr.comboBox_TempIcc.setStyleSheet('background-color: rgb(255, 0, 0)')

    @QtCore.Slot()
    def norma_icc_change(self):
        icc = self.win_system.inputICC.value()
        if self.win_constr.comboBox_IccMax.currentIndex()==0:
            self.win_constr.labelInccNorma.setText('Incc Máximo pela IEC (kA)>'+f'{icc:.1f}'+' kA')
            if float(self.win_constr.comboBox_IccMax.currentText())>icc:
                self.win_constr.comboBox_IccMax.setStyleSheet('background-color: rgb(0, 255, 0)')
            else:
                self.win_constr.comboBox_IccMax.setStyleSheet('background-color: rgb(255, 0, 0)')
        else:
            self.win_constr.labelInccNorma.setText('Incc Máximo pela ANSI (kA)>'+f'{icc:.1f}'+' kA')
            if float(self.win_constr.comboBox_IccMax.currentText())>icc:
                self.win_constr.comboBox_IccMax.setStyleSheet('background-color: rgb(0, 255, 0)')
            else:
                self.win_constr.comboBox_IccMax.setStyleSheet('background-color: rgb(255, 0, 0)')

    @QtCore.Slot()
    def fibra_change(self):
        Pass

    def widgets_sys(self):
        if not (self.sys_data.empty):
            self.win_system.inputFreq.setValue(
                self.sys_data['Frequência (Hz)'][0])
            self.win_system.inputVn.setValue(
                self.sys_data['Tensão nominal (V)'][0])
            self.win_system.inputNBI.setValue(self.sys_data['NBI (kV)'][0])
            self.win_system.normaComboBox.setCurrentText(
                self.sys_data['Norma'][0])
            self.win_system.inputAlt.setValue(
                self.sys_data['Altitude (m)'][0])
            self.win_system.inputVvento.setValue(
                self.sys_data['Velocidade do vento (km/h)'][0])
            self.win_system.inputICC.setValue(
                self.sys_data['Corrente de curto-circuito (kA)'][0])
            self.win_system.inputTempo.setValue(
                self.sys_data['Duração (s)'][0])
            self.win_system.inputPico.setValue(
                self.sys_data['Icc Pico (kA)'][0])
            self.win_system.inputTamb.setValue(
                self.sys_data['Temperatura ambiente (°C)'][0])
            self.win_system.inputTref.setValue(
                self.sys_data['Temperatura de referência (°C)'][0])
            self.win_system.inputLn.setValue(
                self.sys_data['Indutância nominal (mH)'][0])
            self.win_system.inputLtarget.setValue(
                self.sys_data['Indutância alvo (mH)'][0])
            self.win_system.inputInominal.setValue(
                self.sys_data['Corrente nominal (A)'][0])
            self.win_system.inputITotal.setValue(
                self.sys_data['Corrente total (A)'][0])
            self.win_system.InputFatorCorrente.setValue(
                self.sys_data['Fator de correção da corrente pela altitude'][0])
            self.win_system.inputFatorNBI.setValue(
                self.sys_data['Fator de correção do NBI pela altitude'][0])
               

    def widgets_contr(self):
        if not (self.constr_dados.empty):
            self.win_constr.condutorComboBox.setCurrentText(
                self.constr_dados['Material condutor'][0])
            self.win_constr.comboBoxClasse.setCurrentText(
                str(self.constr_dados['Classe de temperatura do isolamento (°C)'][0]))
            self.win_constr.inputAnelInf.setValue(
                self.constr_dados['Anel de fibra inferior (mm)'][0])
            self.win_constr.inputAnelSup.setValue(
                self.constr_dados['Anel de fibra superior (mm)'][0])
            self.win_constr.inputDi.setValue(
                self.constr_dados['Diâmetro interno (mm)'][0])
            self.win_constr.inputDs.setValue(
                self.constr_dados['Diâmetro externo (mm)'][0])
            self.win_constr.inputLargEsp.setValue(
                self.constr_dados['Largura do espaçador (mm)'][0])
            self.win_constr.comboBoxCruzInf.setCurrentText(
                self.constr_dados['Cruzeta inferior'][0])
            self.win_constr.comboBoxCruzSup.setCurrentText(
                self.constr_dados['Cruzeta superior'][0])
            self.win_constr.comboBoxAWG.setCurrentText(
                str(self.constr_dados['AWG 1'][0]))
            self.win_constr.comboBoxFibra.setCurrentText(
                self.constr_dados['Fibra'][0])
            self.win_constr.inputRadial.setValue(
                self.constr_dados['# Fios Radiais'][0])
            self.win_constr.inputAxial.setValue(
                self.constr_dados['# Fios Axiais'][0])
            self.win_constr.input_Folga.setValue(
                self.constr_dados['Fator de folga'][0])
            self.win_constr.inputSecao.setValue(
                self.constr_dados['Seção condutora mm²'][0])
            self.win_constr.inputJdc.setValue(
                self.constr_dados['Densidade de corrente em DC (A/mm²)'][0])
            self.win_constr.inputLcalc.setValue(
                self.constr_dados['Indutância calculada (mH)'][0])
            self.win_constr.inputErro.setValue(
                self.constr_dados['(Lcalc/Ln-1)*100 (%)'][0])
            self.win_constr.inputPerdas.setValue(
                self.constr_dados['Perdas DC (kW)'][0])
            self.win_constr.inputPac.setValue(
                self.constr_dados['Perdas AC (kW)'][0])
            self.win_constr.inputQ.setValue(
                self.constr_dados['Fator Q'][0])
            self.win_constr.InputTmed.setValue(
                self.constr_dados['Tmed (°C)'][0])
            self.win_constr.inputHotSpot.setValue(
                self.constr_dados['Hot Spot (°C)'][0])
            self.win_constr.comboBox_metodo_temp.setCurrentText(
                str(self.constr_dados['Método de cálculo da temperatura'][0]))
            self.win_constr.checkBoxAutoTurn.setChecked(
                self.constr_dados['Auto equilíbrio do número de espiras'][0])
            self.win_constr.checkBoxHe.setChecked(
                self.constr_dados['Auto equilíbrio da altura'][0])
            self.model_cil = DataFrameModel(self.cilindros_dados)
            self.win_constr.tableViewCil.setModel(self.model_cil)
            model_cam = DataFrameModel(self.camadas_dados)
            self.win_constr.tableViewCam.setModel(model_cam)
            self.win_constr.btn_calc.setEnabled(True)
            self.win_constr.btForce.setEnabled(True)
            self.win_constr.btEmpilhado.setEnabled(True)
            self.win_constr.pbCriar.setEnabled(True)
            self.win_constr.pbHibrido.setEnabled(True)
            self.win_constr.pbQRing.setEnabled(True)
            self.win_constr.btnFreeCad.setEnabled(True)
            self.hibrido.inputL1.setValue(self.win_constr.inputLcalc.value())
            self.hibrido.inputLn.setValue(self.win_system.inputLn.value())
            try:
                self.win_constr.comboBoxMontagem.setCurrentText(
                    str(self.constr_dados['Montagem'][0]))
            except:
                pass        

    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        self.fileName, _ = QFileDialog.getOpenFileName(
            self, "Abrir arquivo de projeto", "", "All Files (*);;Python Files (*.py)", options=options)
        if self.fileName:
            try:
                self.dados_projeto = pd.read_excel(self.fileName,5)
                self.win_system.inputCliente.setText(self.dados_projeto['Cliente'][0])
                self.win_system.inputQuant.setValue(self.dados_projeto['Quantidade'][0])
                self.win_system.inputRef.setText(self.dados_projeto['Referência'][0])
                self.win_system.inputItem.setValue(self.dados_projeto['Item'][0])
                self.win_system.comboBoxAplicacao.setText(self.dados_projeto['Aplicação'][0])
                self.win_system.inputTipo.setText(self.dados_projeto['Tipo'][0]) 
                self.win_system.cbInstall.setText(self.dados_projeto['Instalação'][0])
                self.win_system.inputItem.selText(self.dados_projeto['Item'][0])
                
            except:
                pass               
            self.sys_data = pd.read_excel(self.fileName, 0)
            self.widgets_sys()
            self.sys_data.loc[:, ~self.sys_data.columns.str.match("Unnamed")]
            self.constr_dados = pd.read_excel(self.fileName, 1)
            self.cilindros_dados = pd.read_excel(self.fileName, 2)
            self.cilindros_dados.loc[:, ~self.cilindros_dados.columns.str.match(
                "Unnamed")]
            self.cilindros = self.cilindros_dados
            self.camadas_dados = pd.read_excel(self.fileName, 3)
            self.camadas_dados.loc[:, ~
                                   self.camadas_dados.columns.str.match("Unnamed")]
            self.harm_dados = pd.read_excel(self.fileName, 4)
            self.widgets_contr()
            self.constr_dados.loc[:, ~
                                  self.constr_dados.columns.str.match("Unnamed")]
            try:
                self.open_style = pd.read_excel(self.fileName,7)
                self.model_hibrido = DataFrameModel(self.open_style)
                self.hibrido.tv_Taps.setModel(self.model_hibrido)
                self.dados_taps = pd.read_excel(self.fileName,6)
                self.hibrido.inputL1.setValue(self.dados_taps['L1 (mH)'][0])
                self.hibrido.inputEspPerfil.setValue(self.dados_taps['Altura do Perfil (mm'][0])
                self.hibrido.inputLargPerf.setValue(self.dados_taps['Largura Perfil (mm)'][0])
                self.hibrido.inputNumPerf.setValue(self.dados_taps['# Entradas'][0])
                self.hibrido.cbMaterialTap.setCurrentText(self.dados_taps['Material do perfil'][0])
                self.hibrido.cbPosicao.setCurrentText(self.dados_taps['Instalação'][0])
                self.hibrido.inputDBSb.setValue(self.dados_taps['DBST'][0])
                self.hibrido.inputDi.setValue(self.dados_taps['Diâmetro interno (mm)'][0])
                self.hibrido.inputHe.setValue(self.dados_taps['Altura do enrolamento (mm)'][0])
                self.hibrido.inputNesp.setValue(self.dados_taps['Número de espiras'][0])
                self.hibrido.inputCalco.setValue(self.dados_taps['Calço'][0])
            except:
                dict = {'Tap': [],'Lalvo(mH)':[],'L1 (mH)':[],'Lcalc(mH)':[0.0],'Erro(%)':[],'Ltap(mH)':[],'M(mH)':[], 'Z(mm)': [], '# Espiras': [], 'Altura da espira': [], 
                'Altura do enrolamento': [], 'Diâmetro médio': [], 'Rdc @ 20°C (mΩ)': [], 'Rdc Espira': [],}
                self.open_style = pd.DataFrame(dict)
                self.model_hibrido = DataFrameModel(self.open_style)
            try:
                self.pd_obs = pd.read_excel(self.fileName,8)
                # for index, obs in self.pd_obs.iterrows():
                #     self.obs.plainTextEdit.append(obs)
                print(self.pd_obs)
                self.obs.plainTextEdit.setPlainText(self.pd_obs['Observações'][0])
                if len(self.obs.plainTextEdit.toPlainText())!=0:
                    self.win_constr.pbObs.setStyleSheet('background-color: rgb(255, 0, 0)')
                else:
                    self.win_constr.pbObs.setStyleSheet('background-color: rgb(0, 170, 0)')
            except:
                pass                          

    '''def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        files, _ = QFileDialog.getOpenFileNames(self,"Abrir arquivo de projeto", "","All Files (*);;Python Files (*.py)", options=options)
        if files:
            print(files)'''

    def saveFileDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        self.fileName, _ = QFileDialog.getSaveFileName(
            self, "Salvar arquivo de projeto", "", "Excel files (*.xlsx);;All Files (*);;Text Files (*.txt)", options=options)
        if self.fileName:
            self.dados_tap()
            self.dados_proj()
            self.arquivo_rt, extensao = os.path.splitext(self.fileName)
            if extensao == '':
                self.fileName = self.fileName + '.xlsx'
            path = os.path.normpath(self.fileName)
            with pd.ExcelWriter(
                    path, engine="openpyxl", date_format="DD-MM-YYYY", datetime_format="DD-MM-YYYY HH:MM:SS") as writer:
                self.sys_data.to_excel(
                    writer, sheet_name="Dados do sistema", float_format="%.5f", index_label=False, index=False)
                self.constr_dados.to_excel(
                    writer, sheet_name="Dados construtivos", float_format="%.5f", index_label=False, index=False)
                self.cilindros_dados.to_excel(
                    writer, sheet_name="Cilindros", float_format="%.5f", index_label=False, index=False)
                self.camadas_dados.to_excel(
                    writer, sheet_name="Camadas", float_format="%.5f", index_label=False, index=False)
                self.harm_dados.to_excel(
                    writer, sheet_name="Harmônicas", float_format="%.5f", index_label=False, index=False)
                self.dados_projeto.to_excel(writer, sheet_name="Dados do projeto", float_format="%.5f", index_label=False, index=False)                                
                if type(self.dados_taps)==pd.DataFrame:
                    self.dados_taps.to_excel(writer, sheet_name="Dados do Tap", float_format="%.5f", index_label=False, index=False)    
                    try:
                        self.open_style.to_excel(writer, sheet_name="Taps", float_format="%.5f", index_label=False, index=False)
                    except:
                        pass    
                try:    
                    self.pd_obs.to_excel(writer, sheet_name="Observações", float_format="%.5f", index_label=False, index=False)
                except:
                    pass
                #self.resultado_dados.to_excel(writer, sheet_name="Resultados")

    @QtCore.Slot()
    def cruz_inf_change(self):
        self.win_constr.comboBoxCruzSup.setCurrentText(
            self.win_constr.comboBoxCruzInf.currentText())

    @QtCore.Slot()
    def alt_change(self):
        fator_bil, fator_term = reactor.corr_fact(
            self.win_system.inputAlt.value(), self.win_system.normaComboBox.currentText())
        self.win_system.inputFatorNBI.setValue(fator_bil)
        self.win_system.InputFatorCorrente.setValue(fator_term)
        #self.bt_calc_enable()

    @QtCore.Slot(pd.DataFrame)
    def receive_signal_sys(self, data_sys):
        pass

    @QtCore.Slot()
    def system_data(self):
        self.win_system.exec_()

    @QtCore.Slot()
    def system_contr(self):
        # self.set_win_constr()
        self.win_constr.exec_()

    def dados_proj(self):
        dict={'Cliente':[self.win_system.inputCliente.text()],
            'Quantidade':[self.win_system.inputQuant.value()],
            'Referência':[self.win_system.inputRef.text()], 
            'Aplicação':[self.win_system.comboBoxAplicacao.currentText()],
            'Tipo':[self.win_system.inputTipo.text()],
            'Instalação':[self.win_system.cbInstall.currentText()],
            'Item':[self.win_system.inputItem.value()]}
        self.dados_projeto = pd.DataFrame(dict)

    def constr_pd(self):
        dict = {'Material condutor': [self.win_constr.condutorComboBox.currentText()],
                'Classe de temperatura do isolamento (°C)': [self.win_constr.comboBoxClasse.currentText()],
                'Anel de fibra inferior (mm)': [self.win_constr.inputAnelInf.value()],
                'Anel de fibra superior (mm)': [self.win_constr.inputAnelSup.value()],
                'Diâmetro interno (mm)': [self.win_constr.inputDi.value()],
                'Diâmetro externo (mm)': [self.win_constr.inputDs.value()],
                'Largura do espaçador (mm)': [self.win_constr.inputLargEsp.value()],
                'Cruzeta inferior': [self.win_constr.comboBoxCruzInf.currentText()],
                'Cruzeta superior': [self.win_constr.comboBoxCruzSup.currentText()],
                'AWG 1': [self.win_constr.comboBoxAWG.currentText()],
                'Fibra': [self.win_constr.comboBoxFibra.currentText()],
                'Fator de folga': [self.win_constr.input_Folga.value()],
                '# Fios Axiais': [self.win_constr.inputAxial.value()],
                '# Fios Radiais': [self.win_constr.inputRadial.value()],
                'Seção condutora mm²': [self.win_constr.inputSecao.value()],
                'Densidade de corrente em DC (A/mm²)': [self.win_constr.inputJdc.value()],
                'Indutância calculada (mH)': [self.win_constr.inputLcalc.value()],
                '(Lcalc/Ln-1)*100 (%)': [self.win_constr.inputErro.value()],
                'Perdas DC (kW)': [self.win_constr.inputPerdas.value()],
                'Perdas AC (kW)': [self.win_constr.inputPac.value()],
                'Fator Q': [self.win_constr.inputQ.value()],
                'Tmed (°C)': [self.win_constr.InputTmed.value()],
                'Hot Spot (°C)': [self.win_constr.inputHotSpot.value()],
                'Método de cálculo da temperatura': [self.win_constr.comboBox_metodo_temp.currentText()],
                'Auto equilíbrio do número de espiras': [self.win_constr.checkBoxAutoTurn.isChecked()],
                'Auto equilíbrio da altura': [self.win_constr.checkBoxHe.isChecked()],
                'DBS (mm)': [self.win_constr.inputDBS.value()],
                'Altura (mm):': [self.win_constr.inputHs.value()],
                'Montagem': [self.win_constr.comboBoxMontagem.currentText()],
                }
        self.constr_dados = pd.DataFrame(dict)

    @QtCore.Slot()
    def constr_ok(self):
        self.constr_pd()
        self.win_constr.close()

    @QtCore.Slot()
    def sys_data_ok(self):
        fator_bil, fator_term = reactor.corr_fact(
            self.win_system.inputAlt.value(), self.win_system.normaComboBox.currentText())
        self.win_system.inputFatorNBI.setValue(fator_bil)
        self.win_system.InputFatorCorrente.setValue(fator_term)
        dict = {'Frequência (Hz)': [self.win_system.inputFreq.value()],
                'Tensão nominal (V)': [self.win_system.inputVn.value()],
                'NBI (kV)': [self.win_system.inputNBI.value()],
                'Norma': [self.win_system.normaComboBox.currentText()],
                'Altitude (m)': [self.win_system.inputAlt.value()],
                'Velocidade do vento (km/h)': [self.win_system.inputVvento.value()],
                'Corrente de curto-circuito (kA)': [self.win_system.inputICC.value()],
                'Duração (s)': [self.win_system.inputTempo.value()],
                'Icc Pico (kA)': [self.win_system.inputPico.value()],
                'Temperatura ambiente (°C)': [self.win_system.inputTamb.value()],
                'Temperatura de referência (°C)': [self.win_system.inputTref.value()],
                'Indutância nominal (mH)': [self.win_system.inputLn.value()],
                'Indutância alvo (mH)': [self.win_system.inputLtarget.value()],
                'Corrente nominal (A)': [self.win_system.inputInominal.value()],
                'Corrente total (A)': [self.win_system.inputITotal.value()],
                'Fator de correção da corrente pela altitude': [fator_term],
                'Fator de correção do NBI pela altitude': [fator_bil], }
        self.sys_data = pd.DataFrame(dict)
        self.dados_proj()
        self.win_system.close()


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.win_reactor = None  # MainReactor()
        self.wincp = None
        self.bt_CpEn.clicked.connect(self.cp_en)
        self.bt_CpBr.clicked.connect(self.cp_br)
        self.bt_CHF.clicked.connect(self.cf_br)
        self.bt_FI.clicked.connect(self.fusein)
        self.bt_RS.clicked.connect(self.rs)
        self.bt_RSP.clicked.connect(self.rsp)
        self.bt_reactor.clicked.connect(self.show_reator)
        self.winrsp = Resistor()
        self.winfusewin = Fusein()
        self. wincp = CP()
        self.wincf = CF()

    def show_reator(self, checked):
        if self.win_reactor is None:
            self.win_reactor = MainReactor()
        self.win_reactor.show()
        # else:
        #self.win_reactor = None

    @QtCore.Slot()
    def rs(self):        
        self.winrsp.cb_Conex.setCurrentIndex(0)
        self.winrsp.exec()

    @QtCore.Slot()
    def rsp(self):
        self.winrsp.cb_Conex.setCurrentIndex(1)
        self.winrsp.exec()

    @QtCore.Slot()
    def cp_en(self):        
        self.wincp.cb_Eixos.setCurrentIndex(0)
        self.wincp.exec()        

    @QtCore.Slot()
    def cp_br(self):
        wincp = CP()
        wincp.cb_Eixos.setCurrentIndex(1)
        wincp.exec()

    @QtCore.Slot()
    def cf_br(self):
        wincf = CF()
        wincf.exec()

    @QtCore.Slot()
    def fusein(self):        
        self.winfusewin.exec()


# main
if __name__ == "__main__":
    # Create the application
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    #app = QApplication(sys.argv)
    # Create and show the application's main window
    win = MainWindow()
    win.show()
    # Run the application's main loop
    sys.exit(app.exec_())
